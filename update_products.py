"""
Скрипт для парсинга товаров из Excel файла
Читает ссылки, парсит данные, обновляет Excel
"""

import json
import re
import sys
import subprocess
from pathlib import Path
from datetime import datetime

# Фикс кодировки для Windows консоли
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ═══════════════════════════════════════════════════════════
# 📦 АВТОУСТАНОВКА ЗАВИСИМОСТЕЙ
# ═══════════════════════════════════════════════════════════

def install_dependencies():
    """Автоматически устанавливает необходимые пакеты."""
    required_packages = {
        'openpyxl': 'openpyxl==3.1.2',
        'requests': 'requests==2.31.0',
    }

    missing_packages = []

    for package_name, package_spec in required_packages.items():
        try:
            __import__(package_name)
        except ImportError:
            missing_packages.append(package_spec)

    if missing_packages:
        print("📦 Устанавливаю недостающие зависимости...")
        print(f"   Пакеты: {', '.join(missing_packages)}")

        try:
            subprocess.check_call([
                sys.executable,
                '-m',
                'pip',
                'install',
                *missing_packages
            ])
            print("✅ Зависимости установлены успешно!\n")
        except subprocess.CalledProcessError as e:
            print(f"❌ Ошибка установки зависимостей: {e}")
            print("   Попробуйте установить вручную:")
            print(f"   pip install {' '.join(missing_packages)}")
            sys.exit(1)

# Устанавливаем зависимости
install_dependencies()

import openpyxl
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ═══════════════════════════════════════════════════════════
# 📄 СОЗДАНИЕ ШАБЛОНА EXCEL
# ═══════════════════════════════════════════════════════════

def create_template(file_path=None):
    """Создаёт шаблонный Excel файл."""

    # Если путь не указан, используем папку скрипта
    if file_path is None:
        script_dir = Path(__file__).parent
        file_path = script_dir / "products_links.xlsx"
    else:
        file_path = Path(file_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    # Заголовки
    headers = ["URL товара", "Название", "Цена (€)", "Описание", "Группа", "Подгруппа", "Эмодзи", "URL фото", "Локальное фото", "Размеры", "Последнее обновление", "Статус"]
    ws.append(headers)

    # Стилизация заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Устанавливаем ширину колонок
    ws.column_dimensions['A'].width = 50  # URL товара
    ws.column_dimensions['B'].width = 30  # Название
    ws.column_dimensions['C'].width = 12  # Цена
    ws.column_dimensions['D'].width = 40  # Описание
    ws.column_dimensions['E'].width = 15  # Группа
    ws.column_dimensions['F'].width = 15  # Подгруппа
    ws.column_dimensions['G'].width = 10  # Эмодзи
    ws.column_dimensions['H'].width = 45  # URL фото
    ws.column_dimensions['I'].width = 25  # Локальное фото
    ws.column_dimensions['J'].width = 25  # Размеры
    ws.column_dimensions['K'].width = 20  # Обновление
    ws.column_dimensions['L'].width = 15  # Статус

    # Примеры (можно удалить)
    examples = [
        ["https://www.tradeinn.com/volleyball/ru/asics-gel-tactic-2-asics/139269743/p", "", "", "", "Волейбол", "Обувь", "🏐", "", "", "40, 41, 42, 43, 44", "", "Не спаршено"],
        ["", "", "", "", "Теннис", "Ракетки", "🎾", "", "", "", "", "Не заполнено"],
    ]

    for row in examples:
        ws.append(row)

    wb.save(file_path)
    print(f"✅ Создан шаблон Excel: {file_path}")


# ═══════════════════════════════════════════════════════════
# 🕷️ ПАРСИНГ ТОВАРОВ
# ═══════════════════════════════════════════════════════════

def download_image(image_url, save_dir, product_id):
    """Скачивает изображение и сохраняет локально."""
    try:
        response = requests.get(image_url, timeout=10, stream=True)
        response.raise_for_status()

        # Определяем расширение файла
        content_type = response.headers.get('content-type', '')
        ext = '.jpg'
        if 'png' in content_type:
            ext = '.png'
        elif 'webp' in content_type:
            ext = '.webp'
        elif 'jpeg' in content_type or 'jpg' in content_type:
            ext = '.jpg'

        # Генерируем имя файла
        filename = f"product_{product_id}_{hash(image_url) % 10000}{ext}"
        filepath = save_dir / filename

        # Сохраняем файл
        with open(filepath, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        return str(filepath.relative_to(save_dir.parent))
    except Exception as e:
        print(f"      ⚠️ Ошибка скачивания фото: {e}")
        return None


def parse_tradeinn_product(url, script_dir, product_id):
    """Парсит товар с tradeinn.com через HTML."""
    try:
        # Убираем лишние параметры из URL (всё после ?)
        if '?' in url:
            url = url.split('?')[0]
            print(f"      Очищаю URL от параметров...")

        # Автоматически заменяем /en/ на /ru/ для русской версии
        if '/en/' in url:
            url = url.replace('/en/', '/ru/')
            print(f"      Переключаю на русскую версию...")

        # Извлекаем product_id из URL (число перед /p)
        # Например: /141608258/p -> 141608258
        url_product_id = None
        product_id_match = re.search(r'/(\d+)/p/?$', url)
        if product_id_match:
            url_product_id = product_id_match.group(1)
            print(f"      🔑 Product ID: {url_product_id}")

        response = requests.get(url, timeout=10)
        response.raise_for_status()
        html = response.text

        # Ищем название в <h1>
        name_match = re.search(r'<h1[^>]*>([^<]+)</h1>', html, re.IGNORECASE)
        name = name_match.group(1).strip() if name_match else "Без названия"

        # Ищем цену в data-price (сохраняем в евро как есть)
        price_match = re.search(r'data-price="([^"]+)"', html, re.IGNORECASE)
        if price_match:
            try:
                price = float(price_match.group(1))
            except:
                price = 0
        else:
            price = 0

        # Ищем описание (опционально)
        desc_match = re.search(r'<meta name="description" content="([^"]+)"', html, re.IGNORECASE)
        description = desc_match.group(1)[:100] if desc_match else ""

        # Ищем ВСЕ фотки товара
        # Попробуем найти JSON с данными о товаре
        image_urls = []

        # Ищем JSON объект с данными товара (обычно называется dataLayer или similar)
        json_match = re.search(r'var\s+product\s*=\s*(\{[^}]+images[^}]+\})', html, re.DOTALL)
        if not json_match:
            # Пробуем другой формат
            json_match = re.search(r'"images"\s*:\s*(\[[^\]]+\])', html, re.DOTALL)

        if json_match:
            try:
                import json
                images_data = json.loads(json_match.group(1))
                if isinstance(images_data, list):
                    for img in images_data:
                        if isinstance(img, str) and 'tradeinn.com/f/' in img:
                            image_urls.append(img)
                elif isinstance(images_data, dict) and 'images' in images_data:
                    for img in images_data['images']:
                        if isinstance(img, str) and 'tradeinn.com/f/' in img:
                            image_urls.append(img)
            except:
                pass

        # Если JSON не нашли, ищем через селектор галереи
        if not image_urls:
            # Ищем фотки в блоке галереи (они обычно в specific container)
            gallery_match = re.search(r'<div[^>]*class="[^"]*product-gallery[^"]*"[^>]*>(.*?)</div>', html, re.DOTALL | re.IGNORECASE)
            if gallery_match:
                gallery_html = gallery_match.group(1)
                # Ищем все ссылки на фотки в галерее
                gallery_images = re.findall(r'(?:data-zoom-image|data-src|src)="([^"]+/f/\d+/\d+/[^"]+)"', gallery_html, re.IGNORECASE)
                image_urls.extend(gallery_images)

        # Если и это не сработало, используем более широкий поиск
        if not image_urls:
            all_images = re.findall(r'https://[^"\']+/f/\d+/\d+/[^"\']+\.(?:jpg|jpeg|png|webp)', html, re.IGNORECASE)
            # Фильтруем: берем только фотки ЭТОГО товара (с правильным product_id)
            for img_url in all_images:
                # Проверяем, что фотка не добавлена и не является превью/логотипом
                if img_url in image_urls:
                    continue
                if any(x in img_url.lower() for x in ['_thumb', '_small', '_icon', 'logo']):
                    continue

                # ВАЖНО: Проверяем, что URL фотки содержит product_id товара
                if url_product_id and f'/{url_product_id}/' not in img_url:
                    print(f"      ⏭️ Пропускаю (другой товар): ...{img_url[-60:]}")
                    continue

                image_urls.append(img_url)
                print(f"      ✓ Фото: ...{img_url[-60:]}")

        # Если не нашли фотки через data-атрибуты, ищем в Open Graph
        if not image_urls:
            og_image = re.search(r'<meta property="og:image" content="([^"]+)"', html)
            if og_image and og_image.group(1).startswith('http'):
                image_urls.append(og_image.group(1))

        print(f"      📷 Найдено фоток: {len(image_urls)}")

        # Скачиваем все фотки
        images_dir = script_dir / "images"
        images_dir.mkdir(exist_ok=True)

        local_images = []
        for img_url in image_urls:
            local_path = download_image(img_url, images_dir, product_id)
            if local_path:
                local_images.append(local_path)
                print(f"      ✅ Скачано: {local_path}")

        return {
            "name": name,
            "description": description,
            "price": price,
            "image_urls": ", ".join(image_urls) if image_urls else "",
            "local_images": ", ".join(local_images) if local_images else ""
        }, None

    except Exception as e:
        return None, f"Ошибка: {str(e)}"


def parse_generic_product(url, script_dir, product_id):
    """Универсальный парсер для других сайтов."""
    try:
        # Убираем лишние параметры из URL (всё после ?)
        if '?' in url:
            url = url.split('?')[0]

        response = requests.get(url, timeout=10)
        response.raise_for_status()
        html = response.text

        image_urls = []

        # Ищем JSON-LD
        json_ld_match = re.search(r'<script type="application/ld\+json">(.*?)</script>', html, re.DOTALL)

        if json_ld_match:
            try:
                data = json.loads(json_ld_match.group(1))

                # Проверяем тип
                if data.get("@type") == "Product":
                    name = data.get("name", "Без названия")
                    description = data.get("description", "")[:100]

                    offers = data.get("offers", {})
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}

                    price_str = offers.get("price", "0")
                    try:
                        price = float(price_str)
                    except:
                        price = 0

                    # Извлекаем фотки из JSON-LD
                    images = data.get("image", [])
                    if isinstance(images, str):
                        images = [images]
                    elif isinstance(images, dict):
                        images = [images.get("url", "")]

                    for img in images:
                        if isinstance(img, str) and img.startswith('http'):
                            image_urls.append(img)
                        elif isinstance(img, dict) and img.get("url"):
                            image_urls.append(img["url"])

                    # Скачиваем фотки
                    images_dir = script_dir / "images"
                    images_dir.mkdir(exist_ok=True)

                    local_images = []
                    for img_url in image_urls:
                        local_path = download_image(img_url, images_dir, product_id)
                        if local_path:
                            local_images.append(local_path)

                    return {
                        "name": name,
                        "description": description,
                        "price": price,
                        "image_urls": ", ".join(image_urls) if image_urls else "",
                        "local_images": ", ".join(local_images) if local_images else ""
                    }, None
            except:
                pass

        # Если JSON-LD не сработал, пробуем мета-теги
        name_match = re.search(r'<meta property="og:title" content="([^"]+)"', html)
        desc_match = re.search(r'<meta property="og:description" content="([^"]+)"', html)
        price_match = re.search(r'<meta property="product:price:amount" content="([^"]+)"', html)

        name = name_match.group(1) if name_match else "Без названия"
        description = desc_match.group(1)[:100] if desc_match else ""

        price = 0
        if price_match:
            try:
                price = float(price_match.group(1))
            except:
                price = 0

        # Ищем фотки через Open Graph
        og_images = re.findall(r'<meta property="og:image" content="([^"]+)"', html)
        for img in og_images:
            if img.startswith('http'):
                image_urls.append(img)

        # Скачиваем фотки
        images_dir = script_dir / "images"
        images_dir.mkdir(exist_ok=True)

        local_images = []
        for img_url in image_urls:
            local_path = download_image(img_url, images_dir, product_id)
            if local_path:
                local_images.append(local_path)

        return {
            "name": name,
            "description": description,
            "price": price,
            "image_urls": ", ".join(image_urls) if image_urls else "",
            "local_images": ", ".join(local_images) if local_images else ""
        }, None

    except Exception as e:
        return None, f"Ошибка: {str(e)}"


def parse_product(url, script_dir, product_id):
    """Определяет сайт и парсит товар."""

    if not url or not url.startswith("http"):
        return None, "Некорректный URL"

    if "tradeinn.com" in url:
        return parse_tradeinn_product(url, script_dir, product_id)
    else:
        return parse_generic_product(url, script_dir, product_id)


# ═══════════════════════════════════════════════════════════
# 📊 ОБНОВЛЕНИЕ EXCEL
# ═══════════════════════════════════════════════════════════

def update_excel(file_path=None):
    """Читает Excel, парсит товары, обновляет данные."""

    # Если путь не указан, используем папку скрипта
    if file_path is None:
        script_dir = Path(__file__).parent
        file_path = script_dir / "products_links.xlsx"
    else:
        file_path = Path(file_path)
        script_dir = file_path.parent

    if not file_path.exists():
        print(f"❌ Файл {file_path} не найден!")
        print("   Создаю шаблон...")
        create_template(file_path)
        return

    wb = load_workbook(file_path)
    ws = wb.active

    print("\n" + "=" * 60)
    print("ПАРСИНГ ТОВАРОВ")
    print("=" * 60 + "\n")

    updated_count = 0
    error_count = 0

    # Проходим по строкам (начиная со 2-й, т.к. 1-я это заголовки)
    for row_num in range(2, ws.max_row + 1):
        url = ws.cell(row_num, 1).value  # Колонка A - URL

        if not url or not url.startswith("http"):
            ws.cell(row_num, 12).value = "Пропущено (нет URL)"  # L: Статус
            continue

        print(f"[{row_num - 1}] Парсинг: {url[:60]}...")

        product_id = row_num - 1
        product_data, error = parse_product(url, script_dir, product_id)

        if error:
            print(f"    ❌ {error}")
            ws.cell(row_num, 12).value = error  # L: Статус
            error_count += 1
        else:
            print(f"    ✅ {product_data['name']}")
            print(f"       💰 Цена: {product_data['price']} €")

            # Обновляем ячейки
            ws.cell(row_num, 2).value = product_data['name']          # B: Название
            ws.cell(row_num, 3).value = product_data['price']         # C: Цена
            ws.cell(row_num, 4).value = product_data['description']   # D: Описание
            # E: Группа (заполняется вручную)
            # F: Подгруппа (заполняется вручную)
            # G: Эмодзи (заполняется вручную)
            ws.cell(row_num, 8).value = product_data['image_urls']    # H: URL фото
            ws.cell(row_num, 9).value = product_data['local_images']  # I: Локальное фото
            # J: Размеры (заполняется вручную)
            ws.cell(row_num, 11).value = datetime.now().strftime("%Y-%m-%d %H:%M")  # K: Последнее обновление
            ws.cell(row_num, 12).value = "✅ Обновлено"                              # L: Статус

            updated_count += 1

        # Задержка между запросами (2-3 секунды) чтобы не словить бан
        import time
        time.sleep(2)

    # Сохраняем
    wb.save(file_path)

    print("\n" + "=" * 60)
    print(f"✅ Обновлено товаров: {updated_count}")
    print(f"❌ Ошибок: {error_count}")
    print(f"📄 Файл сохранён: {file_path}")
    print("=" * 60 + "\n")


# ═══════════════════════════════════════════════════════════
# 📦 ЭКСПОРТ В PRODUCTS ДЛЯ БОТА
# ═══════════════════════════════════════════════════════════

def export_to_products_list(file_path=None):
    """Экспортирует товары из Excel в список PRODUCTS для mini_app.py."""

    # Если путь не указан, используем папку скрипта
    if file_path is None:
        script_dir = Path(__file__).parent
        file_path = script_dir / "products_links.xlsx"
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"❌ Файл {file_path} не найден!")
        return []

    wb = load_workbook(file_path)
    ws = wb.active

    products = []

    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row_num, 2).value          # B: Название
        price = ws.cell(row_num, 3).value         # C: Цена
        description = ws.cell(row_num, 4).value   # D: Описание
        category = ws.cell(row_num, 5).value      # E: Группа
        subcategory = ws.cell(row_num, 6).value   # F: Подгруппа
        emoji = ws.cell(row_num, 7).value         # G: Эмодзи
        image_urls = ws.cell(row_num, 8).value    # H: URL фото
        local_images = ws.cell(row_num, 9).value  # I: Локальное фото
        sizes = ws.cell(row_num, 10).value        # J: Размеры

        # Пропускаем строки без данных
        if not name or not price:
            continue

        # Преобразуем строку с фотками в список
        images_list = []
        if local_images:
            images_list = [img.strip() for img in local_images.split(',')]
        elif emoji:
            images_list = [emoji]

        # Преобразуем строку с размерами в список
        sizes_list = []
        if sizes:
            sizes_list = [size.strip() for size in sizes.split(',')]

        products.append({
            "id": row_num - 1,
            "name": name,
            "description": description or "",
            "price": float(price) if price else 0,  # Оставляем в евро
            "image": images_list[0] if images_list else "📦",  # Первая фотка
            "images": images_list,  # Все фотки
            "category": category or "",
            "subcategory": subcategory or "",
            "sizes": sizes_list,  # Размеры
        })

    return products


# ═══════════════════════════════════════════════════════════
# 🚀 MAIN
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    # Файл всегда создается в папке со скриптом
    script_dir = Path(__file__).parent
    file_path = script_dir / "products_links.xlsx"

    # Если файла нет - создаём шаблон
    if not file_path.exists():
        create_template()
        print("\n📝 Инструкция:")
        print(f"   1. Открой файл {file_path}")
        print("   2. Вставь ссылки на товары в колонку A")
        print("   3. Заполни эмодзи в колонке E (опционально)")
        print("   4. Запусти снова: python update_products.py\n")
    else:
        # Обновляем товары
        update_excel()

        # Показываем список для mini_app.py
        products = export_to_products_list()

        if products:
            print("\n💡 Товары готовы для использования в боте!")
            print(f"   Найдено товаров: {len(products)}\n")
