"""
ğŸ•·ï¸ GUI ĞŸÑ€Ğ¸Ğ»Ğ¾Ğ¶ĞµĞ½Ğ¸Ğµ Ğ´Ğ»Ñ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ° Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²
ĞĞºĞ¾Ğ½Ğ½Ğ¾Ğµ Ğ¿Ñ€Ğ¸Ğ»Ğ¾Ğ¶ĞµĞ½Ğ¸Ğµ Ñ ĞºĞ½Ğ¾Ğ¿ĞºĞ°Ğ¼Ğ¸ Ğ¸ ÑÑ‚Ğ°Ñ‚ÑƒÑ-Ğ±Ğ°Ñ€Ğ¾Ğ¼
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import sys
import subprocess
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import io
import json
import zipfile

# Ğ¤Ğ¸ĞºÑ ĞºĞ¾Ğ´Ğ¸Ñ€Ğ¾Ğ²ĞºĞ¸ Ğ´Ğ»Ñ Windows ĞºĞ¾Ğ½ÑĞ¾Ğ»Ğ¸
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ“¦ ĞĞ’Ğ¢ĞĞ£Ğ¡Ğ¢ĞĞĞĞ’ĞšĞ Ğ—ĞĞ’Ğ˜Ğ¡Ğ˜ĞœĞĞ¡Ğ¢Ğ•Ğ™
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def install_dependencies():
    """ĞĞ²Ñ‚Ğ¾Ğ¼Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¸ ÑƒÑÑ‚Ğ°Ğ½Ğ°Ğ²Ğ»Ğ¸Ğ²Ğ°ĞµÑ‚ Ğ½ĞµĞ¾Ğ±Ñ…Ğ¾Ğ´Ğ¸Ğ¼Ñ‹Ğµ Ğ¿Ğ°ĞºĞµÑ‚Ñ‹."""
    required_packages = {
        'openpyxl': 'openpyxl==3.1.2',
        'requests': 'requests==2.31.0',
        'yfinance': 'yfinance',
    }

    missing_packages = []

    for package_name, package_spec in required_packages.items():
        try:
            __import__(package_name)
        except ImportError:
            missing_packages.append(package_spec)

    if missing_packages:
        print("ğŸ“¦ Ğ£ÑÑ‚Ğ°Ğ½Ğ°Ğ²Ğ»Ğ¸Ğ²Ğ°Ñ Ğ½ĞµĞ´Ğ¾ÑÑ‚Ğ°ÑÑ‰Ğ¸Ğµ Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸...")
        print(f"   ĞŸĞ°ĞºĞµÑ‚Ñ‹: {', '.join(missing_packages)}")

        try:
            subprocess.check_call([
                sys.executable,
                '-m',
                'pip',
                'install',
                *missing_packages
            ])
            print("âœ… Ğ—Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ ÑƒÑÑ‚Ğ°Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ñ‹ ÑƒÑĞ¿ĞµÑˆĞ½Ğ¾!\n")
        except subprocess.CalledProcessError as e:
            print(f"âŒ ĞÑˆĞ¸Ğ±ĞºĞ° ÑƒÑÑ‚Ğ°Ğ½Ğ¾Ğ²ĞºĞ¸ Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚ĞµĞ¹: {e}")
            sys.exit(1)

# Ğ£ÑÑ‚Ğ°Ğ½Ğ°Ğ²Ğ»Ğ¸Ğ²Ğ°ĞµĞ¼ Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸
install_dependencies()

import openpyxl
import requests
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ“„ Ğ¡ĞĞ—Ğ”ĞĞĞ˜Ğ• ĞšĞ ĞĞ¡Ğ˜Ğ’ĞĞ“Ğ Ğ¨ĞĞ‘Ğ›ĞĞĞ EXCEL
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def create_beautiful_template(file_path=None, brands=None):
    """Ğ¡Ğ¾Ğ·Ğ´Ğ°Ñ‘Ñ‚ ĞºÑ€Ğ°ÑĞ¸Ğ²Ğ¾ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ½Ñ‹Ğ¹ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½ Excel."""

    if file_path is None:
        script_dir = Path(__file__).parent
        file_path = script_dir / "products_links.xlsx"
    else:
        file_path = Path(file_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "ğŸ› Ğ¢Ğ¾Ğ²Ğ°Ñ€Ñ‹"

    # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸ (Ğ±ĞµĞ· ÑĞ¼Ğ¾Ğ´Ğ·Ğ¸, Ğ‘Ğ•Ğ— Ğ¾Ğ¿Ğ¸ÑĞ°Ğ½Ğ¸Ñ)
    headers = ["URL Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°", "ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ", "Ğ¦ĞµĞ½Ğ° (â‚¬)", "Ğ“Ñ€ÑƒĞ¿Ğ¿Ğ°", "ĞŸĞ¾Ğ´Ğ³Ñ€ÑƒĞ¿Ğ¿Ğ°", "ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°", "URL Ñ„Ğ¾Ñ‚Ğ¾", "Ğ›Ğ¾ĞºĞ°Ğ»ÑŒĞ½Ğ¾Ğµ Ñ„Ğ¾Ñ‚Ğ¾", "Ğ Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹", "ĞŸĞ¾ÑĞ»ĞµĞ´Ğ½ĞµĞµ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ", "Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ"]
    ws.append(headers)

    # ĞšÑ€Ğ°ÑĞ¸Ğ²Ñ‹Ğµ ÑÑ‚Ğ¸Ğ»Ğ¸ Ğ´Ğ»Ñ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¾Ğ²
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ğ“Ñ€Ğ°Ğ½Ğ¸Ñ†Ñ‹
    thin_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ ÑÑ‚Ğ¸Ğ»Ğ¸ Ğº Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ°Ğ¼
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Ğ’Ñ‹ÑĞ¾Ñ‚Ğ° ÑÑ‚Ñ€Ğ¾ĞºĞ¸ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ°
    ws.row_dimensions[1].height = 30

    # Ğ¨Ğ¸Ñ€Ğ¸Ğ½Ğ° ĞºĞ¾Ğ»Ğ¾Ğ½Ğ¾Ğº
    ws.column_dimensions['A'].width = 55  # URL Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°
    ws.column_dimensions['B'].width = 35  # ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ
    ws.column_dimensions['C'].width = 12  # Ğ¦ĞµĞ½Ğ°
    ws.column_dimensions['D'].width = 18  # Ğ“Ñ€ÑƒĞ¿Ğ¿Ğ°
    ws.column_dimensions['E'].width = 18  # ĞŸĞ¾Ğ´Ğ³Ñ€ÑƒĞ¿Ğ¿Ğ°
    ws.column_dimensions['F'].width = 20  # ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°
    ws.column_dimensions['G'].width = 45  # URL Ñ„Ğ¾Ñ‚Ğ¾
    ws.column_dimensions['H'].width = 25  # Ğ›Ğ¾ĞºĞ°Ğ»ÑŒĞ½Ğ¾Ğµ Ñ„Ğ¾Ñ‚Ğ¾
    ws.column_dimensions['I'].width = 25  # Ğ Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹
    ws.column_dimensions['J'].width = 22  # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ
    ws.column_dimensions['K'].width = 18  # Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ

    # ĞŸÑ€Ğ¸Ğ¼ĞµÑ€Ñ‹ ÑƒĞ±Ñ€Ğ°Ğ½Ñ‹ - Ğ¿ÑƒÑÑ‚Ğ¾Ğ¹ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½
    examples = []

    # Ğ¦Ğ²ĞµÑ‚Ğ° Ğ´Ğ»Ñ ÑÑ‚Ñ€Ğ¾Ğº (Ñ‡ĞµÑ€ĞµĞ´Ğ¾Ğ²Ğ°Ğ½Ğ¸Ğµ)
    row_colors = ["F2F2F2", "FFFFFF"]

    # Ğ¡Ñ‚Ğ¸Ğ»Ğ¸ Ğ´Ğ»Ñ ÑÑ‡ĞµĞµĞº Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ…
    data_font = Font(size=11, name="Calibri")
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    data_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    for idx, row in enumerate(examples):
        row_num = idx + 2
        ws.append(row)

        # Ğ¦Ğ²ĞµÑ‚ Ñ„Ğ¾Ğ½Ğ° ÑÑ‚Ñ€Ğ¾ĞºĞ¸ (Ğ¿Ğ¾ ÑƒĞ¼Ğ¾Ğ»Ñ‡Ğ°Ğ½Ğ¸Ñ)
        row_fill = PatternFill(start_color=row_colors[idx % 2], end_color=row_colors[idx % 2], fill_type="solid")

        # Ğ—Ğ°Ğ»Ğ¸Ğ²ĞºĞ¸ Ğ´Ğ»Ñ ĞºĞ¾Ğ½ĞºÑ€ĞµÑ‚Ğ½Ñ‹Ñ… ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ²
        name_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Ğ¡ĞµÑ€Ñ‹Ğ¹ Ğ´Ğ»Ñ Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ
        price_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Ğ—ĞµĞ»Ñ‘Ğ½Ñ‹Ğ¹ Ğ´Ğ»Ñ Ñ†ĞµĞ½Ñ‹

        # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ ÑÑ‚Ğ¸Ğ»Ğ¸ Ğº ĞºĞ°Ğ¶Ğ´Ğ¾Ğ¹ ÑÑ‡ĞµĞ¹ĞºĞµ
        for col_num, cell in enumerate(ws[row_num], start=1):
            # Ğ¡Ğ¿ĞµÑ†Ğ¸Ğ°Ğ»ÑŒĞ½Ñ‹Ğµ Ğ·Ğ°Ğ»Ğ¸Ğ²ĞºĞ¸ Ğ´Ğ»Ñ Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ Ğ¸ Ñ†ĞµĞ½Ñ‹
            if col_num == 2:  # B: ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ
                cell.fill = name_fill
            elif col_num == 3:  # C: Ğ¦ĞµĞ½Ğ°
                cell.fill = price_fill
            else:
                cell.fill = row_fill

            cell.font = data_font
            cell.border = data_border

            # Ğ’Ñ‹Ñ€Ğ°Ğ²Ğ½Ğ¸Ğ²Ğ°Ğ½Ğ¸Ğµ Ğ¿Ğ¾ Ñ†ĞµĞ½Ñ‚Ñ€Ñƒ Ğ´Ğ»Ñ Ğ¾Ğ¿Ñ€ĞµĞ´ĞµĞ»Ñ‘Ğ½Ğ½Ñ‹Ñ… ĞºĞ¾Ğ»Ğ¾Ğ½Ğ¾Ğº
            if col_num in [3, 9, 10, 11]:  # Ğ¦ĞµĞ½Ğ°, Ğ Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹, ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ, Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment

        # Ğ’Ñ‹ÑĞ¾Ñ‚Ğ° ÑÑ‚Ñ€Ğ¾ĞºĞ¸
        ws.row_dimensions[row_num].height = 25

    # Ğ—Ğ°ĞºÑ€ĞµĞ¿Ğ»ÑĞµĞ¼ Ğ¿ĞµÑ€Ğ²ÑƒÑ ÑÑ‚Ñ€Ğ¾ĞºÑƒ (Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸)
    ws.freeze_panes = "A2"

    # ĞĞ²Ñ‚Ğ¾Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€ (Ñ‚ĞµĞ¿ĞµÑ€ÑŒ Ğ´Ğ¾ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ¸ S - Ğ²ĞºĞ»ÑÑ‡Ğ°Ñ Ñ€Ğ°ÑÑ‡ĞµÑ‚Ğ½Ñ‹Ğµ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ñ‹)
    ws.auto_filter.ref = f"A1:S1"

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ğŸ“Š Ğ”ĞĞ‘ĞĞ’Ğ›Ğ¯Ğ•Ğœ Ğ ĞĞ¡Ğ§Ğ•Ğ¢ĞĞ«Ğ• Ğ¡Ğ¢ĞĞ›Ğ‘Ğ¦Ğ« (L-S)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    calc_headers = [
        "Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚½)",      # L
        "Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° (â‚½)",       # M
        "ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (%)",      # N
        "ĞĞ°Ñˆ ĞšÑÑ„ (%)",       # O
        "Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚. (â‚½)", # P
        "Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚. (â‚½)", # Q
        "ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ° (â‚½)",    # R
        "ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸ (â‚½)"     # S
    ]

    # ĞÑ€Ğ°Ğ½Ğ¶ĞµĞ²Ğ¾Ğµ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ¸Ğµ Ğ´Ğ»Ñ Ñ€Ğ°ÑÑ‡ĞµÑ‚Ğ½Ñ‹Ñ… Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¾Ğ²
    orange_header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    calc_header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    calc_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    calc_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for col_idx, header in enumerate(calc_headers, start=12):  # L=12
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = orange_header_fill
        cell.font = calc_header_font
        cell.alignment = calc_header_alignment
        cell.border = calc_border

        # Ğ¨Ğ¸Ñ€Ğ¸Ğ½Ğ° ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ²
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = 18

    # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº "Ğ‘Ñ€ĞµĞ½Ğ´" Ğ² ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğµ T (20)
    brand_cell = ws.cell(1, 20)
    brand_cell.value = "Ğ‘Ñ€ĞµĞ½Ğ´"
    brand_cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    brand_cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    brand_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    brand_cell.border = calc_border
    ws.column_dimensions['T'].width = 18

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ğŸ“‹ Ğ’Ğ«ĞŸĞĞ”ĞĞ®Ğ©Ğ˜Ğ™ Ğ¡ĞŸĞ˜Ğ¡ĞĞš ĞšĞĞ¢Ğ•Ğ“ĞĞ Ğ˜Ğ™ Ğ´Ğ»Ñ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ° F
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    # Ğ¡Ğ¿Ğ¸ÑĞ¾Ğº ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²
    categories = [
        "ĞÑ‡ĞºĞ¸",
        "Ğ Ğ°ĞºĞµÑ‚ĞºĞ°",
        "ĞšÑ€Ğ¾ÑÑĞ¾Ğ²ĞºĞ¸",
        "ĞšÑƒÑ€Ñ‚ĞºĞ°",
        "Ğ¨Ñ‚Ğ°Ğ½Ñ‹",
        "Ğ¨Ğ»ĞµĞ¼",
        "Ğ‘Ğ¾Ñ‚Ğ¸Ğ½ĞºĞ¸ Ğ±Ğ¾Ñ€Ğ´",
        "Ğ¢ĞµÑ€Ğ¼Ğ¾",
        "ĞÑ‡ĞºĞ¸ Ğ´Ğ»Ñ ÑĞ½ĞµĞ³Ğ°"
    ]

    # Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµĞ¼ Ğ²Ñ‹Ğ¿Ğ°Ğ´Ğ°ÑÑ‰Ğ¸Ğ¹ ÑĞ¿Ğ¸ÑĞ¾Ğº Ğ´Ğ»Ñ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ° F (ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°)
    categories_formula = f'"{",".join(categories)}"'
    dv_category = DataValidation(
        type="list",
        formula1=categories_formula,
        allow_blank=True,
        showDropDown=False,  # False = Ğ¿Ğ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ ÑÑ‚Ñ€ĞµĞ»ĞºÑƒ Ğ²Ñ‹Ğ¿Ğ°Ğ´Ğ°ÑÑ‰ĞµĞ³Ğ¾ ÑĞ¿Ğ¸ÑĞºĞ°
        showInputMessage=False,  # ĞĞµ Ğ¿Ğ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ‡Ğ°Ğ½Ğ¸Ğµ
        showErrorMessage=True
    )
    dv_category.error = "Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ğ¸Ğ· ÑĞ¿Ğ¸ÑĞºĞ° Ğ´Ğ¾Ğ¿ÑƒÑÑ‚Ğ¸Ğ¼Ñ‹Ñ… Ğ·Ğ½Ğ°Ñ‡ĞµĞ½Ğ¸Ğ¹!"
    dv_category.errorTitle = "âŒ ĞĞµĞ²ĞµÑ€Ğ½Ğ°Ñ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ"

    ws.add_data_validation(dv_category)
    # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğº ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ñƒ F ÑĞ¾ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ 2 Ğ´Ğ¾ 10000
    dv_category.add('F2:F10000')

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # âš™ï¸ Ğ›Ğ˜Ğ¡Ğ¢ ĞĞĞ¡Ğ¢Ğ ĞĞ•Ğš
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    settings_ws = wb.create_sheet("âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸")

    # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
    settings_ws['A1'] = "âš™ï¸ ĞĞĞ¡Ğ¢Ğ ĞĞ™ĞšĞ˜ Ğ ĞĞ¡Ğ§Ğ•Ğ¢ĞĞ’"
    settings_ws['A1'].font = Font(bold=True, size=16, name="Calibri")
    settings_ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    settings_ws['A1'].font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
    settings_ws.merge_cells('A1:C1')

    # ĞšÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹
    settings_ws['A3'] = "ĞšÑƒÑ€Ñ EUR/RUB:"
    settings_ws['B3'] = 100.0  # Ğ—Ğ½Ğ°Ñ‡ĞµĞ½Ğ¸Ğµ Ğ¿Ğ¾ ÑƒĞ¼Ğ¾Ğ»Ñ‡Ğ°Ğ½Ğ¸Ñ
    settings_ws['A3'].font = Font(bold=True, size=12)
    settings_ws['B3'].font = Font(size=12)
    settings_ws['B3'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    settings_ws['A4'] = "ĞĞ°Ğ´Ğ±Ğ°Ğ²ĞºĞ°:"
    settings_ws['B4'] = 0.5  # Ğ—Ğ½Ğ°Ñ‡ĞµĞ½Ğ¸Ğµ Ğ¿Ğ¾ ÑƒĞ¼Ğ¾Ğ»Ñ‡Ğ°Ğ½Ğ¸Ñ
    settings_ws['A4'].font = Font(bold=True, size=12)
    settings_ws['B4'].font = Font(size=12)
    settings_ws['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    settings_ws['A5'] = "Ğ˜Ñ‚Ğ¾Ğ³Ğ¾Ğ²Ñ‹Ğ¹ ĞºÑƒÑ€Ñ:"
    settings_ws['B5'] = "=B3+B4"
    settings_ws['A5'].font = Font(bold=True, size=12)
    settings_ws['B5'].font = Font(bold=True, size=14)
    settings_ws['B5'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
    settings_ws['A7'] = "ğŸ“¦ Ğ¡Ğ¢ĞĞ˜ĞœĞĞ¡Ğ¢Ğ¬ Ğ”ĞĞ¡Ğ¢ĞĞ’ĞšĞ˜ (â‚¬)"
    settings_ws['A7'].font = Font(bold=True, size=14)
    settings_ws['A7'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    settings_ws['A7'].font = Font(bold=True, color="FFFFFF", size=14)
    settings_ws.merge_cells('A7:B7')

    settings_ws['A8'] = "ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ"
    settings_ws['B8'] = "Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚¬)"
    settings_ws['A8'].font = Font(bold=True, size=11)
    settings_ws['B8'].font = Font(bold=True, size=11)
    settings_ws['A8'].fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    settings_ws['B8'].fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")

    # Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹ Ğ¸ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
    delivery_table = [
        ("ĞÑ‡ĞºĞ¸", 12),
        ("Ğ Ğ°ĞºĞµÑ‚ĞºĞ°", 17),
        ("ĞšÑ€Ğ¾ÑÑĞ¾Ğ²ĞºĞ¸", 28),
        ("ĞšÑƒÑ€Ñ‚ĞºĞ°", 17),
        ("Ğ¨Ñ‚Ğ°Ğ½Ñ‹", 17),
        ("Ğ¨Ğ»ĞµĞ¼", 28),
        ("Ğ‘Ğ¾Ñ‚Ğ¸Ğ½ĞºĞ¸ Ğ±Ğ¾Ñ€Ğ´", 25),
        ("Ğ¢ĞµÑ€Ğ¼Ğ¾", 17),
        ("ĞÑ‡ĞºĞ¸ Ğ´Ğ»Ñ ÑĞ½ĞµĞ³Ğ°", 17)
    ]

    for idx, (cat, delivery) in enumerate(delivery_table, start=9):
        settings_ws[f'A{idx}'] = cat
        settings_ws[f'B{idx}'] = delivery
        settings_ws[f'A{idx}'].border = calc_border
        settings_ws[f'B{idx}'].border = calc_border

    # Ğ¡ĞµĞºÑ†Ğ¸Ñ Ğ‘Ğ Ğ•ĞĞ”Ğ« (ÑÑ‚Ğ¾Ğ»Ğ±ĞµÑ† D)
    settings_ws['D1'] = "ğŸ·ï¸ Ğ‘Ğ Ğ•ĞĞ”Ğ«"
    settings_ws['D1'].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    settings_ws['D1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    settings_ws['D2'] = "Ğ‘Ñ€ĞµĞ½Ğ´"
    settings_ws['D2'].font = Font(bold=True, size=11)
    settings_ws['D2'].fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")

    brands_list = brands or [
        "Asics", "Adidas", "Bullpadel", "Drop Shot", "Head",
        "Joma", "Mizuno", "Nike", "Nox", "Oakley", "Puma", "Siux", "Wilson"
    ]
    for idx, brand in enumerate(brands_list, start=3):
        settings_ws[f'D{idx}'] = brand
        settings_ws[f'D{idx}'].border = calc_border

    # Ğ¨Ğ¸Ñ€Ğ¸Ğ½Ğ° ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ² Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
    settings_ws.column_dimensions['A'].width = 25
    settings_ws.column_dimensions['B'].width = 20
    settings_ws.column_dimensions['C'].width = 15
    settings_ws.column_dimensions['D'].width = 20

    wb.save(file_path)
    return file_path


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ•·ï¸ ĞŸĞĞ Ğ¡Ğ˜ĞĞ“ (Ğ¸Ğ· update_products.py)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def get_images_dir(script_dir):
    """
    ĞĞ¿Ñ€ĞµĞ´ĞµĞ»ÑĞµÑ‚ Ğ¿ÑƒÑ‚ÑŒ Ğº Ğ¿Ğ°Ğ¿ĞºĞµ images Ğ² Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ¾Ñ‚ Ğ¾ĞºÑ€ÑƒĞ¶ĞµĞ½Ğ¸Ñ.

    ĞŸÑ€Ğ¸Ğ¾Ñ€Ğ¸Ñ‚ĞµÑ‚:
    1. /data/images/ (ĞµÑĞ»Ğ¸ ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚ Ğ¸ ĞĞ• Ğ¿ÑƒÑÑ‚Ğ°Ñ) - Ğ¿Ğ¾ÑÑ‚Ğ¾ÑĞ½Ğ½Ğ¾Ğµ Ñ…Ñ€Ğ°Ğ½Ğ¸Ğ»Ğ¸Ñ‰Ğµ Amvera
    2. script_dir/images/ - Ğ¸Ğ· Ñ€ĞµĞ¿Ğ¾Ğ·Ğ¸Ñ‚Ğ¾Ñ€Ğ¸Ñ (fallback)
    """
    # ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑĞµĞ¼ /data/images/ Ğ½Ğ° Amvera
    data_path = Path('/data')
    if data_path.exists() and data_path.is_dir():
        data_images_dir = data_path / 'images'
        data_images_dir.mkdir(exist_ok=True)

        # Ğ•ÑĞ»Ğ¸ Ñ‚Ğ°Ğ¼ ÑƒĞ¶Ğµ ĞµÑÑ‚ÑŒ Ñ„Ğ°Ğ¹Ğ»Ñ‹ - Ğ¸ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞµĞ¼ ĞµÑ‘
        if any(data_images_dir.iterdir()):
            return data_images_dir

    # Fallback: Ğ»Ğ¾ĞºĞ°Ğ»ÑŒĞ½Ğ°Ñ Ğ¿Ğ°Ğ¿ĞºĞ° Ğ¸Ğ»Ğ¸ images Ğ¸Ğ· Ñ€ĞµĞ¿Ğ¾Ğ·Ğ¸Ñ‚Ğ¾Ñ€Ğ¸Ñ
    images_dir = script_dir / 'images'
    images_dir.mkdir(exist_ok=True)
    return images_dir


def clean_product_name(name):
    """Ğ£Ğ±Ğ¸Ñ€Ğ°ĞµÑ‚ Ñ€ÑƒÑÑĞºĞ¸Ğµ (ĞºĞ¸Ñ€Ğ¸Ğ»Ğ»Ğ¸Ñ‡ĞµÑĞºĞ¸Ğµ) ÑĞ»Ğ¾Ğ²Ğ° Ğ¸Ğ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°.

    ĞŸÑ€Ğ¸Ğ¼ĞµÑ€:
        "Bullpadel Ñ€Ğ°ĞºĞµÑ‚ĞºĞ° Ğ´Ğ»Ñ Ğ¿Ğ°Ğ´ĞµĞ»Ñ Vertex 04 2025" â†’ "Bullpadel Vertex 04 2025"
    """
    if not name:
        return name

    import re

    # Ğ Ğ°Ğ·Ğ±Ğ¸Ğ²Ğ°ĞµĞ¼ Ğ½Ğ° ÑĞ»Ğ¾Ğ²Ğ°
    words = name.split()

    # ĞÑÑ‚Ğ°Ğ²Ğ»ÑĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ ÑĞ»Ğ¾Ğ²Ğ°, ĞºĞ¾Ñ‚Ğ¾Ñ€Ñ‹Ğµ Ğ½Ğµ ÑĞ¾Ğ´ĞµÑ€Ğ¶Ğ°Ñ‚ ĞºĞ¸Ñ€Ğ¸Ğ»Ğ»Ğ¸Ñ†Ñƒ
    clean_words = []
    for word in words:
        # ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑĞµĞ¼, ĞµÑÑ‚ÑŒ Ğ»Ğ¸ Ğ² ÑĞ»Ğ¾Ğ²Ğµ Ñ…Ğ¾Ñ‚ÑŒ Ğ¾Ğ´Ğ½Ğ° ĞºĞ¸Ñ€Ğ¸Ğ»Ğ»Ğ¸Ñ‡ĞµÑĞºĞ°Ñ Ğ±ÑƒĞºĞ²Ğ°
        if not re.search(r'[Ğ°-ÑĞ-Ğ¯Ñ‘Ğ]', word):
            clean_words.append(word)

    # Ğ¡Ğ¾Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ğ¾Ğ±Ñ€Ğ°Ñ‚Ğ½Ğ¾ Ğ² ÑÑ‚Ñ€Ğ¾ĞºÑƒ
    result = ' '.join(clean_words)

    # Ğ£Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ğ¼Ğ½Ğ¾Ğ¶ĞµÑÑ‚Ğ²ĞµĞ½Ğ½Ñ‹Ğµ Ğ¿Ñ€Ğ¾Ğ±ĞµĞ»Ñ‹
    result = re.sub(r'\s+', ' ', result).strip()

    return result


def download_image(image_url, save_dir, product_id):
    """Ğ¡ĞºĞ°Ñ‡Ğ¸Ğ²Ğ°ĞµÑ‚ Ğ¸Ğ·Ğ¾Ğ±Ñ€Ğ°Ğ¶ĞµĞ½Ğ¸Ğµ Ğ¸ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ÑĞµÑ‚ Ğ»Ğ¾ĞºĞ°Ğ»ÑŒĞ½Ğ¾ (Ğ²ÑĞµĞ³Ğ´Ğ° Ğ¿ĞµÑ€ĞµĞ·Ğ°Ğ¿Ğ¸ÑÑ‹Ğ²Ğ°ĞµÑ‚)."""
    try:
        # Ğ£Ğ´Ğ°Ğ»ÑĞµĞ¼ ÑÑ‚Ğ°Ñ€Ñ‹Ğµ Ñ„Ğ°Ğ¹Ğ»Ñ‹ Ğ´Ğ»Ñ ÑÑ‚Ğ¾Ğ³Ğ¾ product_id (ĞµÑĞ»Ğ¸ ĞµÑÑ‚ÑŒ)
        for old_file in save_dir.glob(f"product_{product_id}.*"):
            old_file.unlink()

        # Ğ¡ĞºĞ°Ñ‡Ğ¸Ğ²Ğ°ĞµĞ¼ Ğ¸Ğ·Ğ¾Ğ±Ñ€Ğ°Ğ¶ĞµĞ½Ğ¸Ğµ
        response = requests.get(image_url, timeout=10, stream=True)
        response.raise_for_status()

        # ĞĞ¿Ñ€ĞµĞ´ĞµĞ»ÑĞµĞ¼ Ñ€Ğ°ÑÑˆĞ¸Ñ€ĞµĞ½Ğ¸Ğµ Ñ„Ğ°Ğ¹Ğ»Ğ°
        content_type = response.headers.get('content-type', '')
        ext = '.jpg'
        if 'png' in content_type:
            ext = '.png'
        elif 'webp' in content_type:
            ext = '.webp'
        elif 'jpeg' in content_type or 'jpg' in content_type:
            ext = '.jpg'

        # Ğ“ĞµĞ½ĞµÑ€Ğ¸Ñ€ÑƒĞµĞ¼ Ğ¸Ğ¼Ñ Ñ„Ğ°Ğ¹Ğ»Ğ°: product_1.webp, product_2.jpg Ğ¸ Ñ‚.Ğ´.
        filename = f"product_{product_id}{ext}"
        filepath = save_dir / filename

        # Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµĞ¼ Ñ„Ğ°Ğ¹Ğ» (Ğ¿ĞµÑ€ĞµĞ·Ğ°Ğ¿Ğ¸ÑÑ‹Ğ²Ğ°ĞµĞ¼ ĞµÑĞ»Ğ¸ ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚)
        with open(filepath, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        return str(filepath.relative_to(save_dir.parent))
    except Exception as e:
        print(f"      âš ï¸ ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞºĞ°Ñ‡Ğ¸Ğ²Ğ°Ğ½Ğ¸Ñ Ñ„Ğ¾Ñ‚Ğ¾: {e}")
        return None


def parse_tradeinn_product(url, script_dir, product_id):
    """ĞŸĞ°Ñ€ÑĞ¸Ñ‚ Ñ‚Ğ¾Ğ²Ğ°Ñ€ Ñ tradeinn.com Ñ‡ĞµÑ€ĞµĞ· HTML."""
    try:
        if '?' in url:
            url = url.split('?')[0]

        if '/en/' in url:
            url = url.replace('/en/', '/ru/')

        # Ğ˜Ğ·Ğ²Ğ»ĞµĞºĞ°ĞµĞ¼ product_id Ğ¸Ğ· URL (Ñ‡Ğ¸ÑĞ»Ğ¾ Ğ¿ĞµÑ€ĞµĞ´ /p)
        url_product_id = None
        product_id_match = re.search(r'/(\d+)/p/?$', url)
        if product_id_match:
            url_product_id = product_id_match.group(1)

        response = requests.get(url, timeout=10)
        response.raise_for_status()
        html = response.text

        name_match = re.search(r'<h1[^>]*>([^<]+)</h1>', html, re.IGNORECASE)
        name = name_match.group(1).strip() if name_match else "Ğ‘ĞµĞ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ"
        # Ğ£Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ñ€ÑƒÑÑĞºĞ¸Ğµ ÑĞ»Ğ¾Ğ²Ğ° Ğ¸Ğ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ
        name = clean_product_name(name)

        price_match = re.search(r'data-price="([^"]+)"', html, re.IGNORECASE)
        if price_match:
            try:
                price = float(price_match.group(1))
            except:
                price = 0
        else:
            price = 0

        # ĞŸĞ°Ñ€ÑĞ¸Ğ¼ Ğ²ÑĞµ Ñ„Ğ¾Ñ‚ĞºĞ¸
        image_urls = []

        # ĞœĞ•Ğ¢ĞĞ” 1: Ğ˜Ñ‰ĞµĞ¼ Ğ³Ğ°Ğ»ĞµÑ€ĞµÑ Ñ data-fancybox="gallery" - ÑĞ°Ğ¼Ñ‹Ğ¹ Ğ½Ğ°Ğ´Ñ‘Ğ¶Ğ½Ñ‹Ğ¹ ÑĞ¿Ğ¾ÑĞ¾Ğ±!
        # Ğ­Ñ‚Ğ¸ ÑÑÑ‹Ğ»ĞºĞ¸ Ğ²ĞµĞ´ÑƒÑ‚ Ğ½Ğ° Ğ¿Ğ¾Ğ»Ğ½Ğ¾Ñ€Ğ°Ğ·Ğ¼ĞµÑ€Ğ½Ñ‹Ğµ Ğ¸Ğ·Ğ¾Ğ±Ñ€Ğ°Ğ¶ĞµĞ½Ğ¸Ñ Ğ´Ğ»Ñ Ğ¿Ñ€Ğ¾ÑĞ¼Ğ¾Ñ‚Ñ€Ğ°
        gallery_links = re.findall(r'data-fancybox="gallery"[^>]*href="([^"]+)"', html, re.IGNORECASE)
        if gallery_links:
            for link in gallery_links:
                # ĞÑ‚Ğ½Ğ¾ÑĞ¸Ñ‚ĞµĞ»ÑŒĞ½Ñ‹Ğµ ÑÑÑ‹Ğ»ĞºĞ¸ Ğ¿Ñ€ĞµĞ¾Ğ±Ñ€Ğ°Ğ·ÑƒĞµĞ¼ Ğ² Ğ°Ğ±ÑĞ¾Ğ»ÑÑ‚Ğ½Ñ‹Ğµ
                if link.startswith('/'):
                    link = 'https://www.tradeinn.com' + link
                if link not in image_urls:
                    image_urls.append(link)

        # ĞœĞ•Ğ¢ĞĞ” 2: Ğ•ÑĞ»Ğ¸ Ğ³Ğ°Ğ»ĞµÑ€ĞµÑ Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½Ğ°, Ğ¸Ñ‰ĞµĞ¼ Ñ‡ĞµÑ€ĞµĞ· Ğ¿Ğ°Ñ‚Ñ‚ĞµÑ€Ğ½ /f/ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ/Ñ‚Ğ¾Ğ²Ğ°Ñ€_X/
        # Ğ£Ñ‡Ğ¸Ñ‚Ñ‹Ğ²Ğ°ĞµĞ¼ ÑÑƒÑ„Ñ„Ğ¸ĞºÑÑ‹ _2, _3, _4 Ğ¸ Ñ‚.Ğ´. Ğ´Ğ»Ñ Ñ€Ğ°Ğ·Ğ½Ñ‹Ñ… Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸Ğ¹
        if not image_urls and url_product_id:
            # Ğ˜Ğ·Ğ²Ğ»ĞµĞºĞ°ĞµĞ¼ ID ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸ Ğ¸Ğ· URL
            category_match = re.search(r'/(\d+)/\d+/p', url)
            if category_match:
                category_id = category_match.group(1)

                # Ğ˜Ñ‰ĞµĞ¼ Ğ²ÑĞµ Ğ¸Ğ·Ğ¾Ğ±Ñ€Ğ°Ğ¶ĞµĞ½Ğ¸Ñ Ñ ÑƒÑ‡Ñ‘Ñ‚Ğ¾Ğ¼ ÑÑƒÑ„Ñ„Ğ¸ĞºÑĞ¾Ğ² _2, _3, _4...
                # ĞŸĞ°Ñ‚Ñ‚ĞµÑ€Ğ½: /f/14160/141608258_2/filename.webp
                pattern = rf'/f/{category_id}/{url_product_id}(?:_\d+)?/[^"\']+\.(?:jpg|jpeg|png|webp)'
                found_images = re.findall(pattern, html, re.IGNORECASE)

                for img in found_images:
                    full_url = 'https://www.tradeinn.com' + img if img.startswith('/') else img
                    if full_url not in image_urls:
                        image_urls.append(full_url)

        # ĞœĞ•Ğ¢ĞĞ” 3: Ğ˜Ñ‰ĞµĞ¼ JSON Ğ¾Ğ±ÑŠĞµĞºÑ‚ Ñ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğ¼Ğ¸ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°
        if not image_urls:
            json_match = re.search(r'var\s+product\s*=\s*(\{[^}]+images[^}]+\})', html, re.DOTALL)
            if not json_match:
                json_match = re.search(r'"images"\s*:\s*(\[[^\]]+\])', html, re.DOTALL)

            if json_match:
                try:
                    import json
                    images_data = json.loads(json_match.group(1))
                    if isinstance(images_data, list):
                        for img in images_data:
                            if isinstance(img, str) and 'tradeinn.com/f/' in img:
                                image_urls.append(img)
                    elif isinstance(images_data, dict) and 'images' in images_data:
                        for img in images_data['images']:
                            if isinstance(img, str) and 'tradeinn.com/f/' in img:
                                image_urls.append(img)
                except:
                    pass

        # ĞœĞ•Ğ¢ĞĞ” 4: Ğ¨Ğ¸Ñ€Ğ¾ĞºĞ¸Ğ¹ Ğ¿Ğ¾Ğ¸ÑĞº Ğ²ÑĞµÑ… Ğ¸Ğ·Ğ¾Ğ±Ñ€Ğ°Ğ¶ĞµĞ½Ğ¸Ğ¹
        if not image_urls:
            all_images = re.findall(r'https://[^"\']+/f/\d+/\d+(?:_\d+)?/[^"\']+\.(?:jpg|jpeg|png|webp)', html, re.IGNORECASE)
            for img_url in all_images:
                if img_url not in image_urls and not any(x in img_url.lower() for x in ['_thumb', '_small', '_icon', 'logo']):
                    image_urls.append(img_url)

        # ĞœĞ•Ğ¢ĞĞ” 5: Open Graph ĞºĞ°Ğº Ğ·Ğ°Ğ¿Ğ°ÑĞ½Ğ¾Ğ¹ Ğ²Ğ°Ñ€Ğ¸Ğ°Ğ½Ñ‚
        if not image_urls:
            og_image = re.search(r'<meta property="og:image" content="([^"]+)"', html)
            if og_image and og_image.group(1).startswith('http'):
                image_urls.append(og_image.group(1))

        # ĞŸĞ°Ñ€ÑĞ¸Ğ¼ Ñ€Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹ (Ğ´Ğ»Ñ Ğ¾Ğ±ÑƒĞ²Ğ¸, Ğ¾Ğ´ĞµĞ¶Ğ´Ñ‹)
        sizes = []

        # ĞœĞ•Ğ¢ĞĞ” 1: JSON-LD ÑÑ‚Ñ€ÑƒĞºÑ‚ÑƒÑ€Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ½Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ (ÑĞ°Ğ¼Ñ‹Ğ¹ Ğ½Ğ°Ğ´ĞµĞ¶Ğ½Ñ‹Ğ¹ Ğ´Ğ»Ñ TradeInn!)
        json_ld_pattern = r'<script type="application/ld\+json">(.*?)</script>'
        json_ld_matches = re.findall(json_ld_pattern, html, re.DOTALL)

        for json_str in json_ld_matches:
            try:
                import json
                data = json.loads(json_str)

                # Ğ˜Ñ‰ĞµĞ¼ Ğ²Ğ°Ñ€Ğ¸Ğ°Ğ½Ñ‚Ñ‹ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ° (hasVariant)
                if isinstance(data, dict) and data.get('@type') == 'Product':
                    variants = data.get('hasVariant', [])
                    if variants:
                        for variant in variants:
                            # Ğ˜Ğ·Ğ²Ğ»ĞµĞºĞ°ĞµĞ¼ Ñ€Ğ°Ğ·Ğ¼ĞµÑ€ Ğ¸Ğ· Ğ¸Ğ¼ĞµĞ½Ğ¸ Ğ²Ğ°Ñ€Ğ¸Ğ°Ğ½Ñ‚Ğ°
                            variant_name = variant.get('name', '')
                            # ĞŸÑ€Ğ¸Ğ¼ĞµÑ€: "EU 42 1/2" Ğ¸Ğ»Ğ¸ "EU 44"
                            size_match = re.search(r'EU\s+(\d+(?:\s*1/2)?)', variant_name)
                            if size_match:
                                size = size_match.group(1).strip()
                                if size not in sizes:
                                    sizes.append(size)
            except:
                pass

        # ĞœĞ•Ğ¢ĞĞ” 2: Ğ˜Ñ‰ĞµĞ¼ Ñ€Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹ Ğ² select ÑĞ»ĞµĞ¼ĞµĞ½Ñ‚Ğµ
        if not sizes:
            size_select_match = re.findall(r'<option[^>]*value="size:([^"]+)"[^>]*>([^<]+)</option>', html, re.IGNORECASE)
            if size_select_match:
                for size_value, size_label in size_select_match:
                    size_clean = size_label.strip()
                    if size_clean and size_clean.lower() not in ['Ğ²Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ Ñ€Ğ°Ğ·Ğ¼ĞµÑ€', 'choose size', 'select']:
                        sizes.append(size_clean)

        # ĞœĞ•Ğ¢ĞĞ” 3: Ğ˜Ñ‰ĞµĞ¼ Ñ€Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹ Ğ² data-Ğ°Ñ‚Ñ€Ğ¸Ğ±ÑƒÑ‚Ğ°Ñ…
        if not sizes:
            size_data_match = re.findall(r'data-size="([^"]+)"', html, re.IGNORECASE)
            if size_data_match:
                for size in size_data_match:
                    size_clean = size.strip()
                    if size_clean and len(size_clean) <= 10:
                        sizes.append(size_clean)

        # ĞœĞ•Ğ¢ĞĞ” 4: Ğ˜Ñ‰ĞµĞ¼ Ñ‚ĞµĞºÑÑ‚Ğ¾Ğ²Ñ‹Ğµ Ğ¿Ğ°Ñ‚Ñ‚ĞµÑ€Ğ½Ñ‹ "EU 42", "Size 42" (Ğ ĞĞ‘ĞĞ¢ĞĞ•Ğ¢ Ğ”Ğ›Ğ¯ TRADEINN!)
        if not sizes:
            text_patterns = [
                r'(?:EU|Size|Ğ Ğ°Ğ·Ğ¼ĞµÑ€)\s+(\d{2}(?:\s*1/2)?)',  # EU 42, EU 42 1/2
                r'size["\']?\s*:\s*["\'](\d{2}(?:\s*1/2)?)["\']',  # JSON: "size":"42"
            ]

            for pattern in text_patterns:
                matches = re.findall(pattern, html, re.IGNORECASE)
                if matches:
                    for match in set(matches):
                        if match not in sizes:
                            sizes.append(match)

        # Ğ£Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ğ´ÑƒĞ±Ğ»Ğ¸ĞºĞ°Ñ‚Ñ‹ Ğ¸ ÑĞ¾Ñ€Ñ‚Ğ¸Ñ€ÑƒĞµĞ¼
        if sizes:
            sizes = list(dict.fromkeys(sizes))  # Ğ£Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ğ´ÑƒĞ±Ğ»Ğ¸ĞºĞ°Ñ‚Ñ‹, ÑĞ¾Ñ…Ñ€Ğ°Ğ½ÑÑ Ğ¿Ğ¾Ñ€ÑĞ´Ğ¾Ğº
            # ĞŸÑ‹Ñ‚Ğ°ĞµĞ¼ÑÑ Ğ¾Ñ‚ÑĞ¾Ñ€Ñ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ñ‚ÑŒ Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ½Ğ¾ (ÑƒÑ‡Ğ¸Ñ‚Ñ‹Ğ²Ğ°ĞµĞ¼ Ğ´Ñ€Ğ¾Ğ±Ğ¸ Ñ‚Ğ¸Ğ¿Ğ° "42 1/2")
            def parse_size(s):
                # ĞŸÑ€ĞµĞ¾Ğ±Ñ€Ğ°Ğ·ÑƒĞµĞ¼ "42 1/2" Ğ² 42.5
                if '1/2' in s:
                    base = float(s.replace('1/2', '').strip())
                    return base + 0.5
                try:
                    return float(s.replace(',', '.'))
                except:
                    return 999

            try:
                sizes = sorted(set(sizes), key=parse_size)
            except:
                pass

        # Ğ¡ĞºĞ°Ñ‡Ğ¸Ğ²Ğ°ĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ ĞŸĞ•Ğ Ğ’Ğ£Ğ® Ñ„Ğ¾Ñ‚ĞºÑƒ (ÑĞºĞ¾Ğ½Ğ¾Ğ¼Ğ¸Ğ¼ Ğ¼ĞµÑÑ‚Ğ¾ Ğ¸ Ñ‚Ñ€Ğ°Ñ„Ğ¸Ğº)
        images_dir = get_images_dir(script_dir)

        local_images = []
        if image_urls:
            # Ğ‘ĞµÑ€ĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ¿ĞµÑ€Ğ²ÑƒÑ Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸Ñ
            local_path = download_image(image_urls[0], images_dir, product_id)
            if local_path:
                local_images.append(local_path)

        return {
            "name": name,
            "price": price,
            "image_urls": ", ".join(image_urls) if image_urls else "",
            "local_images": ", ".join(local_images) if local_images else "",
            "sizes": ", ".join(sizes) if sizes else ""
        }, None

    except Exception as e:
        return None, f"ĞÑˆĞ¸Ğ±ĞºĞ°: {str(e)}"


def parse_generic_product(url, script_dir, product_id):
    """Ğ£Ğ½Ğ¸Ğ²ĞµÑ€ÑĞ°Ğ»ÑŒĞ½Ñ‹Ğ¹ Ğ¿Ğ°Ñ€ÑĞµÑ€ Ğ´Ğ»Ñ Ğ´Ñ€ÑƒĞ³Ğ¸Ñ… ÑĞ°Ğ¹Ñ‚Ğ¾Ğ²."""
    try:
        if '?' in url:
            url = url.split('?')[0]

        # Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµĞ¼ session Ğ´Ğ»Ñ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ğ¸Ñ cookies Ğ¼ĞµĞ¶Ğ´Ñƒ Ğ·Ğ°Ğ¿Ñ€Ğ¾ÑĞ°Ğ¼Ğ¸
        session = requests.Session()

        # ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ° headers Ğ´Ğ»Ñ Ğ¸Ğ¼Ğ¸Ñ‚Ğ°Ñ†Ğ¸Ğ¸ Ñ€ĞµĞ°Ğ»ÑŒĞ½Ğ¾Ğ³Ğ¾ Ğ±Ñ€Ğ°ÑƒĞ·ĞµÑ€Ğ°
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
        }
        session.headers.update(headers)

        # Ğ”Ğ»Ñ TradeInn ÑƒÑÑ‚Ğ°Ğ½Ğ°Ğ²Ğ»Ğ¸Ğ²Ğ°ĞµĞ¼ ÑÑ‚Ñ€Ğ°Ğ½Ñƒ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ Armenia (id_pais=9)
        if 'tradeinn.com' in url:
            # Ğ”ĞµĞ»Ğ°ĞµĞ¼ Ğ·Ğ°Ğ¿Ñ€Ğ¾Ñ Ğ´Ğ»Ñ ÑƒÑÑ‚Ğ°Ğ½Ğ¾Ğ²ĞºĞ¸ ÑÑ‚Ñ€Ğ°Ğ½Ñ‹ Armenia Ñ‡ĞµÑ€ĞµĞ· ÑĞ¿ĞµÑ†Ğ¸Ğ°Ğ»ÑŒĞ½Ñ‹Ğ¹ endpoint
            try:
                # id_tienda=27 - ÑÑ‚Ğ¾ volleyball Ğ¼Ğ°Ğ³Ğ°Ğ·Ğ¸Ğ½, id_pais=9 - ÑÑ‚Ğ¾ Armenia
                country_setup_url = "https://www.tradeinn.com/get_dades.php?id_tienda=27&idioma=rus&id_pais=9&country_code_url="
                session.get(country_setup_url, timeout=5)
            except:
                pass  # Ğ•ÑĞ»Ğ¸ Ğ½Ğµ Ğ¿Ğ¾Ğ»ÑƒÑ‡Ğ¸Ğ»Ğ¾ÑÑŒ - Ğ¿Ñ€Ğ¾Ğ´Ğ¾Ğ»Ğ¶Ğ°ĞµĞ¼ Ğ±ĞµĞ· ÑƒÑÑ‚Ğ°Ğ½Ğ¾Ğ²ĞºĞ¸ ÑÑ‚Ñ€Ğ°Ğ½Ñ‹

        # ĞÑĞ½Ğ¾Ğ²Ğ½Ğ¾Ğ¹ Ğ·Ğ°Ğ¿Ñ€Ğ¾Ñ ÑÑ‚Ñ€Ğ°Ğ½Ğ¸Ñ†Ñ‹
        response = session.get(url, timeout=10)
        response.raise_for_status()
        html = response.text

        image_urls = []
        json_ld_match = re.search(r'<script type="application/ld\+json">(.*?)</script>', html, re.DOTALL)

        if json_ld_match:
            try:
                import json
                data = json.loads(json_ld_match.group(1))

                if data.get("@type") == "Product":
                    name = data.get("name", "Ğ‘ĞµĞ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ")
                    description = data.get("description", "")[:100]

                    offers = data.get("offers", {})
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}

                    price_str = offers.get("price", "0")
                    try:
                        price = float(price_str)
                    except:
                        price = 0

                    # ĞŸĞ°Ñ€ÑĞ¸Ğ¼ Ñ„Ğ¾Ñ‚ĞºĞ¸ Ğ¸Ğ· JSON-LD
                    images = data.get("image", [])
                    if isinstance(images, str):
                        images = [images]
                    elif isinstance(images, dict):
                        images = [images.get("url", "")]

                    for img in images:
                        if isinstance(img, str) and img.startswith('http'):
                            image_urls.append(img)
                        elif isinstance(img, dict) and img.get("url"):
                            image_urls.append(img["url"])

                    # Ğ¡ĞºĞ°Ñ‡Ğ¸Ğ²Ğ°ĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ ĞŸĞ•Ğ Ğ’Ğ£Ğ® Ñ„Ğ¾Ñ‚ĞºÑƒ (ÑĞºĞ¾Ğ½Ğ¾Ğ¼Ğ¸Ğ¼ Ğ¼ĞµÑÑ‚Ğ¾ Ğ¸ Ñ‚Ñ€Ğ°Ñ„Ğ¸Ğº)
                    images_dir = script_dir / "images"
                    images_dir.mkdir(exist_ok=True)

                    local_images = []
                    if image_urls:
                        # Ğ‘ĞµÑ€ĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ¿ĞµÑ€Ğ²ÑƒÑ Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸Ñ
                        local_path = download_image(image_urls[0], images_dir, product_id)
                        if local_path:
                            local_images.append(local_path)

                    return {
                        "name": name,
                        "description": description,
                        "price": price,
                        "image_urls": ", ".join(image_urls) if image_urls else "",
                        "local_images": ", ".join(local_images) if local_images else ""
                    }, None
            except:
                pass

        name_match = re.search(r'<meta property="og:title" content="([^"]+)"', html)
        price_match = re.search(r'<meta property="product:price:amount" content="([^"]+)"', html)

        name = name_match.group(1) if name_match else "Ğ‘ĞµĞ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ"
        # Ğ£Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ñ€ÑƒÑÑĞºĞ¸Ğµ ÑĞ»Ğ¾Ğ²Ğ° Ğ¸Ğ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ
        name = clean_product_name(name)

        price = 0
        if price_match:
            try:
                price = float(price_match.group(1))
            except:
                price = 0

        # ĞŸĞ°Ñ€ÑĞ¸Ğ¼ Ñ„Ğ¾Ñ‚ĞºĞ¸ Ñ‡ĞµÑ€ĞµĞ· Open Graph
        og_images = re.findall(r'<meta property="og:image" content="([^"]+)"', html)
        for img in og_images:
            if img.startswith('http'):
                image_urls.append(img)

        # Ğ¡ĞºĞ°Ñ‡Ğ¸Ğ²Ğ°ĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ ĞŸĞ•Ğ Ğ’Ğ£Ğ® Ñ„Ğ¾Ñ‚ĞºÑƒ (ÑĞºĞ¾Ğ½Ğ¾Ğ¼Ğ¸Ğ¼ Ğ¼ĞµÑÑ‚Ğ¾ Ğ¸ Ñ‚Ñ€Ğ°Ñ„Ğ¸Ğº)
        images_dir = get_images_dir(script_dir)

        local_images = []
        if image_urls:
            # Ğ‘ĞµÑ€ĞµĞ¼ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ¿ĞµÑ€Ğ²ÑƒÑ Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸Ñ
            local_path = download_image(image_urls[0], images_dir, product_id)
            if local_path:
                local_images.append(local_path)

        return {
            "name": name,
            "price": price,
            "image_urls": ", ".join(image_urls) if image_urls else "",
            "local_images": ", ".join(local_images) if local_images else ""
        }, None

    except Exception as e:
        return None, f"ĞÑˆĞ¸Ğ±ĞºĞ°: {str(e)}"


def parse_product(url, script_dir, product_id):
    """ĞĞ¿Ñ€ĞµĞ´ĞµĞ»ÑĞµÑ‚ ÑĞ°Ğ¹Ñ‚ Ğ¸ Ğ¿Ğ°Ñ€ÑĞ¸Ñ‚ Ñ‚Ğ¾Ğ²Ğ°Ñ€."""
    if not url or not url.startswith("http"):
        return None, "ĞĞµĞºĞ¾Ñ€Ñ€ĞµĞºÑ‚Ğ½Ñ‹Ğ¹ URL"

    if "tradeinn.com" in url:
        return parse_tradeinn_product(url, script_dir, product_id)
    else:
        return parse_generic_product(url, script_dir, product_id)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ–¥ï¸ GUI ĞŸĞ Ğ˜Ğ›ĞĞ–Ğ•ĞĞ˜Ğ•
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

class ParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ğŸ•·ï¸ ĞŸĞ°Ñ€ÑĞµÑ€ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ² Ğ´Ğ»Ñ Telegram Ğ¼Ğ°Ğ³Ğ°Ğ·Ğ¸Ğ½Ğ°")
        self.root.geometry("900x650")
        self.root.resizable(True, True)

        # ĞŸĞ°Ğ¿ĞºĞ° ÑĞºÑ€Ğ¸Ğ¿Ñ‚Ğ°
        self.script_dir = Path(__file__).parent
        self.file_path = self.script_dir / "products_links.xlsx"
        self.settings_file = self.script_dir / "parser_settings.json"

        # Ğ¡Ñ‚Ğ¸Ğ»Ğ¸
        style = ttk.Style()
        style.theme_use('clam')

        # Ğ“Ğ»Ğ°Ğ²Ğ½Ñ‹Ğ¹ Ñ„Ñ€ĞµĞ¹Ğ¼
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº Ğ¿Ñ€Ğ¸Ğ»Ğ¾Ğ¶ĞµĞ½Ğ¸Ñ
        title_label = tk.Label(
            main_frame,
            text="ğŸ•·ï¸ ĞŸĞ°Ñ€ÑĞµÑ€ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ² Ğ´Ğ»Ñ Telegram Ğ¼Ğ°Ğ³Ğ°Ğ·Ğ¸Ğ½Ğ°",
            font=("Segoe UI", 16, "bold"),
            fg="#1F4E78"
        )
        title_label.pack(pady=(0, 20))

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ“‘ Ğ’ĞšĞ›ĞĞ”ĞšĞ˜ (Notebook)
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ“‘ Ğ’ĞšĞ›ĞĞ”ĞšĞ 1: ĞŸĞĞ Ğ¡Ğ˜ĞĞ“ Ğ¢ĞĞ’ĞĞ ĞĞ’
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        tab_parser = ttk.Frame(notebook, padding="10")
        notebook.add(tab_parser, text="ğŸ•·ï¸ ĞŸĞ°Ñ€ÑĞ¸Ğ½Ğ³ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²")

        # Ğ¤Ñ€ĞµĞ¹Ğ¼ Ğ´Ğ»Ñ ĞºĞ½Ğ¾Ğ¿Ğ¾Ğº
        button_frame = ttk.Frame(tab_parser)
        button_frame.pack(pady=10)

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ° ÑĞ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ñ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°
        self.create_btn = tk.Button(
            button_frame,
            text="ğŸ“„ Ğ¡Ğ¾Ğ·Ğ´Ğ°Ñ‚ÑŒ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ½Ñ‹Ğ¹ Ñ„Ğ°Ğ¹Ğ»",
            command=self.create_template_clicked,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            padx=20,
            pady=10,
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        )
        self.create_btn.pack(side=tk.LEFT, padx=10)

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ° Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°
        self.parse_btn = tk.Button(
            button_frame,
            text="ğŸš€ Ğ¡Ğ¿Ğ°Ñ€ÑĞ¸Ñ‚ÑŒ Ğ¸Ğ· Excel Ñ„Ğ°Ğ¹Ğ»Ğ°",
            command=self.parse_clicked,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            padx=20,
            pady=10,
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        )
        self.parse_btn.pack(side=tk.LEFT, padx=10)

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ° Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ñ
        self.archive_btn = tk.Button(
            button_frame,
            text="ğŸ“¦ Ğ—Ğ°Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ¸Ñ€Ğ¾Ğ²Ğ°Ñ‚ÑŒ Ğ´Ğ»Ñ Ğ±Ğ¾Ñ‚Ğ°",
            command=self.archive_clicked,
            bg="#FF9800",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            padx=20,
            pady=10,
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        )
        self.archive_btn.pack(side=tk.LEFT, padx=10)

        # Ğ’Ñ‹Ğ±Ğ¾Ñ€ ĞºĞ¾Ğ»Ğ¸Ñ‡ĞµÑÑ‚Ğ²Ğ° Ğ¿Ğ¾Ñ‚Ğ¾ĞºĞ¾Ğ²
        threads_frame = tk.Frame(tab_parser, bg="#f5f5f5")
        threads_frame.pack(pady=5)

        tk.Label(
            threads_frame,
            text="âš¡ ĞŸĞ¾Ñ‚Ğ¾ĞºĞ¸:",
            font=("Segoe UI", 11, "bold"),
            bg="#f5f5f5",
            fg="#333"
        ).pack(side=tk.LEFT, padx=(0, 8))

        self.threads_var = tk.IntVar(value=saved.get("threads", 5))
        self.threads_spinbox = tk.Spinbox(
            threads_frame,
            from_=1,
            to=10,
            textvariable=self.threads_var,
            width=3,
            font=("Segoe UI", 12, "bold"),
            justify=tk.CENTER,
            state="readonly"
        )
        self.threads_spinbox.pack(side=tk.LEFT)

        tk.Label(
            threads_frame,
            text="(1 = Ğ¼ĞµĞ´Ğ»ĞµĞ½Ğ½Ğ¾, 5 = Ğ¾Ğ¿Ñ‚Ğ¸Ğ¼Ğ°Ğ»ÑŒĞ½Ğ¾, 10 = Ğ¼Ğ°ĞºÑĞ¸Ğ¼ÑƒĞ¼)",
            font=("Segoe UI", 9),
            bg="#f5f5f5",
            fg="#888"
        ).pack(side=tk.LEFT, padx=(8, 0))

        # Ğ˜Ğ½Ñ„Ğ¾Ñ€Ğ¼Ğ°Ñ†Ğ¸Ğ¾Ğ½Ğ½Ğ°Ñ Ğ¿Ğ°Ğ½ĞµĞ»ÑŒ
        info_frame = ttk.LabelFrame(tab_parser, text="â„¹ï¸ Ğ˜Ğ½Ñ„Ğ¾Ñ€Ğ¼Ğ°Ñ†Ğ¸Ñ", padding="10")
        info_frame.pack(fill=tk.X, pady=10)

        self.info_label = tk.Label(
            info_frame,
            text=f"ğŸ“ Ğ¤Ğ°Ğ¹Ğ»: {self.file_path.name}\nğŸ“‚ ĞŸĞ°Ğ¿ĞºĞ°: {self.script_dir}",
            font=("Segoe UI", 10),
            justify=tk.LEFT,
            fg="#555"
        )
        self.info_label.pack(anchor=tk.W)

        # Ğ›Ğ¾Ğ³-Ğ¿Ğ°Ğ½ĞµĞ»ÑŒ
        log_frame = ttk.LabelFrame(tab_parser, text="ğŸ“‹ Ğ–ÑƒÑ€Ğ½Ğ°Ğ» Ñ€Ğ°Ğ±Ğ¾Ñ‚Ñ‹", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Ğ¢ĞµĞºÑÑ‚Ğ¾Ğ²Ğ¾Ğµ Ğ¿Ğ¾Ğ»Ğµ Ğ´Ğ»Ñ Ğ»Ğ¾Ğ³Ğ¾Ğ²
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg="#F5F5F5",
            fg="#333",
            height=15
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ Ğ±Ğ°Ñ€
        self.status_bar = tk.Label(
            root,
            text="Ğ“Ğ¾Ñ‚Ğ¾Ğ² Ğº Ñ€Ğ°Ğ±Ğ¾Ñ‚Ğµ",
            bd=1,
            relief=tk.SUNKEN,
            anchor=tk.W,
            bg="#1F4E78",
            fg="white",
            font=("Segoe UI", 10),
            padx=10,
            pady=5
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # ĞŸÑ€Ğ¸Ğ²ĞµÑ‚ÑÑ‚Ğ²ĞµĞ½Ğ½Ğ¾Ğµ ÑĞ¾Ğ¾Ğ±Ñ‰ĞµĞ½Ğ¸Ğµ
        self.log("=" * 80)
        self.log("ğŸ‰ Ğ”Ğ¾Ğ±Ñ€Ğ¾ Ğ¿Ğ¾Ğ¶Ğ°Ğ»Ğ¾Ğ²Ğ°Ñ‚ÑŒ Ğ² ĞŸĞ°Ñ€ÑĞµÑ€ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²!")
        self.log("=" * 80)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ“‘ Ğ’ĞšĞ›ĞĞ”ĞšĞ 2: ĞšĞ£Ğ Ğ¡ Ğ’ĞĞ›Ğ®Ğ¢Ğ«
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        tab_currency = ttk.Frame(notebook, padding="0")
        notebook.add(tab_currency, text="ğŸ’± ĞšÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹")

        # Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµĞ¼ Canvas Ñ Ğ¿Ñ€Ğ¾ĞºÑ€ÑƒÑ‚ĞºĞ¾Ğ¹ Ğ´Ğ»Ñ Ğ²ÑĞµĞ³Ğ¾ ÑĞ¾Ğ´ĞµÑ€Ğ¶Ğ¸Ğ¼Ğ¾Ğ³Ğ¾
        currency_canvas = tk.Canvas(tab_currency, highlightthickness=0)
        currency_scrollbar = ttk.Scrollbar(tab_currency, orient="vertical", command=currency_canvas.yview)
        currency_scrollable_frame = ttk.Frame(currency_canvas, padding="20")

        currency_scrollable_frame.bind(
            "<Configure>",
            lambda _: currency_canvas.configure(scrollregion=currency_canvas.bbox("all"))
        )

        currency_canvas.create_window((0, 0), window=currency_scrollable_frame, anchor="nw")
        currency_canvas.configure(yscrollcommand=currency_scrollbar.set)

        currency_canvas.pack(side="left", fill="both", expand=True)
        currency_scrollbar.pack(side="right", fill="y")

        # ĞŸÑ€Ğ¸Ğ²ÑĞ·Ñ‹Ğ²Ğ°ĞµĞ¼ Ğ¿Ñ€Ğ¾ĞºÑ€ÑƒÑ‚ĞºÑƒ Ğ¼Ñ‹ÑˆÑŒÑ
        def _on_mousewheel(event):
            currency_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        currency_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº Ğ²ĞºĞ»Ğ°Ğ´ĞºĞ¸
        currency_title = tk.Label(
            currency_scrollable_frame,
            text="ğŸ’± Ğ£Ğ¿Ñ€Ğ°Ğ²Ğ»ĞµĞ½Ğ¸Ğµ ĞºÑƒÑ€ÑĞ¾Ğ¼ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹ EUR â†’ RUB",
            font=("Segoe UI", 14, "bold"),
            fg="#1F4E78"
        )
        currency_title.pack(pady=(0, 15))

        # Ğ¤Ñ€ĞµĞ¹Ğ¼ Ğ´Ğ»Ñ Ñ‚ĞµĞºÑƒÑ‰ĞµĞ³Ğ¾ ĞºÑƒÑ€ÑĞ° Ğ¸ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº (ĞºĞ¾Ğ¼Ğ¿Ğ°ĞºÑ‚Ğ½Ğ¾)
        current_settings_frame = ttk.LabelFrame(currency_scrollable_frame, text="ğŸ“Š ĞšÑƒÑ€Ñ Ğ¸ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸", padding="15")
        current_settings_frame.pack(fill=tk.X, pady=10)

        self.currency_rate_label = tk.Label(
            current_settings_frame,
            text="ĞšÑƒÑ€Ñ EUR/RUB: Ğ·Ğ°Ğ³Ñ€ÑƒĞ·ĞºĞ°...",
            font=("Segoe UI", 11, "bold"),
            fg="#2196F3"
        )
        self.currency_rate_label.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=5)

        self.last_update_label = tk.Label(
            current_settings_frame,
            text="ĞŸĞ¾ÑĞ»ĞµĞ´Ğ½ĞµĞµ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ: -",
            font=("Segoe UI", 9),
            fg="#666"
        )
        self.last_update_label.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))

        # ĞĞ°Ğ´Ğ±Ğ°Ğ²ĞºĞ° Ğº ĞºÑƒÑ€ÑÑƒ
        markup_label = tk.Label(
            current_settings_frame,
            text="ĞĞ°Ğ´Ğ±Ğ°Ğ²ĞºĞ° Ğº ĞºÑƒÑ€ÑÑƒ (+):",
            font=("Segoe UI", 10)
        )
        markup_label.grid(row=2, column=0, sticky=tk.W, pady=5, padx=(0, 10))

        self.markup_entry = tk.Entry(
            current_settings_frame,
            font=("Segoe UI", 10),
            width=10
        )
        self.markup_entry.insert(0, str(saved.get("markup", 0.5)))
        self.markup_entry.grid(row=2, column=1, sticky=tk.W, pady=5)

        markup_hint = tk.Label(
            current_settings_frame,
            text="â‚½",
            font=("Segoe UI", 10),
            fg="#666"
        )
        markup_hint.grid(row=2, column=2, sticky=tk.W, pady=5, padx=(5, 0))

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ“‹ ĞšĞĞ¢Ğ•Ğ“ĞĞ Ğ˜Ğ˜ Ğ¢ĞĞ’ĞĞ ĞĞ’ Ğ˜ Ğ¡Ğ¢ĞĞ˜ĞœĞĞ¡Ğ¢Ğ¬ Ğ”ĞĞ¡Ğ¢ĞĞ’ĞšĞ˜
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        # Ğ¤Ñ€ĞµĞ¹Ğ¼ Ğ´Ğ»Ñ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹
        self.categories_main_frame = ttk.LabelFrame(currency_scrollable_frame, text="ğŸ“‹ ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸ Ğ¸ ÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ (â‚¬)", padding="15")
        self.categories_main_frame.pack(fill=tk.X, pady=10)

        # Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµĞ¼ ÑĞ¾Ñ…Ñ€Ğ°Ğ½Ñ‘Ğ½Ğ½Ñ‹Ğµ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸ (Ğ¸Ğ»Ğ¸ Ğ´ĞµÑ„Ğ¾Ğ»Ñ‚Ğ½Ñ‹Ğµ)
        saved = self._load_settings()

        self.categories_data = saved.get("categories", [
            {"name": "ĞÑ‡ĞºĞ¸", "delivery": 12},
            {"name": "Ğ Ğ°ĞºĞµÑ‚ĞºĞ°", "delivery": 17},
            {"name": "ĞšÑ€Ğ¾ÑÑĞ¾Ğ²ĞºĞ¸", "delivery": 28},
            {"name": "ĞšÑƒÑ€Ñ‚ĞºĞ°", "delivery": 17},
            {"name": "Ğ¨Ñ‚Ğ°Ğ½Ñ‹", "delivery": 17},
            {"name": "Ğ¨Ğ»ĞµĞ¼", "delivery": 28},
            {"name": "Ğ‘Ğ¾Ñ‚Ğ¸Ğ½ĞºĞ¸ Ğ±Ğ¾Ñ€Ğ´", "delivery": 25},
            {"name": "Ğ¢ĞµÑ€Ğ¼Ğ¾", "delivery": 17},
            {"name": "ĞÑ‡ĞºĞ¸ Ğ´Ğ»Ñ ÑĞ½ĞµĞ³Ğ°", "delivery": 17}
        ])

        # Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµĞ¼ Ñ„Ñ€ĞµĞ¹Ğ¼ Ğ´Ğ»Ñ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñ‹ (Ğ±ÑƒĞ´ĞµÑ‚ Ğ¿ĞµÑ€ĞµÑ€Ğ¸ÑĞ¾Ğ²Ñ‹Ğ²Ğ°Ñ‚ÑŒÑÑ)
        self.categories_table_frame = tk.Frame(self.categories_main_frame)
        self.categories_table_frame.pack(fill=tk.X)

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ¸ ÑƒĞ¿Ñ€Ğ°Ğ²Ğ»ĞµĞ½Ğ¸Ñ
        buttons_frame = tk.Frame(self.categories_main_frame)
        buttons_frame.pack(fill=tk.X, pady=10)

        add_category_btn = tk.Button(
            buttons_frame,
            text="â• Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ¸Ñ‚ÑŒ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ",
            command=self.add_category_dialog,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=5,
            cursor="hand2"
        )
        add_category_btn.pack(side=tk.LEFT, padx=5)

        save_categories_btn = tk.Button(
            buttons_frame,
            text="ğŸ’¾ Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½Ğ¸Ñ‚ÑŒ Ğ¸Ğ·Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ",
            command=self.save_category_changes,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=5,
            cursor="hand2"
        )
        save_categories_btn.pack(side=tk.LEFT, padx=5)

        # ĞÑ‚Ñ€Ğ¸ÑĞ¾Ğ²Ñ‹Ğ²Ğ°ĞµĞ¼ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñƒ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹
        self.refresh_categories_table()

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ·ï¸ Ğ‘Ğ Ğ•ĞĞ”Ğ«
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        self.brands_data = saved.get("brands", [
            "Asics", "Adidas", "Bullpadel", "Drop Shot", "Head",
            "Joma", "Mizuno", "Nike", "Nox", "Oakley", "Puma", "Siux", "Wilson"
        ])

        self.brands_main_frame = ttk.LabelFrame(currency_scrollable_frame, text="ğŸ·ï¸ Ğ‘Ñ€ĞµĞ½Ğ´Ñ‹ (Ğ´Ğ»Ñ Ğ°Ğ²Ñ‚Ğ¾-Ğ¾Ğ¿Ñ€ĞµĞ´ĞµĞ»ĞµĞ½Ğ¸Ñ Ğ¸Ğ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğ¹)", padding="15")
        self.brands_main_frame.pack(fill=tk.X, pady=10)

        self.brands_table_frame = tk.Frame(self.brands_main_frame)
        self.brands_table_frame.pack(fill=tk.X)

        brands_buttons_frame = tk.Frame(self.brands_main_frame)
        brands_buttons_frame.pack(fill=tk.X, pady=10)

        add_brand_btn = tk.Button(
            brands_buttons_frame,
            text="â• Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ¸Ñ‚ÑŒ Ğ±Ñ€ĞµĞ½Ğ´",
            command=self.add_brand_dialog,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=5,
            cursor="hand2"
        )
        add_brand_btn.pack(side=tk.LEFT, padx=5)

        self.refresh_brands_table()

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ“Š ĞšĞĞ­Ğ¤Ğ¤Ğ˜Ğ¦Ğ˜Ğ•ĞĞ¢Ğ« ĞĞĞ¦Ğ•ĞĞšĞ˜
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        coef_frame = ttk.LabelFrame(currency_scrollable_frame, text="ğŸ“Š ĞšĞ¾ÑÑ„Ñ„Ğ¸Ñ†Ğ¸ĞµĞ½Ñ‚Ñ‹ Ğ½Ğ°Ñ†ĞµĞ½ĞºĞ¸ (%)", padding="15")
        coef_frame.pack(fill=tk.X, pady=10)

        # ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸
        tk.Label(coef_frame, text="ğŸ“Š ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (Ğ½Ğ° Ğ¾ÑĞ½Ğ¾Ğ²Ğµ Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ¸)", font=("Segoe UI", 10, "bold"), fg="#4CAF50").grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° < 15,000â‚½:", font=("Segoe UI", 9)).grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="10%", font=("Segoe UI", 9, "bold")).grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)

        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° â‰¤ 30,000â‚½:", font=("Segoe UI", 9)).grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="9%", font=("Segoe UI", 9, "bold")).grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)

        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° > 30,000â‚½:", font=("Segoe UI", 9)).grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="8%", font=("Segoe UI", 9, "bold")).grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)

        # Ğ Ğ°Ğ·Ğ´ĞµĞ»Ğ¸Ñ‚ĞµĞ»ÑŒ
        ttk.Separator(coef_frame, orient="horizontal").grid(row=4, column=0, columnspan=3, sticky="ew", pady=10)

        # ĞĞ°Ñˆ ĞšÑÑ„
        tk.Label(coef_frame, text="ğŸ’° ĞĞ°Ñˆ ĞšÑÑ„ (Ğ½Ğ° Ğ¾ÑĞ½Ğ¾Ğ²Ğµ Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ¸)", font=("Segoe UI", 10, "bold"), fg="#2196F3").grid(row=5, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° < 10,000â‚½:", font=("Segoe UI", 9)).grid(row=6, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="17%", font=("Segoe UI", 9, "bold")).grid(row=6, column=1, sticky=tk.W, padx=5, pady=2)

        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° â‰¤ 20,000â‚½:", font=("Segoe UI", 9)).grid(row=7, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="15%", font=("Segoe UI", 9, "bold")).grid(row=7, column=1, sticky=tk.W, padx=5, pady=2)

        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° â‰¤ 30,000â‚½:", font=("Segoe UI", 9)).grid(row=8, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="14%", font=("Segoe UI", 9, "bold")).grid(row=8, column=1, sticky=tk.W, padx=5, pady=2)

        tk.Label(coef_frame, text="Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° > 30,000â‚½:", font=("Segoe UI", 9)).grid(row=9, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Label(coef_frame, text="13%", font=("Segoe UI", 9, "bold")).grid(row=9, column=1, sticky=tk.W, padx=5, pady=2)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ“ Ğ¤ĞĞ ĞœĞ£Ğ›Ğ« Ğ ĞĞ¡Ğ§Ğ•Ğ¢Ğ (ĞºÑ€Ğ°Ñ‚ĞºĞ¸Ğ¹ ÑĞ¿Ñ€Ğ°Ğ²Ğ¾Ñ‡Ğ½Ğ¸Ğº)
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        formulas_frame = ttk.LabelFrame(currency_scrollable_frame, text="ğŸ“ Excel Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ (ĞºÑ€Ğ°Ñ‚ĞºĞ¸Ğ¹ ÑĞ¿Ñ€Ğ°Ğ²Ğ¾Ñ‡Ğ½Ğ¸Ğº)", padding="15")
        formulas_frame.pack(fill=tk.X, pady=10)

        formulas_text = tk.Label(
            formulas_frame,
            text=(
                "L (Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ°â‚½)         = VLOOKUP(ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ, Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ°_Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸) Ã— ĞšÑƒÑ€Ñ\n"
                "M (Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°â‚½)          = Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° + (Ğ¦ĞµĞ½Ğ°â‚¬ Ã— ĞšÑƒÑ€Ñ)\n"
                "N (ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ %)        = IFS(Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°<15000, 10%, Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°â‰¤30000, 9%, Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°>30000, 8%)\n"
                "O (ĞĞ°Ñˆ ĞšÑÑ„ %)         = IFS(Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°<10000, 17%, Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°â‰¤20000, 15%, Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°â‰¤30000, 14%, Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°>30000, 13%)\n"
                "P (Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚.â‚½)     = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° Ã— (1 + ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸ + ĞĞ°Ñˆ_ĞšÑÑ„)\n"
                "Q (Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚.â‚½)   = Ğ¦ĞµĞ½Ğ°_Ñ_Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹ - Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ°\n"
                "R (ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ°â‚½)       = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° Ã— ĞĞ°Ñˆ_ĞšÑÑ„\n"
                "S (ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸â‚½)       = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° Ã— ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸"
            ),
            font=("Consolas", 9),
            fg="#333",
            justify=tk.LEFT
        )
        formulas_text.pack(pady=5)

        # ĞŸĞ¾ÑÑĞ½ĞµĞ½Ğ¸Ğµ
        formulas_hint = tk.Label(
            formulas_frame,
            text="ğŸ’¡ Ğ’ÑĞµ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ°Ğ²Ñ‚Ğ¾Ğ¼Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¸ Ğ²ÑÑ‚Ğ°Ğ²Ğ»ÑÑÑ‚ÑÑ Ğ² Excel Ğ¿Ñ€Ğ¸ Ğ½Ğ°Ğ¶Ğ°Ñ‚Ğ¸Ğ¸ 'ğŸ“Š ĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğº Excel'",
            font=("Segoe UI", 9),
            fg="#666"
        )
        formulas_hint.pack(pady=(10, 0))

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # ğŸ¯ ĞšĞĞĞŸĞšĞ˜ Ğ£ĞŸĞ ĞĞ’Ğ›Ğ•ĞĞ˜Ğ¯
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ¸ ÑƒĞ¿Ñ€Ğ°Ğ²Ğ»ĞµĞ½Ğ¸Ñ ĞºÑƒÑ€ÑĞ¾Ğ¼ Ğ¸ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ğ°Ğ¼Ğ¸
        currency_buttons_frame = ttk.Frame(currency_scrollable_frame)
        currency_buttons_frame.pack(pady=20)

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ° Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ñ ĞºÑƒÑ€ÑĞ°
        self.update_rate_btn = tk.Button(
            currency_buttons_frame,
            text="ğŸ”„ ĞĞ±Ğ½Ğ¾Ğ²Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ",
            command=self.update_currency_rate,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=8,
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        )
        self.update_rate_btn.pack(side=tk.LEFT, padx=10)

        # ĞšĞ½Ğ¾Ğ¿ĞºĞ° Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ» Ğº Excel
        self.apply_formulas_btn = tk.Button(
            currency_buttons_frame,
            text="ğŸ“Š ĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğº Excel",
            command=self.apply_formulas_to_excel,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=8,
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        )
        self.apply_formulas_btn.pack(side=tk.LEFT, padx=10)

        # Ğ˜Ğ½Ğ¸Ñ†Ğ¸Ğ°Ğ»Ğ¸Ğ·Ğ°Ñ†Ğ¸Ñ: Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµĞ¼ ĞºÑƒÑ€Ñ Ğ¿Ñ€Ğ¸ Ğ·Ğ°Ğ¿ÑƒÑĞºĞµ
        self.current_eur_rub = 0
        self.update_currency_rate()
        self.log("")
        self.log("ğŸ“ Ğ˜Ğ½ÑÑ‚Ñ€ÑƒĞºÑ†Ğ¸Ñ:")
        self.log("1. ĞĞ°Ğ¶Ğ¼Ğ¸ 'ğŸ“„ Ğ¡Ğ¾Ğ·Ğ´Ğ°Ñ‚ÑŒ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ½Ñ‹Ğ¹ Ñ„Ğ°Ğ¹Ğ»' ĞµÑĞ»Ğ¸ Ñ„Ğ°Ğ¹Ğ»Ğ° ĞµÑ‰Ñ‘ Ğ½ĞµÑ‚")
        self.log("2. ĞÑ‚ĞºÑ€Ğ¾Ğ¹ Excel Ñ„Ğ°Ğ¹Ğ» Ğ¸ Ğ²ÑÑ‚Ğ°Ğ²ÑŒ ÑÑÑ‹Ğ»ĞºĞ¸ Ğ½Ğ° Ñ‚Ğ¾Ğ²Ğ°Ñ€Ñ‹ Ğ² ĞºĞ¾Ğ»Ğ¾Ğ½ĞºÑƒ A")
        self.log("3. Ğ—Ğ°Ğ¿Ğ¾Ğ»Ğ½Ğ¸ Ğ“Ñ€ÑƒĞ¿Ğ¿Ñƒ, ĞŸĞ¾Ğ´Ğ³Ñ€ÑƒĞ¿Ğ¿Ñƒ Ğ¸ Ğ­Ğ¼Ğ¾Ğ´Ğ·Ğ¸ (Ğ¾Ğ¿Ñ†Ğ¸Ğ¾Ğ½Ğ°Ğ»ÑŒĞ½Ğ¾)")
        self.log("4. ĞĞ°Ğ¶Ğ¼Ğ¸ 'ğŸš€ Ğ¡Ğ¿Ğ°Ñ€ÑĞ¸Ñ‚ÑŒ Ğ¸Ğ· Excel Ñ„Ğ°Ğ¹Ğ»Ğ°'")
        self.log("")

        if self.file_path.exists():
            self.log(f"âœ… Ğ¤Ğ°Ğ¹Ğ» {self.file_path.name} ÑƒĞ¶Ğµ ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚")
            self.update_status("Ğ¤Ğ°Ğ¹Ğ» Ğ½Ğ°Ğ¹Ğ´ĞµĞ½, Ğ³Ğ¾Ñ‚Ğ¾Ğ² Ğº Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ñƒ")
        else:
            self.log(f"âš ï¸ Ğ¤Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½ - ÑĞ¾Ğ·Ğ´Ğ°Ğ¹ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½")
            self.update_status("Ğ¡Ğ¾Ğ·Ğ´Ğ°Ğ¹ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ½Ñ‹Ğ¹ Ñ„Ğ°Ğ¹Ğ» Ğ´Ğ»Ñ Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° Ñ€Ğ°Ğ±Ğ¾Ñ‚Ñ‹")

    def log(self, message, color=None):
        """Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµÑ‚ ÑĞ¾Ğ¾Ğ±Ñ‰ĞµĞ½Ğ¸Ğµ Ğ² Ğ»Ğ¾Ğ³."""
        if color:
            tag = f"color_{color}"
            self.log_text.tag_configure(tag, foreground=color)
            self.log_text.insert(tk.END, message + "\n", tag)
        else:
            self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def update_status(self, message):
        """ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµÑ‚ ÑÑ‚Ğ°Ñ‚ÑƒÑ Ğ±Ğ°Ñ€."""
        self.status_bar.config(text=message)
        self.root.update()

    def create_template_clicked(self):
        """ĞĞ±Ñ€Ğ°Ğ±Ğ¾Ñ‚Ñ‡Ğ¸Ğº ĞºĞ½Ğ¾Ğ¿ĞºĞ¸ ÑĞ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ñ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°."""
        self.update_status("Ğ¡Ğ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ğµ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°...")
        self.log("\n" + "=" * 80)
        self.log("ğŸ“„ Ğ¡ĞĞ—Ğ”ĞĞĞ˜Ğ• Ğ¨ĞĞ‘Ğ›ĞĞĞ")
        self.log("=" * 80)

        try:
            if self.file_path.exists():
                response = messagebox.askyesno(
                    "Ğ¤Ğ°Ğ¹Ğ» ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚",
                    f"Ğ¤Ğ°Ğ¹Ğ» {self.file_path.name} ÑƒĞ¶Ğµ ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚.\nĞŸĞµÑ€ĞµĞ·Ğ°Ğ¿Ğ¸ÑĞ°Ñ‚ÑŒ?"
                )
                if not response:
                    self.log("âŒ ĞÑ‚Ğ¼ĞµĞ½ĞµĞ½Ğ¾ Ğ¿Ğ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ĞµĞ¼")
                    self.update_status("ĞÑ‚Ğ¼ĞµĞ½ĞµĞ½Ğ¾")
                    return

            file_path = create_beautiful_template(self.file_path, brands=self.get_brands_from_ui())
            self.log(f"âœ… Ğ¨Ğ°Ğ±Ğ»Ğ¾Ğ½ ÑĞ¾Ğ·Ğ´Ğ°Ğ½: {file_path}")
            self.log("")
            self.log("ğŸ“ Ğ¡Ğ»ĞµĞ´ÑƒÑÑ‰Ğ¸Ğµ ÑˆĞ°Ğ³Ğ¸:")
            self.log(f"   1. ĞÑ‚ĞºÑ€Ğ¾Ğ¹ Ñ„Ğ°Ğ¹Ğ»: {file_path}")
            self.log("   2. Ğ’ÑÑ‚Ğ°Ğ²ÑŒ ÑÑÑ‹Ğ»ĞºĞ¸ Ğ² ĞºĞ¾Ğ»Ğ¾Ğ½ĞºÑƒ 'URL'")
            self.log("   3. Ğ—Ğ°Ğ¿Ğ¾Ğ»Ğ½Ğ¸ Ğ“Ñ€ÑƒĞ¿Ğ¿Ñƒ, ĞŸĞ¾Ğ´Ğ³Ñ€ÑƒĞ¿Ğ¿Ñƒ, Ğ­Ğ¼Ğ¾Ğ´Ğ·Ğ¸")
            self.log("   4. ĞĞ°Ğ¶Ğ¼Ğ¸ 'Ğ¡Ğ¿Ğ°Ñ€ÑĞ¸Ñ‚ÑŒ Ğ¸Ğ· Excel Ñ„Ğ°Ğ¹Ğ»Ğ°'")
            self.log("")

            self.update_status("âœ… Ğ¨Ğ°Ğ±Ğ»Ğ¾Ğ½ ÑƒÑĞ¿ĞµÑˆĞ½Ğ¾ ÑĞ¾Ğ·Ğ´Ğ°Ğ½")

            messagebox.showinfo(
                "Ğ£ÑĞ¿ĞµÑ…",
                f"Ğ¨Ğ°Ğ±Ğ»Ğ¾Ğ½ ÑĞ¾Ğ·Ğ´Ğ°Ğ½!\n\nğŸ“ {file_path}\n\nĞ¢ĞµĞ¿ĞµÑ€ÑŒ Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½Ğ¸ ÑÑÑ‹Ğ»ĞºĞ¸ Ğ¸ Ğ·Ğ°Ğ¿ÑƒÑÑ‚Ğ¸ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³."
            )

        except Exception as e:
            self.log(f"âŒ ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ñ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°: {e}")
            self.update_status("âŒ ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ñ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°")
            messagebox.showerror("ĞÑˆĞ¸Ğ±ĞºĞ°", f"ĞĞµ ÑƒĞ´Ğ°Ğ»Ğ¾ÑÑŒ ÑĞ¾Ğ·Ğ´Ğ°Ñ‚ÑŒ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½:\n{e}")

    def archive_clicked(self):
        """ĞĞ±Ñ€Ğ°Ğ±Ğ¾Ñ‚Ñ‡Ğ¸Ğº ĞºĞ½Ğ¾Ğ¿ĞºĞ¸ Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ñ."""
        self.update_status("ğŸ“¦ Ğ¡Ğ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ğµ Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ°...")
        self.log("\n" + "=" * 80)
        self.log("ğŸ“¦ Ğ¡ĞĞ—Ğ”ĞĞĞ˜Ğ• ZIP ĞĞ Ğ¥Ğ˜Ğ’Ğ Ğ”Ğ›Ğ¯ Ğ‘ĞĞ¢Ğ")
        self.log("=" * 80)
        self.log("")

        try:
            # ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑĞµĞ¼ Ğ½Ğ°Ğ»Ğ¸Ñ‡Ğ¸Ğµ Ñ„Ğ°Ğ¹Ğ»Ğ¾Ğ²
            if not self.file_path.exists():
                messagebox.showwarning(
                    "Ğ¤Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½",
                    f"Excel Ñ„Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½!\n\nĞ¡Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° ÑĞ¿Ğ°Ñ€ÑĞ¸ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ñ‹."
                )
                self.log("âŒ Excel Ñ„Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½")
                self.update_status("âŒ Excel Ñ„Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½")
                return

            images_dir = get_images_dir(self.script_dir)
            if not images_dir.exists() or not any(images_dir.iterdir()):
                messagebox.showwarning(
                    "ĞŸĞ°Ğ¿ĞºĞ° images Ğ¿ÑƒÑÑ‚Ğ°",
                    "ĞŸĞ°Ğ¿ĞºĞ° images/ Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½Ğ° Ğ¸Ğ»Ğ¸ Ğ¿ÑƒÑÑ‚Ğ°!\n\nĞ¡Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° ÑĞ¿Ğ°Ñ€ÑĞ¸ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ñ‹."
                )
                self.log("âŒ ĞŸĞ°Ğ¿ĞºĞ° images/ Ğ¿ÑƒÑÑ‚Ğ°")
                self.update_status("âŒ ĞŸĞ°Ğ¿ĞºĞ° images/ Ğ¿ÑƒÑÑ‚Ğ°")
                return

            # Ğ¡Ğ¾Ğ·Ğ´Ğ°Ñ‘Ğ¼ ZIP Ğ°Ñ€Ñ…Ğ¸Ğ²
            archive_name = f"catalog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            archive_path = self.script_dir / archive_name

            self.log(f"ğŸ“¦ Ğ¡Ğ¾Ğ·Ğ´Ğ°Ñ Ğ°Ñ€Ñ…Ğ¸Ğ²: {archive_name}")
            self.log("")

            with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Excel Ñ„Ğ°Ğ¹Ğ»
                self.log(f"   âœ… Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑÑ {self.file_path.name}")
                zipf.write(self.file_path, self.file_path.name)

                # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Ğ²ÑĞµ Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸Ğ¸ Ğ¸Ğ· Ğ¿Ğ°Ğ¿ĞºĞ¸ images
                image_count = 0
                for image_file in images_dir.iterdir():
                    if image_file.is_file():
                        # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Ñ Ğ¿ÑƒÑ‚Ñ‘Ğ¼ images/filename
                        arcname = f"images/{image_file.name}"
                        zipf.write(image_file, arcname)
                        image_count += 1

                self.log(f"   âœ… Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ğ¾ Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸Ğ¹: {image_count}")

            archive_size = archive_path.stat().st_size / 1024 / 1024  # MB

            self.log("")
            self.log("=" * 80)
            self.log(f"âœ… ĞÑ€Ñ…Ğ¸Ğ² ÑĞ¾Ğ·Ğ´Ğ°Ğ½: {archive_name}")
            self.log(f"ğŸ“ ĞŸÑƒÑ‚ÑŒ: {archive_path}")
            self.log(f"ğŸ’¾ Ğ Ğ°Ğ·Ğ¼ĞµÑ€: {archive_size:.2f} MB")
            self.log("")
            self.log("ğŸ“¤ Ğ¡Ğ»ĞµĞ´ÑƒÑÑ‰Ğ¸Ğµ ÑˆĞ°Ğ³Ğ¸:")
            self.log("   1. ĞÑ‚Ğ¿Ñ€Ğ°Ğ²ÑŒ ÑÑ‚Ğ¾Ñ‚ ZIP Ğ°Ñ€Ñ…Ğ¸Ğ² ÑĞ²Ğ¾ĞµĞ¼Ñƒ Telegram Ğ±Ğ¾Ñ‚Ñƒ")
            self.log("   2. Ğ‘Ğ¾Ñ‚ Ğ°Ğ²Ñ‚Ğ¾Ğ¼Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¸ Ñ€Ğ°ÑĞ¿Ğ°ĞºÑƒĞµÑ‚ Ğ¸ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ¸Ñ‚ ĞºĞ°Ñ‚Ğ°Ğ»Ğ¾Ğ³")
            self.log("   3. Ğ˜ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞ¹ /shop Ñ‡Ñ‚Ğ¾Ğ±Ñ‹ Ğ¾Ñ‚ĞºÑ€Ñ‹Ñ‚ÑŒ Ğ¼Ğ°Ğ³Ğ°Ğ·Ğ¸Ğ½ Ñ Ñ„Ğ¾Ñ‚Ğ¾Ğ³Ñ€Ğ°Ñ„Ğ¸ÑĞ¼Ğ¸")
            self.log("=" * 80)
            self.log("")

            self.update_status(f"âœ… ĞÑ€Ñ…Ğ¸Ğ² ÑĞ¾Ğ·Ğ´Ğ°Ğ½: {archive_name}")

            messagebox.showinfo(
                "ĞÑ€Ñ…Ğ¸Ğ² ÑĞ¾Ğ·Ğ´Ğ°Ğ½!",
                f"âœ… ZIP Ğ°Ñ€Ñ…Ğ¸Ğ² ÑĞ¾Ğ·Ğ´Ğ°Ğ½!\n\n"
                f"ğŸ“ {archive_name}\n"
                f"ğŸ’¾ Ğ Ğ°Ğ·Ğ¼ĞµÑ€: {archive_size:.2f} MB\n\n"
                f"ĞÑ‚Ğ¿Ñ€Ğ°Ğ²ÑŒ ÑÑ‚Ğ¾Ñ‚ Ğ°Ñ€Ñ…Ğ¸Ğ² ÑĞ²Ğ¾ĞµĞ¼Ñƒ Telegram Ğ±Ğ¾Ñ‚Ñƒ\n"
                f"Ğ´Ğ»Ñ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ñ ĞºĞ°Ñ‚Ğ°Ğ»Ğ¾Ğ³Ğ°!"
            )

            # ĞÑ‚ĞºÑ€Ñ‹Ğ²Ğ°ĞµĞ¼ Ğ¿Ğ°Ğ¿ĞºÑƒ Ñ Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ¾Ğ¼
            import subprocess
            subprocess.Popen(f'explorer /select,"{archive_path}"')

        except Exception as e:
            self.log(f"\nâŒ ĞĞ¨Ğ˜Ğ‘ĞšĞ: {e}")
            self.update_status("âŒ ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ñ Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ°")
            messagebox.showerror("ĞÑˆĞ¸Ğ±ĞºĞ°", f"ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞ¾Ğ·Ğ´Ğ°Ğ½Ğ¸Ñ Ğ°Ñ€Ñ…Ğ¸Ğ²Ğ°:\n{e}")

    def parse_clicked(self):
        """ĞĞ±Ñ€Ğ°Ğ±Ğ¾Ñ‚Ñ‡Ğ¸Ğº ĞºĞ½Ğ¾Ğ¿ĞºĞ¸ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°."""
        if not self.file_path.exists():
            messagebox.showwarning(
                "Ğ¤Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½",
                f"Ğ¤Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½!\n\nĞ¡Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° ÑĞ¾Ğ·Ğ´Ğ°Ğ¹ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½."
            )
            return

        # Ğ—Ğ°Ğ¿ÑƒÑĞºĞ°ĞµĞ¼ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³ Ğ² Ğ¾Ñ‚Ğ´ĞµĞ»ÑŒĞ½Ğ¾Ğ¼ Ğ¿Ğ¾Ñ‚Ğ¾ĞºĞµ
        thread = threading.Thread(target=self.parse_excel, daemon=True)
        thread.start()

    def parse_excel(self):
        """ĞŸĞ°Ñ€ÑĞ¸Ñ‚ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ñ‹ Ğ¸Ğ· Excel (Ğ¼Ğ½Ğ¾Ğ³Ğ¾Ğ¿Ğ¾Ñ‚Ğ¾Ñ‡Ğ½Ğ¾)."""
        # Ğ‘Ğ»Ğ¾ĞºĞ¸Ñ€ÑƒĞµĞ¼ ĞºĞ½Ğ¾Ğ¿ĞºĞ¸
        self.create_btn.config(state=tk.DISABLED)
        self.parse_btn.config(state=tk.DISABLED)

        num_threads = self.threads_var.get()
        self.update_status(f"ğŸ•·ï¸ ĞŸĞ°Ñ€ÑĞ¸Ğ½Ğ³ Ğ² Ğ¿Ñ€Ğ¾Ñ†ĞµÑÑĞµ ({num_threads} Ğ¿Ğ¾Ñ‚Ğ¾ĞºĞ¾Ğ²)...")
        self.log("\n" + "=" * 80)
        self.log(f"ğŸ•·ï¸ ĞŸĞĞ Ğ¡Ğ˜ĞĞ“ Ğ¢ĞĞ’ĞĞ ĞĞ’ ({num_threads} Ğ¿Ğ¾Ñ‚Ğ¾ĞºĞ¾Ğ²)")
        self.log("=" * 80)
        self.log("")

        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            updated_count = 0
            error_count = 0
            total_rows = ws.max_row - 1  # ĞœĞ¸Ğ½ÑƒÑ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº

            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            # 1. Ğ¡Ğ¾Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ Ğ·Ğ°Ğ´Ğ°Ñ‡Ğ¸ Ğ´Ğ»Ñ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°
            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            tasks = []
            for row_num in range(2, ws.max_row + 1):
                url = ws.cell(row_num, 1).value
                if not url or not url.startswith("http"):
                    self.log(f"[{row_num - 1}/{total_rows}] â­ï¸ ĞŸÑ€Ğ¾Ğ¿ÑƒÑ‰ĞµĞ½Ğ¾ (Ğ½ĞµÑ‚ URL)")
                    ws.cell(row_num, 11).value = "ĞŸÑ€Ğ¾Ğ¿ÑƒÑ‰ĞµĞ½Ğ¾ (Ğ½ĞµÑ‚ URL)"
                    continue
                tasks.append((row_num, url, row_num - 1))

            self.log(f"ğŸ“‹ ĞĞ°Ğ¹Ğ´ĞµĞ½Ğ¾ {len(tasks)} ÑÑÑ‹Ğ»Ğ¾Ğº Ğ´Ğ»Ñ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°\n")

            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            # 2. ĞŸĞ°Ñ€Ğ°Ğ»Ğ»ĞµĞ»ÑŒĞ½Ñ‹Ğ¹ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³ Ñ‡ĞµÑ€ĞµĞ· ThreadPoolExecutor
            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            results = {}  # {row_num: (product_data, error)}
            completed = 0

            with ThreadPoolExecutor(max_workers=num_threads) as executor:
                future_to_row = {
                    executor.submit(parse_product, url, self.script_dir, pid): (row_num, url, pid)
                    for row_num, url, pid in tasks
                }

                for future in as_completed(future_to_row):
                    row_num, url, pid = future_to_row[future]
                    completed += 1

                    try:
                        product_data, error = future.result()
                        results[row_num] = (product_data, error)

                        if error:
                            self.log(f"[{completed}/{len(tasks)}] âŒ #{pid}: {error}", color="red")
                        else:
                            photos = len(product_data['image_urls'].split(',')) if product_data.get('image_urls') else 0
                            price = product_data.get('price')

                            if not price or not photos:
                                missing = []
                                if not price: missing.append("Ğ½ĞµÑ‚ Ñ†ĞµĞ½Ñ‹")
                                if not photos: missing.append("Ğ½ĞµÑ‚ Ñ„Ğ¾Ñ‚Ğ¾")
                                self.log(f"[{completed}/{len(tasks)}] âš ï¸ #{pid}: {product_data['name']} | {price or '???'}â‚¬ | ğŸ“·{photos} â€” {', '.join(missing)}", color="red")
                            else:
                                self.log(f"[{completed}/{len(tasks)}] âœ… #{pid}: {product_data['name']} | {price}â‚¬ | ğŸ“·{photos}")
                    except Exception as e:
                        results[row_num] = (None, str(e))
                        self.log(f"[{completed}/{len(tasks)}] âŒ #{pid}: {e}", color="red")

                    self.update_status(f"ğŸ•·ï¸ ĞŸĞ°Ñ€ÑĞ¸Ğ½Ğ³: {completed}/{len(tasks)} ({num_threads} Ğ¿Ğ¾Ñ‚Ğ¾ĞºĞ¾Ğ²)")

            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            # 3. Ğ§Ğ¸Ñ‚Ğ°ĞµĞ¼ ÑĞ¿Ğ¸ÑĞ¾Ğº Ğ±Ñ€ĞµĞ½Ğ´Ğ¾Ğ² Ğ¸Ğ· Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            brands_list = []
            if "âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸" in wb.sheetnames:
                settings_ws = wb["âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸"]
                for row in range(3, 100):
                    brand = settings_ws[f'D{row}'].value
                    if brand and str(brand).strip():
                        brands_list.append(str(brand).strip())
                    elif row > 10:
                        break
            self.log(f"\nğŸ·ï¸ Ğ‘Ñ€ĞµĞ½Ğ´Ğ¾Ğ²: {len(brands_list)} ({', '.join(brands_list[:5])}{'...' if len(brands_list) > 5 else ''})")

            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            # 4. Ğ—Ğ°Ğ¿Ğ¸ÑÑ‹Ğ²Ğ°ĞµĞ¼ Ñ€ĞµĞ·ÑƒĞ»ÑŒÑ‚Ğ°Ñ‚Ñ‹ Ğ² Excel (Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒĞ½Ğ¾)
            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            self.log("ğŸ“ Ğ—Ğ°Ğ¿Ğ¸ÑÑ‹Ğ²Ğ°Ñ Ñ€ĞµĞ·ÑƒĞ»ÑŒÑ‚Ğ°Ñ‚Ñ‹ Ğ² Excel...")

            data_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
            data_font = Font(size=11, name="Calibri")
            left_alignment = Alignment(horizontal="left", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")

            for row_num, (product_data, error) in sorted(results.items()):
                if error:
                    ws.cell(row_num, 11).value = error
                    error_count += 1
                else:
                    # ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑĞµĞ¼, ĞµÑÑ‚ÑŒ Ğ»Ğ¸ Ğ²Ñ€ÑƒÑ‡Ğ½ÑƒÑ Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½ĞµĞ½Ğ½Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ
                    existing_category = ws.cell(row_num, 6).value
                    if existing_category:
                        self.log(f"   ğŸ“‹ #{row_num-1}: ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ğ°: {existing_category}")

                    # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ Ğ¢ĞĞ›Ğ¬ĞšĞ Ğ°Ğ²Ñ‚Ğ¾Ğ¼Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¸ Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½ÑĞµĞ¼Ñ‹Ğµ Ğ¿Ğ¾Ğ»Ñ
                    ws.cell(row_num, 2).value = product_data['name']           # B: ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ
                    ws.cell(row_num, 3).value = product_data['price']          # C: Ğ¦ĞµĞ½Ğ°
                    # D: Ğ“Ñ€ÑƒĞ¿Ğ¿Ğ° (ĞĞ• Ğ¢Ğ ĞĞ“ĞĞ•Ğœ - Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½ÑĞµÑ‚ÑÑ Ğ²Ñ€ÑƒÑ‡Ğ½ÑƒÑ)
                    # E: ĞŸĞ¾Ğ´Ğ³Ñ€ÑƒĞ¿Ğ¿Ğ° (ĞĞ• Ğ¢Ğ ĞĞ“ĞĞ•Ğœ - Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½ÑĞµÑ‚ÑÑ Ğ²Ñ€ÑƒÑ‡Ğ½ÑƒÑ)
                    # F: ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ° (ĞĞ• Ğ¢Ğ ĞĞ“ĞĞ•Ğœ - Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½ÑĞµÑ‚ÑÑ Ğ²Ñ€ÑƒÑ‡Ğ½ÑƒÑ)
                    ws.cell(row_num, 7).value = product_data['image_urls']     # G: URL Ñ„Ğ¾Ñ‚Ğ¾
                    ws.cell(row_num, 8).value = product_data['local_images']   # H: Ğ›Ğ¾ĞºĞ°Ğ»ÑŒĞ½Ğ¾Ğµ Ñ„Ğ¾Ñ‚Ğ¾
                    ws.cell(row_num, 9).value = product_data.get('sizes', '')  # I: Ğ Ğ°Ğ·Ğ¼ĞµÑ€Ñ‹
                    ws.cell(row_num, 10).value = datetime.now().strftime("%Y-%m-%d %H:%M")  # J: ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ
                    ws.cell(row_num, 11).value = "âœ… ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¾"                # K: Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ

                    # T(20): Ğ‘Ñ€ĞµĞ½Ğ´ â€” Ğ¾Ğ¿Ñ€ĞµĞ´ĞµĞ»ÑĞµĞ¼ Ğ¸Ğ· Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ñ
                    detected_brand = ""
                    name_lower = product_data['name'].lower()
                    for brand in brands_list:
                        if brand.lower() in name_lower:
                            detected_brand = brand
                            break
                    ws.cell(row_num, 20).value = detected_brand  # T: Ğ‘Ñ€ĞµĞ½Ğ´

                    # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ¸Ğµ Ğº ÑÑ‡ĞµĞ¹ĞºĞ°Ğ¼
                    cell_a = ws.cell(row_num, 1)
                    cell_a.border = data_border
                    cell_a.font = data_font
                    cell_a.alignment = left_alignment

                    cell_b = ws.cell(row_num, 2)
                    cell_b.border = data_border
                    cell_b.font = data_font
                    cell_b.alignment = left_alignment
                    cell_b.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                    cell_c = ws.cell(row_num, 3)
                    cell_c.border = data_border
                    cell_c.font = data_font
                    cell_c.alignment = center_alignment
                    cell_c.number_format = '#,##0.00'
                    cell_c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                    for col in [4, 5, 6]:
                        cell = ws.cell(row_num, col)
                        cell.border = data_border
                        cell.font = data_font
                        cell.alignment = left_alignment

                    cell_g = ws.cell(row_num, 7)
                    cell_g.border = data_border
                    cell_g.font = Font(size=9, name="Calibri")
                    cell_g.alignment = left_alignment

                    cell_h = ws.cell(row_num, 8)
                    cell_h.border = data_border
                    cell_h.font = data_font
                    cell_h.alignment = left_alignment

                    cell_i = ws.cell(row_num, 9)
                    cell_i.border = data_border
                    cell_i.font = data_font
                    cell_i.alignment = center_alignment

                    cell_j = ws.cell(row_num, 10)
                    cell_j.border = data_border
                    cell_j.font = Font(size=10, name="Calibri")
                    cell_j.alignment = center_alignment

                    cell_k = ws.cell(row_num, 11)
                    cell_k.border = data_border
                    cell_k.font = data_font
                    cell_k.alignment = center_alignment

                    for col in range(12, 20):
                        cell = ws.cell(row_num, col)
                        if cell.value is not None:
                            cell.border = data_border

                    updated_count += 1

            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            # ğŸ“‹ ĞĞ‘ĞĞĞ’Ğ›Ğ¯Ğ•Ğœ Ğ’Ğ«ĞŸĞĞ”ĞĞ®Ğ©Ğ˜Ğ™ Ğ¡ĞŸĞ˜Ğ¡ĞĞš ĞšĞĞ¢Ğ•Ğ“ĞĞ Ğ˜Ğ™
            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

            # Ğ£Ğ´Ğ°Ğ»ÑĞµĞ¼ ÑÑ‚Ğ°Ñ€ÑƒÑ Ğ²Ğ°Ğ»Ğ¸Ğ´Ğ°Ñ†Ğ¸Ñ (ĞµÑĞ»Ğ¸ ĞµÑÑ‚ÑŒ)
            ws.data_validations.dataValidation = [
                dv for dv in ws.data_validations.dataValidation
                if dv.sqref and 'F' not in str(dv.sqref).split(':')[0]
            ]

            # Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµĞ¼ Ğ½Ğ¾Ğ²ÑƒÑ Ğ²Ğ°Ğ»Ğ¸Ğ´Ğ°Ñ†Ğ¸Ñ Ñ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸ÑĞ¼Ğ¸ Ğ¸Ğ· Ğ»Ğ¸ÑÑ‚Ğ° Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº (ĞµÑĞ»Ğ¸ ĞµÑÑ‚ÑŒ)
            if "âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸" in wb.sheetnames:
                # Ğ’Ñ‹Ñ‡Ğ¸ÑĞ»ÑĞµĞ¼ ĞºĞ¾Ğ½ĞµÑ‡Ğ½ÑƒÑ ÑÑ‚Ñ€Ğ¾ĞºÑƒ Ğ´Ğ¸Ğ½Ğ°Ğ¼Ğ¸Ñ‡ĞµÑĞºĞ¸ Ğ½Ğ° Ğ¾ÑĞ½Ğ¾Ğ²Ğµ ĞºĞ¾Ğ»Ğ¸Ñ‡ĞµÑÑ‚Ğ²Ğ° ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹
                settings_ws = wb["âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸"]
                last_row = 8  # Ğ¡Ñ‚Ñ€Ğ¾ĞºĞ° Ğ¿ĞµÑ€ĞµĞ´ Ğ¿ĞµÑ€Ğ²Ğ¾Ğ¹ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸ĞµĞ¹

                # Ğ˜Ñ‰ĞµĞ¼ Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½ÑÑ Ğ·Ğ°Ğ¿Ğ¾Ğ»Ğ½ĞµĞ½Ğ½ÑƒÑ ÑÑ‚Ñ€Ğ¾ĞºÑƒ Ñ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸ĞµĞ¹ (Ğ½Ğ°Ñ‡Ğ¸Ğ½Ğ°Ñ ÑĞ¾ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ 9)
                for row in range(9, 100):
                    if settings_ws[f'A{row}'].value:
                        last_row = row
                    else:
                        break

                # Ğ˜ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞµĞ¼ ÑÑÑ‹Ğ»ĞºÑƒ Ğ½Ğ° Ğ»Ğ¸ÑÑ‚ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº Ğ´Ğ»Ñ Ğ´Ğ¸Ğ½Ğ°Ğ¼Ğ¸Ñ‡ĞµÑĞºĞ¾Ğ³Ğ¾ ÑĞ¿Ğ¸ÑĞºĞ°
                categories_formula = f"'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$A$9:$A${last_row}"
                dv_category = DataValidation(
                    type="list",
                    formula1=categories_formula,
                    allow_blank=True,
                    showDropDown=False,  # False = Ğ¿Ğ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ ÑÑ‚Ñ€ĞµĞ»ĞºÑƒ Ğ²Ñ‹Ğ¿Ğ°Ğ´Ğ°ÑÑ‰ĞµĞ³Ğ¾ ÑĞ¿Ğ¸ÑĞºĞ°
                    showInputMessage=False,  # ĞĞµ Ğ¿Ğ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ‡Ğ°Ğ½Ğ¸Ğµ
                    showErrorMessage=True
                )
            else:
                # Ğ•ÑĞ»Ğ¸ Ğ»Ğ¸ÑÑ‚Ğ° Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº Ğ½ĞµÑ‚, Ğ¸ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞµĞ¼ ÑÑ‚Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¸Ğ¹ ÑĞ¿Ğ¸ÑĞ¾Ğº
                categories = ["ĞÑ‡ĞºĞ¸", "Ğ Ğ°ĞºĞµÑ‚ĞºĞ°", "ĞšÑ€Ğ¾ÑÑĞ¾Ğ²ĞºĞ¸", "ĞšÑƒÑ€Ñ‚ĞºĞ°", "Ğ¨Ñ‚Ğ°Ğ½Ñ‹", "Ğ¨Ğ»ĞµĞ¼", "Ğ‘Ğ¾Ñ‚Ğ¸Ğ½ĞºĞ¸ Ğ±Ğ¾Ñ€Ğ´", "Ğ¢ĞµÑ€Ğ¼Ğ¾", "ĞÑ‡ĞºĞ¸ Ğ´Ğ»Ñ ÑĞ½ĞµĞ³Ğ°"]
                categories_formula = f'"{",".join(categories)}"'
                dv_category = DataValidation(
                    type="list",
                    formula1=categories_formula,
                    allow_blank=True,
                    showDropDown=False,  # False = Ğ¿Ğ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ ÑÑ‚Ñ€ĞµĞ»ĞºÑƒ Ğ²Ñ‹Ğ¿Ğ°Ğ´Ğ°ÑÑ‰ĞµĞ³Ğ¾ ÑĞ¿Ğ¸ÑĞºĞ°
                    showInputMessage=False,  # ĞĞµ Ğ¿Ğ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ‡Ğ°Ğ½Ğ¸Ğµ
                    showErrorMessage=True
                )

            dv_category.error = "Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ğ¸Ğ· ÑĞ¿Ğ¸ÑĞºĞ° Ğ´Ğ¾Ğ¿ÑƒÑÑ‚Ğ¸Ğ¼Ñ‹Ñ… Ğ·Ğ½Ğ°Ñ‡ĞµĞ½Ğ¸Ğ¹!"
            dv_category.errorTitle = "âŒ ĞĞµĞ²ĞµÑ€Ğ½Ğ°Ñ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ"

            ws.add_data_validation(dv_category)
            # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğº ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ñƒ F Ğ´Ğ»Ñ Ğ²ÑĞµÑ… ÑÑ‚Ñ€Ğ¾Ğº (Ğ²ĞºĞ»ÑÑ‡Ğ°Ñ Ğ½Ğ¾Ğ²Ñ‹Ğµ)
            max_row = ws.max_row if ws.max_row > 2 else 10000
            dv_category.add(f'F2:F{max_row}')

            # Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµĞ¼
            wb.save(self.file_path)

            self.log("")
            self.log("=" * 80)
            self.log(f"âœ… ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¾ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²: {updated_count}")
            self.log(f"âŒ ĞÑˆĞ¸Ğ±Ğ¾Ğº: {error_count}")
            self.log(f"ğŸ“„ Ğ¤Ğ°Ğ¹Ğ» ÑĞ¾Ñ…Ñ€Ğ°Ğ½Ñ‘Ğ½: {self.file_path}")
            self.log(f"ğŸ“ Ğ¤Ğ¾Ñ‚ĞºĞ¸ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ñ‹ Ğ²: {self.script_dir / 'images'}")
            self.log("=" * 80)
            self.log("")

            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
            # ğŸ“Š ĞĞ’Ğ¢ĞĞœĞĞ¢Ğ˜Ğ§Ğ•Ğ¡ĞšĞĞ• ĞŸĞ Ğ˜ĞœĞ•ĞĞ•ĞĞ˜Ğ• Ğ¤ĞĞ ĞœĞ£Ğ› ĞŸĞĞ¡Ğ›Ğ• ĞŸĞĞ Ğ¡Ğ˜ĞĞ“Ğ
            # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

            self.log("ğŸ“Š ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑÑ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ñ€Ğ°ÑÑ‡ĞµÑ‚Ğ° Ğº Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°Ğ¼...")
            self.update_status("ğŸ“Š ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ğµ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»...")

            # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ±ĞµĞ· messagebox (Ñ‚Ğ¸Ñ…Ğ¾)
            try:
                self.apply_formulas_silently()
                self.log("âœ… Ğ¤Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ñ‹! Ğ¡Ñ‚Ğ¾Ğ»Ğ±Ñ†Ñ‹ L-S Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ñ‹.")
            except Exception as e:
                self.log(f"âš ï¸ ĞĞµ ÑƒĞ´Ğ°Ğ»Ğ¾ÑÑŒ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹: {e}")
                self.log("   ĞœĞ¾Ğ¶Ğ½Ğ¾ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ²Ñ€ÑƒÑ‡Ğ½ÑƒÑ Ğ½Ğ° Ğ²ĞºĞ»Ğ°Ğ´ĞºĞµ 'ĞšÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹'")

            self.log("")
            self.update_status(f"âœ… ĞŸĞ°Ñ€ÑĞ¸Ğ½Ğ³ Ğ·Ğ°Ğ²ĞµÑ€ÑˆÑ‘Ğ½: {updated_count} Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ² Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¾")

            messagebox.showinfo(
                "ĞŸĞ°Ñ€ÑĞ¸Ğ½Ğ³ Ğ·Ğ°Ğ²ĞµÑ€ÑˆÑ‘Ğ½",
                f"âœ… ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¾ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²: {updated_count}\nâŒ ĞÑˆĞ¸Ğ±Ğ¾Ğº: {error_count}\n\nğŸ“Š Ğ¤Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ñ‹ Ğ°Ğ²Ñ‚Ğ¾Ğ¼Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¸!\n\nğŸ“„ {self.file_path}\nğŸ“ Ğ¤Ğ¾Ñ‚ĞºĞ¸: {self.script_dir / 'images'}"
            )

        except Exception as e:
            self.log(f"\nâŒ ĞĞ¨Ğ˜Ğ‘ĞšĞ: {e}")
            self.update_status("âŒ ĞÑˆĞ¸Ğ±ĞºĞ° Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°")
            messagebox.showerror("ĞÑˆĞ¸Ğ±ĞºĞ°", f"ĞÑˆĞ¸Ğ±ĞºĞ° Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°:\n{e}")

        finally:
            # Ğ Ğ°Ğ·Ğ±Ğ»Ğ¾ĞºĞ¸Ñ€ÑƒĞµĞ¼ ĞºĞ½Ğ¾Ğ¿ĞºĞ¸
            self.create_btn.config(state=tk.NORMAL)
            self.parse_btn.config(state=tk.NORMAL)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ğŸ’± ĞœĞ•Ğ¢ĞĞ”Ğ« Ğ”Ğ›Ğ¯ Ğ ĞĞ‘ĞĞ¢Ğ« Ğ¡ ĞšĞ£Ğ Ğ¡ĞĞœ Ğ’ĞĞ›Ğ®Ğ¢Ğ«
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    def _load_settings(self):
        """Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµÑ‚ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸ Ğ¸Ğ· JSON Ñ„Ğ°Ğ¹Ğ»Ğ°."""
        try:
            if self.settings_file.exists():
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def _save_settings(self):
        """Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµÑ‚ Ñ‚ĞµĞºÑƒÑ‰Ğ¸Ğµ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸ Ğ² JSON Ñ„Ğ°Ğ¹Ğ»."""
        try:
            # Ğ¡Ñ‡Ğ¸Ñ‚Ñ‹Ğ²Ğ°ĞµĞ¼ Ğ°ĞºÑ‚ÑƒĞ°Ğ»ÑŒĞ½Ñ‹Ğµ Ğ±Ñ€ĞµĞ½Ğ´Ñ‹ Ğ¸Ğ· Ğ¿Ğ¾Ğ»ĞµĞ¹ Ğ²Ğ²Ğ¾Ğ´Ğ°
            brands = self.get_brands_from_ui()

            # Ğ¡Ñ‡Ğ¸Ñ‚Ñ‹Ğ²Ğ°ĞµĞ¼ Ğ½Ğ°Ğ´Ğ±Ğ°Ğ²ĞºÑƒ
            try:
                markup = float(self.markup_entry.get())
            except ValueError:
                markup = 0.5

            settings = {
                "categories": self.categories_data,
                "brands": brands,
                "markup": markup,
                "threads": self.threads_var.get(),
            }
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def refresh_categories_table(self):
        """ĞŸĞµÑ€ĞµÑ€Ğ¸ÑĞ¾Ğ²Ñ‹Ğ²Ğ°ĞµÑ‚ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñƒ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹."""
        # ĞÑ‡Ğ¸Ñ‰Ğ°ĞµĞ¼ ÑÑ‚Ğ°Ñ€ÑƒÑ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñƒ
        for widget in self.categories_table_frame.winfo_children():
            widget.destroy()

        # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸
        tk.Label(self.categories_table_frame, text="ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        tk.Label(self.categories_table_frame, text="Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚¬)", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(self.categories_table_frame, text="", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, padx=5, pady=5)

        # Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµĞ¼ Ğ·Ğ°Ğ¿Ğ¸ÑĞ¸ Ğ´Ğ»Ñ ĞºĞ°Ğ¶Ğ´Ğ¾Ğ¹ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸
        self.category_entries = {}

        for idx, cat_data in enumerate(self.categories_data, start=1):
            # ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸ (Ñ€ĞµĞ´Ğ°ĞºÑ‚Ğ¸Ñ€ÑƒĞµĞ¼Ğ¾Ğµ)
            name_entry = tk.Entry(self.categories_table_frame, font=("Segoe UI", 10), width=20)
            name_entry.insert(0, cat_data["name"])
            name_entry.grid(row=idx, column=0, padx=5, pady=2, sticky=tk.W)

            # Ğ¡Ñ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ (Ñ€ĞµĞ´Ğ°ĞºÑ‚Ğ¸Ñ€ÑƒĞµĞ¼Ğ¾Ğµ)
            delivery_entry = tk.Entry(self.categories_table_frame, font=("Segoe UI", 10), width=10)
            delivery_entry.insert(0, str(cat_data["delivery"]))
            delivery_entry.grid(row=idx, column=1, padx=5, pady=2)

            # Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµĞ¼ ÑÑÑ‹Ğ»ĞºĞ¸ Ğ½Ğ° entry
            self.category_entries[idx - 1] = {
                "name": name_entry,
                "delivery": delivery_entry
            }

            # ĞšĞ½Ğ¾Ğ¿ĞºĞ° ÑƒĞ´Ğ°Ğ»ĞµĞ½Ğ¸Ñ
            delete_btn = tk.Button(
                self.categories_table_frame,
                text="ğŸ—‘ï¸",
                command=lambda i=idx-1: self.delete_category(i),
                bg="#f44336",
                fg="white",
                font=("Segoe UI", 9),
                width=3,
                cursor="hand2"
            )
            delete_btn.grid(row=idx, column=2, padx=5, pady=2)

    def add_category_dialog(self):
        """Ğ”Ğ¸Ğ°Ğ»Ğ¾Ğ³ Ğ´Ğ»Ñ Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ğ¸Ñ Ğ½Ğ¾Ğ²Ğ¾Ğ¹ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸."""
        from tkinter import simpledialog

        # Ğ—Ğ°Ğ¿Ñ€Ğ°ÑˆĞ¸Ğ²Ğ°ĞµĞ¼ Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸
        category_name = simpledialog.askstring(
            "Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ¸Ñ‚ÑŒ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ",
            "Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸:",
            parent=self.root
        )

        if not category_name or not category_name.strip():
            return

        # Ğ—Ğ°Ğ¿Ñ€Ğ°ÑˆĞ¸Ğ²Ğ°ĞµĞ¼ ÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
        delivery_cost = simpledialog.askstring(
            "Ğ¡Ñ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸",
            f"Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ Ğ´Ğ»Ñ '{category_name}' (â‚¬):",
            parent=self.root
        )

        if not delivery_cost:
            return

        try:
            delivery_cost = float(delivery_cost)
        except ValueError:
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                "ĞĞµĞ²ĞµÑ€Ğ½Ñ‹Ğ¹ Ñ„Ğ¾Ñ€Ğ¼Ğ°Ñ‚ ÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸!\nĞ˜ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞ¹Ñ‚Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾ (Ğ½Ğ°Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ€: 17)"
            )
            return

        # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Ğ½Ğ¾Ğ²ÑƒÑ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ
        self.categories_data.append({
            "name": category_name.strip(),
            "delivery": delivery_cost
        })

        # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñƒ
        self.refresh_categories_table()
        self._save_settings()

        messagebox.showinfo(
            "Ğ“Ğ¾Ñ‚Ğ¾Ğ²Ğ¾!",
            f"âœ… ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ '{category_name}' Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ğ°!"
        )

    def delete_category(self, index):
        """Ğ£Ğ´Ğ°Ğ»ÑĞµÑ‚ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ Ğ¿Ğ¾ Ğ¸Ğ½Ğ´ĞµĞºÑÑƒ."""
        if len(self.categories_data) <= 1:
            messagebox.showwarning(
                "ĞĞµĞ»ÑŒĞ·Ñ ÑƒĞ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ",
                "Ğ”Ğ¾Ğ»Ğ¶Ğ½Ğ° Ğ¾ÑÑ‚Ğ°Ñ‚ÑŒÑÑ Ñ…Ğ¾Ñ‚Ñ Ğ±Ñ‹ Ğ¾Ğ´Ğ½Ğ° ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ!"
            )
            return

        category_name = self.categories_data[index]["name"]

        result = messagebox.askyesno(
            "Ğ£Ğ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ?",
            f"Ğ’Ñ‹ ÑƒĞ²ĞµÑ€ĞµĞ½Ñ‹, Ñ‡Ñ‚Ğ¾ Ñ…Ğ¾Ñ‚Ğ¸Ñ‚Ğµ ÑƒĞ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ '{category_name}'?\n\nĞ­Ñ‚Ğ¾ Ğ´ĞµĞ¹ÑÑ‚Ğ²Ğ¸Ğµ Ğ½ĞµĞ»ÑŒĞ·Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ."
        )

        if result:
            self.categories_data.pop(index)
            self.refresh_categories_table()
            self._save_settings()

            messagebox.showinfo(
                "Ğ£Ğ´Ğ°Ğ»ĞµĞ½Ğ¾!",
                f"âœ… ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ '{category_name}' ÑƒĞ´Ğ°Ğ»ĞµĞ½Ğ°!"
            )

    def refresh_brands_table(self):
        """ĞŸĞµÑ€ĞµÑ€Ğ¸ÑĞ¾Ğ²Ñ‹Ğ²Ğ°ĞµÑ‚ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñƒ Ğ±Ñ€ĞµĞ½Ğ´Ğ¾Ğ²."""
        for widget in self.brands_table_frame.winfo_children():
            widget.destroy()

        self.brand_entries = {}

        # Ğ Ğ°Ğ·Ğ¼ĞµÑ‰Ğ°ĞµĞ¼ Ğ±Ñ€ĞµĞ½Ğ´Ñ‹ Ğ² 3 ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ¸
        for idx, brand in enumerate(self.brands_data):
            row = idx // 3
            col = (idx % 3) * 2  # 2 ÑÑ‡ĞµĞ¹ĞºĞ¸ Ğ½Ğ° Ğ±Ñ€ĞµĞ½Ğ´ (Entry + ĞºĞ½Ğ¾Ğ¿ĞºĞ°)

            entry = tk.Entry(self.brands_table_frame, font=("Segoe UI", 10), width=15)
            entry.insert(0, brand)
            entry.grid(row=row, column=col, padx=3, pady=2, sticky=tk.W)
            self.brand_entries[idx] = entry

            delete_btn = tk.Button(
                self.brands_table_frame,
                text="âœ•",
                command=lambda i=idx: self.delete_brand(i),
                bg="#f44336",
                fg="white",
                font=("Segoe UI", 8),
                width=2,
                cursor="hand2"
            )
            delete_btn.grid(row=row, column=col + 1, padx=(0, 10), pady=2)

    def add_brand_dialog(self):
        """Ğ”Ğ¸Ğ°Ğ»Ğ¾Ğ³ Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ğ¸Ñ Ğ±Ñ€ĞµĞ½Ğ´Ğ°."""
        from tkinter import simpledialog
        brand_name = simpledialog.askstring("Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ¸Ñ‚ÑŒ Ğ±Ñ€ĞµĞ½Ğ´", "Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ Ğ±Ñ€ĞµĞ½Ğ´Ğ°:", parent=self.root)
        if not brand_name or not brand_name.strip():
            return
        self.brands_data.append(brand_name.strip())
        self.refresh_brands_table()
        self._save_settings()

    def delete_brand(self, index):
        """Ğ£Ğ´Ğ°Ğ»ÑĞµÑ‚ Ğ±Ñ€ĞµĞ½Ğ´ Ğ¿Ğ¾ Ğ¸Ğ½Ğ´ĞµĞºÑÑƒ."""
        self.brands_data.pop(index)
        self.refresh_brands_table()
        self._save_settings()

    def get_brands_from_ui(self):
        """Ğ¡Ñ‡Ğ¸Ñ‚Ñ‹Ğ²Ğ°ĞµÑ‚ Ğ°ĞºÑ‚ÑƒĞ°Ğ»ÑŒĞ½Ñ‹Ğµ Ğ±Ñ€ĞµĞ½Ğ´Ñ‹ Ğ¸Ğ· Ğ¿Ğ¾Ğ»ĞµĞ¹ Ğ²Ğ²Ğ¾Ğ´Ğ°."""
        brands = []
        for idx, entry in self.brand_entries.items():
            val = entry.get().strip()
            if val:
                brands.append(val)
        self.brands_data = brands
        return brands

    def save_category_changes(self):
        """Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµÑ‚ Ğ¸Ğ·Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹ Ğ¸ ÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ Ğ² Excel."""
        try:
            # Ğ¡Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ self.categories_data Ğ¸Ğ· Ğ¿Ğ¾Ğ»ĞµĞ¹ Ğ²Ğ²Ğ¾Ğ´Ğ°
            for idx, entries in self.category_entries.items():
                try:
                    new_name = entries["name"].get().strip()
                    new_delivery = float(entries["delivery"].get())

                    if not new_name:
                        raise ValueError("ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸ Ğ½Ğµ Ğ¼Ğ¾Ğ¶ĞµÑ‚ Ğ±Ñ‹Ñ‚ÑŒ Ğ¿ÑƒÑÑ‚Ñ‹Ğ¼")

                    self.categories_data[idx]["name"] = new_name
                    self.categories_data[idx]["delivery"] = new_delivery

                except ValueError as e:
                    messagebox.showerror(
                        "ĞÑˆĞ¸Ğ±ĞºĞ°",
                        f"ĞĞµĞ²ĞµÑ€Ğ½Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ Ğ² ÑÑ‚Ñ€Ğ¾ĞºĞµ {idx + 1}:\n{e}\n\nĞ˜ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞ¹Ñ‚Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾ Ğ´Ğ»Ñ ÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ (Ğ½Ğ°Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ€: 17)"
                    )
                    return

            if not self.file_path.exists():
                messagebox.showwarning(
                    "Ğ¤Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½",
                    f"Excel Ñ„Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½!\nĞ¡Ğ¾Ğ·Ğ´Ğ°Ğ¹Ñ‚Ğµ Ñ„Ğ°Ğ¹Ğ» ÑĞ½Ğ°Ñ‡Ğ°Ğ»Ğ°."
                )
                return

            wb = load_workbook(self.file_path)

            # ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑĞµĞ¼ Ğ½Ğ°Ğ»Ğ¸Ñ‡Ğ¸Ğµ Ğ»Ğ¸ÑÑ‚Ğ° Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
            if "âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸" not in wb.sheetnames:
                messagebox.showwarning(
                    "Ğ›Ğ¸ÑÑ‚ Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½",
                    "Ğ›Ğ¸ÑÑ‚ 'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸' Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½!\nĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚Ğµ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ ÑĞ½Ğ°Ñ‡Ğ°Ğ»Ğ°."
                )
                wb.close()
                return

            settings_ws = wb["âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸"]

            # ĞÑ‡Ğ¸Ñ‰Ğ°ĞµĞ¼ ÑÑ‚Ğ°Ñ€Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹ (ÑÑ‚Ñ€Ğ¾ĞºĞ¸ 9 Ğ¸ Ğ´Ğ°Ğ»ĞµĞµ)
            for row in range(9, 100):  # ĞÑ‡Ğ¸Ñ‰Ğ°ĞµĞ¼ Ğ´Ğ¾ 100 ÑÑ‚Ñ€Ğ¾ĞºĞ¸
                settings_ws[f'A{row}'] = None
                settings_ws[f'B{row}'] = None

            # Ğ—Ğ°Ğ¿Ğ¸ÑÑ‹Ğ²Ğ°ĞµĞ¼ Ğ½Ğ¾Ğ²Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            for idx, cat_data in enumerate(self.categories_data, start=9):
                settings_ws[f'A{idx}'] = cat_data["name"]
                settings_ws[f'B{idx}'] = cat_data["delivery"]
                settings_ws[f'A{idx}'].border = thin_border
                settings_ws[f'B{idx}'].border = thin_border

            wb.save(self.file_path)

            # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ñƒ Ğ² GUI
            self.refresh_categories_table()
            self._save_settings()

            messagebox.showinfo(
                "Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ğ¾!",
                f"âœ… Ğ˜Ğ·Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ñ‹!\n\n"
                f"ğŸ“‹ ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹: {len(self.categories_data)}\n\n"
                f"Ğ˜Ğ·Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ Ğ±ÑƒĞ´ÑƒÑ‚ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ñ‹ Ğ¿Ñ€Ğ¸ ÑĞ»ĞµĞ´ÑƒÑÑ‰ĞµĞ¼ Ğ¿ĞµÑ€ĞµÑÑ‡ĞµÑ‚Ğµ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»."
            )

        except Exception as e:
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                f"ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ğ¸Ñ Ğ¸Ğ·Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ğ¹:\n{e}"
            )

    def update_currency_rate(self):
        """ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµÑ‚ Ñ€Ñ‹Ğ½Ğ¾Ñ‡Ğ½Ñ‹Ğ¹ ĞºÑƒÑ€Ñ EUR/RUB Ğ¸Ğ· Yahoo Finance."""
        try:
            self.currency_rate_label.config(text="ĞšÑƒÑ€Ñ EUR/RUB: Ğ·Ğ°Ğ³Ñ€ÑƒĞ·ĞºĞ°...")

            # ĞŸĞ¾Ğ»ÑƒÑ‡Ğ°ĞµĞ¼ ĞºÑƒÑ€Ñ Ğ¸Ğ· Yahoo Finance (Ğ±Ğ¸Ñ€Ğ¶ĞµĞ²Ğ¾Ğ¹ ĞºÑƒÑ€Ñ, Ğ±Ğ»Ğ¸Ğ·ĞºĞ¸Ğ¹ Ğº Google Finance)
            import yfinance as yf

            # Ğ¢Ğ¸ĞºĞµÑ€ Ğ´Ğ»Ñ Ğ¿Ğ°Ñ€Ñ‹ EUR/RUB
            ticker = yf.Ticker("EURRUB=X")

            # ĞŸĞ¾Ğ»ÑƒÑ‡Ğ°ĞµĞ¼ Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½Ğ¸Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ Ğ·Ğ° 1 Ğ´ĞµĞ½ÑŒ
            data = ticker.history(period="1d")

            if not data.empty:
                # Ğ‘ĞµÑ€ĞµĞ¼ Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½ÑÑ Ñ†ĞµĞ½Ñƒ Ğ·Ğ°ĞºÑ€Ñ‹Ñ‚Ğ¸Ñ
                self.current_eur_rub = float(data['Close'].iloc[-1])
            else:
                raise Exception("ĞĞµÑ‚ Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ… Ğ¾Ñ‚ Yahoo Finance")

            if self.current_eur_rub > 0:
                # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ Ğ¸Ğ½Ñ‚ĞµÑ€Ñ„ĞµĞ¹Ñ
                self.currency_rate_label.config(
                    text=f"ĞšÑƒÑ€Ñ EUR/RUB: {self.current_eur_rub:.4f} â‚½ (Ñ€Ñ‹Ğ½Ğ¾Ñ‡Ğ½Ñ‹Ğ¹)",
                    fg="#2196F3"
                )

                from datetime import datetime
                self.last_update_label.config(
                    text=f"ĞŸĞ¾ÑĞ»ĞµĞ´Ğ½ĞµĞµ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Yahoo Finance"
                )

                messagebox.showinfo(
                    "ĞšÑƒÑ€Ñ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½",
                    f"âœ… Ğ Ñ‹Ğ½Ğ¾Ñ‡Ğ½Ñ‹Ğ¹ ĞºÑƒÑ€Ñ EUR/RUB Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½!\n\n"
                    f"ğŸ’± {self.current_eur_rub:.4f} â‚½\n"
                    f"ğŸ• {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
                    f"ğŸ“Š Ğ˜ÑÑ‚Ğ¾Ñ‡Ğ½Ğ¸Ğº: Yahoo Finance (Ğ±Ğ¸Ñ€Ğ¶ĞµĞ²Ğ¾Ğ¹ ĞºÑƒÑ€Ñ)"
                )
            else:
                raise Exception("ĞĞµ ÑƒĞ´Ğ°Ğ»Ğ¾ÑÑŒ Ğ¿Ğ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ EUR")

        except Exception as e:
            self.currency_rate_label.config(
                text=f"ĞÑˆĞ¸Ğ±ĞºĞ° Ğ·Ğ°Ğ³Ñ€ÑƒĞ·ĞºĞ¸ ĞºÑƒÑ€ÑĞ°: {str(e)[:50]}",
                fg="#f44336"
            )
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                f"ĞĞµ ÑƒĞ´Ğ°Ğ»Ğ¾ÑÑŒ Ğ¿Ğ¾Ğ»ÑƒÑ‡Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ EUR/RUB:\n{e}\n\nĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑŒÑ‚Ğµ Ğ¿Ğ¾Ğ´ĞºĞ»ÑÑ‡ĞµĞ½Ğ¸Ğµ Ğº Ğ¸Ğ½Ñ‚ĞµÑ€Ğ½ĞµÑ‚Ñƒ."
            )

    def apply_currency_to_prices(self):
        """ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµÑ‚ ĞºÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹ Ğº Ñ†ĞµĞ½Ğ°Ğ¼ Ğ² Excel."""
        if self.current_eur_rub <= 0:
            messagebox.showwarning(
                "ĞšÑƒÑ€Ñ Ğ½Ğµ Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶ĞµĞ½",
                "Ğ¡Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ¸Ñ‚Ğµ ĞºÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹!\n\nĞĞ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ 'ğŸ”„ ĞĞ±Ğ½Ğ¾Ğ²Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ'"
            )
            return

        if not self.file_path.exists():
            messagebox.showwarning(
                "Ğ¤Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½",
                f"Excel Ñ„Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½!"
            )
            return

        try:
            # ĞŸĞ¾Ğ»ÑƒÑ‡Ğ°ĞµĞ¼ Ğ½Ğ°Ğ´Ğ±Ğ°Ğ²ĞºÑƒ
            markup = float(self.markup_entry.get())
            final_rate = self.current_eur_rub + markup
            use_peti = self.use_peti_coef.get()

            # Ğ¤Ğ¾Ñ€Ğ¼Ğ¸Ñ€ÑƒĞµĞ¼ Ñ‚ĞµĞºÑÑ‚ Ğ¿Ğ¾Ğ´Ñ‚Ğ²ĞµÑ€Ğ¶Ğ´ĞµĞ½Ğ¸Ñ
            formula_text = f"Ğ¦ĞµĞ½Ğ°â‚½ = Ğ¦ĞµĞ½Ğ°â‚¬ Ã— {final_rate:.2f}"
            if use_peti:
                formula_text += " Ã— (1 + ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸)"

            # ĞŸĞ¾Ğ´Ñ‚Ğ²ĞµÑ€Ğ¶Ğ´ĞµĞ½Ğ¸Ğµ
            result = messagebox.askyesno(
                "ĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ?",
                f"ĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ Ğº Ñ†ĞµĞ½Ğ°Ğ¼ Ğ² Excel?\n\n"
                f"ğŸ’± ĞšÑƒÑ€Ñ: {self.current_eur_rub:.2f} â‚½\n"
                f"â• ĞĞ°Ğ´Ğ±Ğ°Ğ²ĞºĞ°: {markup} â‚½\n"
                f"ğŸ“Š ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸: {'Ğ’ĞºĞ»ÑÑ‡ĞµĞ½' if use_peti else 'Ğ’Ñ‹ĞºĞ»ÑÑ‡ĞµĞ½'}\n"
                f"â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•\n"
                f"ğŸ“ Ğ¤Ğ¾Ñ€Ğ¼ÑƒĞ»Ğ°: {formula_text}\n\n"
                + ("ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸: <15Ğšâ†’10%, â‰¤30Ğšâ†’9%, >30Ğšâ†’8%\n\n" if use_peti else "")
                + f"Ğ’ÑĞµ Ñ†ĞµĞ½Ñ‹ Ğ±ÑƒĞ´ÑƒÑ‚ Ğ¿ĞµÑ€ĞµÑÑ‡Ğ¸Ñ‚Ğ°Ğ½Ñ‹."
            )

            if not result:
                return

            # Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµĞ¼ Excel
            wb = load_workbook(self.file_path)
            ws = wb.active

            updated_count = 0
            total_peti_markup = 0  # Ğ”Ğ»Ñ ÑÑ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ¸

            # ĞĞ±Ñ…Ğ¾Ğ´Ğ¸Ğ¼ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ (Ğ½Ğ°Ñ‡Ğ¸Ğ½Ğ°Ñ ÑĞ¾ 2-Ğ¹)
            for row_num in range(2, ws.max_row + 1):
                price_eur = ws.cell(row_num, 3).value  # C: Ğ¦ĞµĞ½Ğ° Ğ² â‚¬

                if price_eur and isinstance(price_eur, (int, float)) and price_eur > 0:
                    # Ğ‘Ğ°Ğ·Ğ¾Ğ²Ğ°Ñ Ñ†ĞµĞ½Ğ° Ğ² Ñ€ÑƒĞ±Ğ»ÑÑ…
                    price_rub_base = price_eur * final_rate

                    # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸, ĞµÑĞ»Ğ¸ Ğ²ĞºĞ»ÑÑ‡ĞµĞ½
                    if use_peti:
                        # Ğ Ğ°ÑÑÑ‡Ğ¸Ñ‚Ñ‹Ğ²Ğ°ĞµĞ¼ ĞºĞ¾ÑÑ„Ñ„Ğ¸Ñ†Ğ¸ĞµĞ½Ñ‚ Ğ² Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ¾Ñ‚ Ñ†ĞµĞ½Ñ‹
                        if price_rub_base < 15000:
                            peti_coef = 0.10  # 10%
                        elif price_rub_base <= 30000:
                            peti_coef = 0.09  # 9%
                        else:
                            peti_coef = 0.08  # 8%

                        # Ğ˜Ñ‚Ğ¾Ğ³Ğ¾Ğ²Ğ°Ñ Ñ†ĞµĞ½Ğ° Ñ Ğ½Ğ°Ñ†ĞµĞ½ĞºĞ¾Ğ¹
                        price_rub_final = price_rub_base * (1 + peti_coef)
                        total_peti_markup += (price_rub_final - price_rub_base)
                    else:
                        price_rub_final = price_rub_base

                    # Ğ—Ğ°Ğ¿Ğ¸ÑÑ‹Ğ²Ğ°ĞµĞ¼ Ğ¸Ñ‚Ğ¾Ğ³Ğ¾Ğ²ÑƒÑ Ñ†ĞµĞ½Ñƒ
                    ws.cell(row_num, 3).value = round(price_rub_final, 2)
                    updated_count += 1

            # Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµĞ¼
            wb.save(self.file_path)

            # Ğ¤Ğ¾Ñ€Ğ¼Ğ¸Ñ€ÑƒĞµĞ¼ ÑĞ¾Ğ¾Ğ±Ñ‰ĞµĞ½Ğ¸Ğµ Ğ¾Ğ± ÑƒÑĞ¿ĞµÑ…Ğµ
            success_message = (
                f"âœ… Ğ¦ĞµĞ½Ñ‹ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ñ‹!\n\n"
                f"ğŸ“Š ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¾ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²: {updated_count}\n"
                f"ğŸ’± ĞšÑƒÑ€Ñ: {final_rate:.2f} â‚½\n"
            )

            if use_peti and updated_count > 0:
                avg_peti_markup = total_peti_markup / updated_count
                success_message += (
                    f"ğŸ“ˆ ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸: Ğ’ĞºĞ»ÑÑ‡ĞµĞ½\n"
                    f"ğŸ’° Ğ¡Ñ€ĞµĞ´Ğ½ÑÑ Ğ½Ğ°Ñ†ĞµĞ½ĞºĞ°: {avg_peti_markup:.2f} â‚½\n"
                )

            success_message += "\nExcel Ñ„Ğ°Ğ¹Ğ» ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½."

            messagebox.showinfo("Ğ“Ğ¾Ñ‚Ğ¾Ğ²Ğ¾!", success_message)

        except ValueError:
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                "ĞĞµĞ²ĞµÑ€Ğ½Ñ‹Ğ¹ Ñ„Ğ¾Ñ€Ğ¼Ğ°Ñ‚ Ğ½Ğ°Ğ´Ğ±Ğ°Ğ²ĞºĞ¸!\n\nĞ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾ (Ğ½Ğ°Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ€: 0.5)"
            )
        except Exception as e:
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                f"ĞÑˆĞ¸Ğ±ĞºĞ° Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ ĞºÑƒÑ€ÑĞ°:\n{e}"
            )

    def apply_formulas_silently(self):
        """ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµÑ‚ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ñ‚Ğ¸Ñ…Ğ¾ (Ğ±ĞµĞ· Ğ´Ğ¸Ğ°Ğ»Ğ¾Ğ³Ğ¾Ğ²Ñ‹Ñ… Ğ¾ĞºĞ¾Ğ½), Ğ´Ğ»Ñ Ğ°Ğ²Ñ‚Ğ¾Ğ¼Ğ°Ñ‚Ğ¸Ñ‡ĞµÑĞºĞ¾Ğ³Ğ¾ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ Ğ¿Ğ¾ÑĞ»Ğµ Ğ¿Ğ°Ñ€ÑĞ¸Ğ½Ğ³Ğ°."""
        if self.current_eur_rub <= 0:
            raise Exception("ĞšÑƒÑ€Ñ EUR/RUB Ğ½Ğµ Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶ĞµĞ½")

        if not self.file_path.exists():
            raise Exception(f"Excel Ñ„Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½")

        # ĞŸĞ¾Ğ»ÑƒÑ‡Ğ°ĞµĞ¼ Ñ‚ĞµĞºÑƒÑ‰Ğ¸Ğµ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸
        markup = float(self.markup_entry.get())
        final_rate = self.current_eur_rub + markup

        # Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµĞ¼ Excel
        wb = load_workbook(self.file_path)
        ws = wb.active

        # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ Ğ»Ğ¸ÑÑ‚ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº Ñ Ğ°ĞºÑ‚ÑƒĞ°Ğ»ÑŒĞ½Ñ‹Ğ¼ ĞºÑƒÑ€ÑĞ¾Ğ¼
        if "âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸" in wb.sheetnames:
            settings_ws = wb["âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸"]
            settings_ws['B3'] = self.current_eur_rub
            settings_ws['B4'] = markup
        else:
            # Ğ•ÑĞ»Ğ¸ Ğ»Ğ¸ÑÑ‚Ğ° Ğ½ĞµÑ‚, ÑĞ¾Ğ·Ğ´Ğ°ĞµĞ¼ ĞµĞ³Ğ¾
            self._create_settings_sheet(wb, self.current_eur_rub, markup)

        # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸ Ğ½Ğ¾Ğ²Ñ‹Ñ… ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ² (ĞµÑĞ»Ğ¸ Ğ¸Ñ… ĞµÑ‰Ğµ Ğ½ĞµÑ‚)
        new_headers = [
            "Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚½)",      # L
            "Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° (â‚½)",       # M
            "ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (%)",      # N
            "ĞĞ°Ñˆ ĞšÑÑ„ (%)",       # O
            "Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚. (â‚½)", # P
            "Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚. (â‚½)", # Q
            "ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ° (â‚½)",    # R
            "ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸ (â‚½)"     # S
        ]

        # Ğ¡Ñ‚Ğ¸Ğ»Ğ¸ Ğ´Ğ»Ñ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ¸Ñ
        orange_header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        green_value_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        value_font = Font(size=11, name="Calibri")
        value_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for col_idx, header in enumerate(new_headers, start=12):  # ĞĞ°Ñ‡Ğ¸Ğ½Ğ°ĞµĞ¼ Ñ L (12)
            if not ws.cell(1, col_idx).value:
                ws.cell(1, col_idx).value = header

            # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğ¾Ñ€Ğ°Ğ½Ğ¶ĞµĞ²Ñ‹Ğ¹ ÑÑ‚Ğ¸Ğ»ÑŒ Ğº Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºÑƒ
            header_cell = ws.cell(1, col_idx)
            header_cell.fill = orange_header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment
            header_cell.border = thin_border

            # Ğ£ÑÑ‚Ğ°Ğ½Ğ°Ğ²Ğ»Ğ¸Ğ²Ğ°ĞµĞ¼ ÑˆĞ¸Ñ€Ğ¸Ğ½Ñƒ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ²
            col_letter = header_cell.column_letter
            ws.column_dimensions[col_letter].width = 18

        # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº "Ğ‘Ñ€ĞµĞ½Ğ´" Ğ² ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğµ T (20)
        if not ws.cell(1, 20).value:
            ws.cell(1, 20).value = "Ğ‘Ñ€ĞµĞ½Ğ´"
        brand_cell = ws.cell(1, 20)
        brand_cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        brand_cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        brand_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        brand_cell.border = thin_border
        ws.column_dimensions['T'].width = 18

        processed_count = 0

        # ĞĞ±Ñ…Ğ¾Ğ´Ğ¸Ğ¼ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°Ğ¼Ğ¸ (Ğ½Ğ°Ñ‡Ğ¸Ğ½Ğ°Ñ ÑĞ¾ 2-Ğ¹)
        for row_num in range(2, ws.max_row + 1):
            price_eur = ws.cell(row_num, 3).value  # C: Ğ¦ĞµĞ½Ğ° (â‚¬)

            # ĞŸÑ€Ğ¾Ğ¿ÑƒÑĞºĞ°ĞµĞ¼ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ Ğ±ĞµĞ· Ñ†ĞµĞ½Ñ‹
            if not price_eur:
                continue

            # === Ğ’Ğ¡Ğ¢ĞĞ’Ğ›Ğ¯Ğ•Ğœ Ğ¤ĞĞ ĞœĞ£Ğ›Ğ« Ğ’ĞœĞ•Ğ¡Ğ¢Ğ Ğ—ĞĞĞ§Ğ•ĞĞ˜Ğ™ ===

            # L: Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° - VLOOKUP Ğ¿Ğ¾ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸ Ğ¸Ğ· Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
            formula_delivery = f"=IFERROR(VLOOKUP(F{row_num},'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$A$9:$B$17,2,FALSE)*'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$B$5,0)"

            # M: Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° = Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° + (Ğ¦ĞµĞ½Ğ°_EUR * ĞšÑƒÑ€Ñ)
            formula_zakupka = f"=L{row_num}+(C{row_num}*'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$B$5)"

            # N: ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (10%, 9%, 8% Ğ² Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ¾Ñ‚ Ğ·Ğ°ĞºÑƒĞ¿ĞºĞ¸)
            formula_peti_coef = f"=IF(M{row_num}<15000,10%,IF(M{row_num}<=30000,9%,8%))"

            # O: ĞĞ°Ñˆ ĞšÑÑ„ (17%, 15%, 14%, 13% Ğ² Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ¾Ñ‚ Ğ·Ğ°ĞºÑƒĞ¿ĞºĞ¸)
            formula_nash_coef = f"=IF(M{row_num}<10000,17%,IF(M{row_num}<=20000,15%,IF(M{row_num}<=30000,14%,13%)))"

            # P: Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹ = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° * (1 + ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸ + ĞĞ°Ñˆ_ĞšÑÑ„)
            formula_price_with_delivery = f"=M{row_num}*(1+N{row_num}+O{row_num})"

            # Q: Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ = Ğ¦ĞµĞ½Ğ°_Ñ_Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹ - Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ°
            formula_price_without_delivery = f"=P{row_num}-L{row_num}"

            # R: ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ° = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° * ĞĞ°Ñˆ_ĞšÑÑ„
            formula_margin_nash = f"=M{row_num}*O{row_num}"

            # S: ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸ = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° * ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸
            formula_margin_peti = f"=M{row_num}*N{row_num}"

            # Ğ’ÑÑ‚Ğ°Ğ²Ğ»ÑĞµĞ¼ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ¸ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğ·ĞµĞ»ĞµĞ½Ğ¾Ğµ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ¸Ğµ
            formulas = [
                (12, formula_delivery),              # L: Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ°
                (13, formula_zakupka),               # M: Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°
                (14, formula_peti_coef),             # N: ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸
                (15, formula_nash_coef),             # O: ĞĞ°Ñˆ ĞšÑÑ„
                (16, formula_price_with_delivery),   # P: Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹
                (17, formula_price_without_delivery),# Q: Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
                (18, formula_margin_nash),           # R: ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ°
                (19, formula_margin_peti)            # S: ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸
            ]

            for col_idx, formula in formulas:
                cell = ws.cell(row_num, col_idx)
                cell.value = formula  # Ğ’ÑÑ‚Ğ°Ğ²Ğ»ÑĞµĞ¼ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñƒ
                cell.fill = green_value_fill
                cell.font = value_font
                cell.alignment = value_alignment
                cell.border = thin_border
                # Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚ Ñ‡Ğ¸ÑĞ»Ğ° Ğ´Ğ»Ñ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ² Ñ Ğ¿Ñ€Ğ¾Ñ†ĞµĞ½Ñ‚Ğ°Ğ¼Ğ¸ (N, O)
                if col_idx in [14, 15]:
                    cell.number_format = '0%'
                else:
                    cell.number_format = '#,##0.00'

            processed_count += 1

        # Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµĞ¼
        wb.save(self.file_path)
        return processed_count

    def _create_settings_sheet(self, wb, eur_rate, markup):
        """Ğ¡Ğ¾Ğ·Ğ´Ğ°ĞµÑ‚ Ğ»Ğ¸ÑÑ‚ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº Ñ ĞºÑƒÑ€ÑĞ¾Ğ¼ Ğ¸ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†ĞµĞ¹ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸."""
        settings_ws = wb.create_sheet("âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸")

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
        settings_ws['A1'] = "âš™ï¸ ĞĞĞ¡Ğ¢Ğ ĞĞ™ĞšĞ˜ Ğ ĞĞ¡Ğ§Ğ•Ğ¢ĞĞ’"
        settings_ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        settings_ws['A1'].font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
        settings_ws.merge_cells('A1:C1')

        # ĞšÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹
        settings_ws['A3'] = "ĞšÑƒÑ€Ñ EUR/RUB:"
        settings_ws['B3'] = eur_rate
        settings_ws['A3'].font = Font(bold=True, size=12)
        settings_ws['B3'].font = Font(size=12)
        settings_ws['B3'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        settings_ws['A4'] = "ĞĞ°Ğ´Ğ±Ğ°Ğ²ĞºĞ°:"
        settings_ws['B4'] = markup
        settings_ws['A4'].font = Font(bold=True, size=12)
        settings_ws['B4'].font = Font(size=12)
        settings_ws['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        settings_ws['A5'] = "Ğ˜Ñ‚Ğ¾Ğ³Ğ¾Ğ²Ñ‹Ğ¹ ĞºÑƒÑ€Ñ:"
        settings_ws['B5'] = "=B3+B4"
        settings_ws['A5'].font = Font(bold=True, size=12)
        settings_ws['B5'].font = Font(bold=True, size=14)
        settings_ws['B5'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
        settings_ws['A7'] = "ğŸ“¦ Ğ¡Ğ¢ĞĞ˜ĞœĞĞ¡Ğ¢Ğ¬ Ğ”ĞĞ¡Ğ¢ĞĞ’ĞšĞ˜ (â‚¬)"
        settings_ws['A7'].font = Font(bold=True, color="FFFFFF", size=14)
        settings_ws['A7'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        settings_ws.merge_cells('A7:B7')

        settings_ws['A8'] = "ĞšĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ñ"
        settings_ws['B8'] = "Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚¬)"
        settings_ws['A8'].font = Font(bold=True, size=11)
        settings_ws['B8'].font = Font(bold=True, size=11)
        settings_ws['A8'].fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        settings_ws['B8'].fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")

        # Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¹ Ğ¸ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
        delivery_table = [
            ("ĞÑ‡ĞºĞ¸", 12),
            ("Ğ Ğ°ĞºĞµÑ‚ĞºĞ°", 17),
            ("ĞšÑ€Ğ¾ÑÑĞ¾Ğ²ĞºĞ¸", 28),
            ("ĞšÑƒÑ€Ñ‚ĞºĞ°", 17),
            ("Ğ¨Ñ‚Ğ°Ğ½Ñ‹", 17),
            ("Ğ¨Ğ»ĞµĞ¼", 28),
            ("Ğ‘Ğ¾Ñ‚Ğ¸Ğ½ĞºĞ¸ Ğ±Ğ¾Ñ€Ğ´", 25),
            ("Ğ¢ĞµÑ€Ğ¼Ğ¾", 17),
            ("ĞÑ‡ĞºĞ¸ Ğ´Ğ»Ñ ÑĞ½ĞµĞ³Ğ°", 17)
        ]

        for idx, (cat, delivery) in enumerate(delivery_table, start=9):
            settings_ws[f'A{idx}'] = cat
            settings_ws[f'B{idx}'] = delivery
            settings_ws[f'A{idx}'].border = thin_border
            settings_ws[f'B{idx}'].border = thin_border

        # Ğ¡ĞµĞºÑ†Ğ¸Ñ Ğ‘Ğ Ğ•ĞĞ”Ğ« (ÑÑ‚Ğ¾Ğ»Ğ±ĞµÑ† D)
        settings_ws['D1'] = "ğŸ·ï¸ Ğ‘Ğ Ğ•ĞĞ”Ğ«"
        settings_ws['D1'].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
        settings_ws['D1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        settings_ws['D2'] = "Ğ‘Ñ€ĞµĞ½Ğ´"
        settings_ws['D2'].font = Font(bold=True, size=11)
        settings_ws['D2'].fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")

        for idx, brand in enumerate(self.brands_data, start=3):
            settings_ws[f'D{idx}'] = brand
            settings_ws[f'D{idx}'].border = thin_border

        # Ğ¨Ğ¸Ñ€Ğ¸Ğ½Ğ° ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ² Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
        settings_ws.column_dimensions['A'].width = 25
        settings_ws.column_dimensions['B'].width = 20
        settings_ws.column_dimensions['C'].width = 15
        settings_ws.column_dimensions['D'].width = 20

    def apply_formulas_to_excel(self):
        """ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµÑ‚ Ğ²ÑĞµ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ñ€Ğ°ÑÑ‡ĞµÑ‚Ğ° Ğº Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°Ğ¼ Ğ² Excel."""
        if self.current_eur_rub <= 0:
            messagebox.showwarning(
                "ĞšÑƒÑ€Ñ Ğ½Ğµ Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶ĞµĞ½",
                "Ğ¡Ğ½Ğ°Ñ‡Ğ°Ğ»Ğ° Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ¸Ñ‚Ğµ ĞºÑƒÑ€Ñ Ğ²Ğ°Ğ»ÑÑ‚Ñ‹!\n\nĞĞ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ 'ğŸ”„ ĞĞ±Ğ½Ğ¾Ğ²Ğ¸Ñ‚ÑŒ ĞºÑƒÑ€Ñ'"
            )
            return

        if not self.file_path.exists():
            messagebox.showwarning(
                "Ğ¤Ğ°Ğ¹Ğ» Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½",
                f"Excel Ñ„Ğ°Ğ¹Ğ» {self.file_path.name} Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½!"
            )
            return

        try:
            # ĞŸĞ¾Ğ»ÑƒÑ‡Ğ°ĞµĞ¼ Ñ‚ĞµĞºÑƒÑ‰Ğ¸Ğµ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸
            markup = float(self.markup_entry.get())
            final_rate = self.current_eur_rub + markup

            # ĞŸĞ¾Ğ´Ñ‚Ğ²ĞµÑ€Ğ¶Ğ´ĞµĞ½Ğ¸Ğµ
            result = messagebox.askyesno(
                "ĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹?",
                f"ĞŸÑ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ Ğ²ÑĞµ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ñ€Ğ°ÑÑ‡ĞµÑ‚Ğ° Ğº Excel?\n\n"
                f"ğŸ’± ĞšÑƒÑ€Ñ: {final_rate:.2f} â‚½\n\n"
                f"Ğ‘ÑƒĞ´ÑƒÑ‚ Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ñ‹ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ñ‹:\n"
                f"â€¢ L: Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚½)\n"
                f"â€¢ M: Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° (â‚½)\n"
                f"â€¢ N: ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (%)\n"
                f"â€¢ O: ĞĞ°Ñˆ ĞšÑÑ„ (%)\n"
                f"â€¢ P: Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹ (â‚½)\n"
                f"â€¢ Q: Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ (â‚½)\n"
                f"â€¢ R: ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ° (â‚½)\n"
                f"â€¢ S: ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸ (â‚½)"
            )

            if not result:
                return

            # Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµĞ¼ Excel
            wb = load_workbook(self.file_path)
            ws = wb.active

            # ĞĞ±Ğ½Ğ¾Ğ²Ğ»ÑĞµĞ¼ Ğ»Ğ¸ÑÑ‚ Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº Ñ Ğ°ĞºÑ‚ÑƒĞ°Ğ»ÑŒĞ½Ñ‹Ğ¼ ĞºÑƒÑ€ÑĞ¾Ğ¼
            if "âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸" in wb.sheetnames:
                settings_ws = wb["âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸"]
                settings_ws['B3'] = self.current_eur_rub
                settings_ws['B4'] = markup
            else:
                # Ğ•ÑĞ»Ğ¸ Ğ»Ğ¸ÑÑ‚Ğ° Ğ½ĞµÑ‚, ÑĞ¾Ğ·Ğ´Ğ°ĞµĞ¼ ĞµĞ³Ğ¾
                self._create_settings_sheet(wb, self.current_eur_rub, markup)

            # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸ Ğ½Ğ¾Ğ²Ñ‹Ñ… ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ² (ĞµÑĞ»Ğ¸ Ğ¸Ñ… ĞµÑ‰Ğµ Ğ½ĞµÑ‚)
            new_headers = [
                "Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° (â‚½)",      # L
                "Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° (â‚½)",       # M
                "ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (%)",      # N
                "ĞĞ°Ñˆ ĞšÑÑ„ (%)",       # O
                "Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚. (â‚½)", # P
                "Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚. (â‚½)", # Q
                "ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ° (â‚½)",    # R
                "ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸ (â‚½)"     # S
            ]

            # Ğ¡Ñ‚Ğ¸Ğ»Ğ¸ Ğ´Ğ»Ñ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ¸Ñ
            orange_header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            green_value_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            value_font = Font(size=11, name="Calibri")
            value_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            for col_idx, header in enumerate(new_headers, start=12):  # ĞĞ°Ñ‡Ğ¸Ğ½Ğ°ĞµĞ¼ Ñ L (12)
                if not ws.cell(1, col_idx).value:
                    ws.cell(1, col_idx).value = header

                # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğ¾Ñ€Ğ°Ğ½Ğ¶ĞµĞ²Ñ‹Ğ¹ ÑÑ‚Ğ¸Ğ»ÑŒ Ğº Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºÑƒ
                header_cell = ws.cell(1, col_idx)
                header_cell.fill = orange_header_fill
                header_cell.font = header_font
                header_cell.alignment = header_alignment
                header_cell.border = thin_border

                # Ğ£ÑÑ‚Ğ°Ğ½Ğ°Ğ²Ğ»Ğ¸Ğ²Ğ°ĞµĞ¼ ÑˆĞ¸Ñ€Ğ¸Ğ½Ñƒ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ²
                col_letter = header_cell.column_letter
                ws.column_dimensions[col_letter].width = 18

            # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²Ğ¾Ğº "Ğ‘Ñ€ĞµĞ½Ğ´" Ğ² ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğµ T (20)
            if not ws.cell(1, 20).value:
                ws.cell(1, 20).value = "Ğ‘Ñ€ĞµĞ½Ğ´"
            brand_cell = ws.cell(1, 20)
            brand_cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
            brand_cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
            brand_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            brand_cell.border = thin_border
            ws.column_dimensions['T'].width = 18

            processed_count = 0
            skipped_count = 0

            # ĞĞ±Ñ…Ğ¾Ğ´Ğ¸Ğ¼ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ Ñ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ°Ğ¼Ğ¸ (Ğ½Ğ°Ñ‡Ğ¸Ğ½Ğ°Ñ ÑĞ¾ 2-Ğ¹)
            for row_num in range(2, ws.max_row + 1):
                price_eur = ws.cell(row_num, 3).value  # C: Ğ¦ĞµĞ½Ğ° (â‚¬)

                # ĞŸÑ€Ğ¾Ğ¿ÑƒÑĞºĞ°ĞµĞ¼ ÑÑ‚Ñ€Ğ¾ĞºĞ¸ Ğ±ĞµĞ· Ñ†ĞµĞ½Ñ‹
                if not price_eur:
                    skipped_count += 1
                    continue

                # === Ğ’Ğ¡Ğ¢ĞĞ’Ğ›Ğ¯Ğ•Ğœ Ğ¤ĞĞ ĞœĞ£Ğ›Ğ« Ğ’ĞœĞ•Ğ¡Ğ¢Ğ Ğ—ĞĞĞ§Ğ•ĞĞ˜Ğ™ ===

                # L: Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° - VLOOKUP Ğ¿Ğ¾ ĞºĞ°Ñ‚ĞµĞ³Ğ¾Ñ€Ğ¸Ğ¸ Ğ¸Ğ· Ğ½Ğ°ÑÑ‚Ñ€Ğ¾ĞµĞº
                formula_delivery = f"=IFERROR(VLOOKUP(F{row_num},'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$A$9:$B$17,2,FALSE)*'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$B$5,0)"

                # M: Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° = Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ° + (Ğ¦ĞµĞ½Ğ°_EUR * ĞšÑƒÑ€Ñ)
                formula_zakupka = f"=L{row_num}+(C{row_num}*'âš™ï¸ ĞĞ°ÑÑ‚Ñ€Ğ¾Ğ¹ĞºĞ¸'!$B$5)"

                # N: ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸ (10%, 9%, 8% Ğ² Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ¾Ñ‚ Ğ·Ğ°ĞºÑƒĞ¿ĞºĞ¸)
                formula_peti_coef = f"=IF(M{row_num}<15000,10%,IF(M{row_num}<=30000,9%,8%))"

                # O: ĞĞ°Ñˆ ĞšÑÑ„ (17%, 15%, 14%, 13% Ğ² Ğ·Ğ°Ğ²Ğ¸ÑĞ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ¾Ñ‚ Ğ·Ğ°ĞºÑƒĞ¿ĞºĞ¸)
                formula_nash_coef = f"=IF(M{row_num}<10000,17%,IF(M{row_num}<=20000,15%,IF(M{row_num}<=30000,14%,13%)))"

                # P: Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹ = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° * (1 + ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸ + ĞĞ°Ñˆ_ĞšÑÑ„)
                formula_price_with_delivery = f"=M{row_num}*(1+N{row_num}+O{row_num})"

                # Q: Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸ = Ğ¦ĞµĞ½Ğ°_Ñ_Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹ - Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ°
                formula_price_without_delivery = f"=P{row_num}-L{row_num}"

                # R: ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ° = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° * ĞĞ°Ñˆ_ĞšÑÑ„
                formula_margin_nash = f"=M{row_num}*O{row_num}"

                # S: ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸ = Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° * ĞšÑÑ„_ĞŸĞµÑ‚Ğ¸
                formula_margin_peti = f"=M{row_num}*N{row_num}"

                # Ğ’ÑÑ‚Ğ°Ğ²Ğ»ÑĞµĞ¼ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ¸ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ğ·ĞµĞ»ĞµĞ½Ğ¾Ğµ Ğ¾Ñ„Ğ¾Ñ€Ğ¼Ğ»ĞµĞ½Ğ¸Ğµ
                formulas = [
                    (12, formula_delivery),              # L: Ğ”Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ°
                    (13, formula_zakupka),               # M: Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ°
                    (14, formula_peti_coef),             # N: ĞšÑÑ„ ĞŸĞµÑ‚Ğ¸
                    (15, formula_nash_coef),             # O: ĞĞ°Ñˆ ĞšÑÑ„
                    (16, formula_price_with_delivery),   # P: Ğ¦ĞµĞ½Ğ° Ñ Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¾Ğ¹
                    (17, formula_price_without_delivery),# Q: Ğ¦ĞµĞ½Ğ° Ğ±ĞµĞ· Ğ´Ğ¾ÑÑ‚Ğ°Ğ²ĞºĞ¸
                    (18, formula_margin_nash),           # R: ĞĞ°ÑˆĞ° ĞœĞ°Ñ€Ğ¶Ğ°
                    (19, formula_margin_peti)            # S: ĞœĞ°Ñ€Ğ¶Ğ° ĞŸĞµÑ‚Ğ¸
                ]

                for col_idx, formula in formulas:
                    cell = ws.cell(row_num, col_idx)
                    cell.value = formula  # Ğ’ÑÑ‚Ğ°Ğ²Ğ»ÑĞµĞ¼ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»Ñƒ
                    cell.fill = green_value_fill
                    cell.font = value_font
                    cell.alignment = value_alignment
                    cell.border = thin_border
                    # Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚ Ñ‡Ğ¸ÑĞ»Ğ° Ğ´Ğ»Ñ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ğ¾Ğ² Ñ Ğ¿Ñ€Ğ¾Ñ†ĞµĞ½Ñ‚Ğ°Ğ¼Ğ¸ (N, O)
                    if col_idx in [14, 15]:
                        cell.number_format = '0%'
                    else:
                        cell.number_format = '#,##0.00'

                processed_count += 1

            # Ğ¡Ğ¾Ñ…Ñ€Ğ°Ğ½ÑĞµĞ¼
            wb.save(self.file_path)

            messagebox.showinfo(
                "Ğ“Ğ¾Ñ‚Ğ¾Ğ²Ğ¾!",
                f"âœ… Ğ¤Ğ¾Ñ€Ğ¼ÑƒĞ»Ñ‹ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ñ‹!\n\n"
                f"ğŸ“Š ĞĞ±Ñ€Ğ°Ğ±Ğ¾Ñ‚Ğ°Ğ½Ğ¾ Ñ‚Ğ¾Ğ²Ğ°Ñ€Ğ¾Ğ²: {processed_count}\n"
                f"â­ï¸ ĞŸÑ€Ğ¾Ğ¿ÑƒÑ‰ĞµĞ½Ğ¾ (Ğ½ĞµÑ‚ Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ…): {skipped_count}\n\n"
                f"Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ñ‹ ÑÑ‚Ğ¾Ğ»Ğ±Ñ†Ñ‹ Ñ Ñ€Ğ°ÑÑ‡ĞµÑ‚Ğ°Ğ¼Ğ¸ (L-S)\n"
                f"Excel Ñ„Ğ°Ğ¹Ğ» ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½."
            )

        except ValueError:
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                "ĞĞµĞ²ĞµÑ€Ğ½Ñ‹Ğ¹ Ñ„Ğ¾Ñ€Ğ¼Ğ°Ñ‚ Ğ½Ğ°Ğ´Ğ±Ğ°Ğ²ĞºĞ¸!\n\nĞ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾ (Ğ½Ğ°Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ€: 0.5)"
            )
        except Exception as e:
            messagebox.showerror(
                "ĞÑˆĞ¸Ğ±ĞºĞ°",
                f"ĞÑˆĞ¸Ğ±ĞºĞ° Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½ĞµĞ½Ğ¸Ñ Ñ„Ğ¾Ñ€Ğ¼ÑƒĞ»:\n{e}"
            )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸš€ MAIN
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

if __name__ == "__main__":
    root = tk.Tk()
    app = ParserApp(root)
    root.mainloop()
