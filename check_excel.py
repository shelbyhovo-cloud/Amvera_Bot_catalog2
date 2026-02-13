import sys
import io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

file_path = Path(__file__).parent / "products_links.xlsx"

print(f"Checking: {file_path}\n")

wb = load_workbook(file_path)
ws = wb.active

print(f"{'Row':<5} {'Name':<40} {'Local Images'}")
print("-" * 100)

for row_num in range(2, min(ws.max_row + 1, 10)):  # First 8 products
    name = ws.cell(row_num, 2).value          # B: Name
    local_images = ws.cell(row_num, 8).value  # H: Local Images

    if name:
        print(f"{row_num:<5} {str(name)[:40]:<40} {str(local_images)[:60]}")

print("\nDone!")
