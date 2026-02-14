"""
Telegram Mini App - Магазин/Каталог для группы
Всё в одном файле: бот + веб-сервер + HTML интерфейс
"""

# ═══════════════════════════════════════════════════════════
# 📦 АВТОУСТАНОВКА ЗАВИСИМОСТЕЙ
# ═══════════════════════════════════════════════════════════

import subprocess
import sys
import platform
import time
import io
from pathlib import Path

# Фикс кодировки для Windows
if platform.system() == 'Windows':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def install_dependencies():
    """Автоматически устанавливает необходимые пакеты."""
    required_packages = {
        'aiogram': 'aiogram==3.13.1',
        'aiohttp': 'aiohttp==3.10.5',
        'openpyxl': 'openpyxl==3.1.2',
    }

    missing_packages = []

    for package_name, package_spec in required_packages.items():
        try:
            __import__(package_name)
        except ImportError:
            missing_packages.append(package_spec)

    if missing_packages:
        print("📦 Устанавливаю недостающие зависимости...")
        print(f"   Пакеты: {', '.join(missing_packages)}")

        try:
            subprocess.check_call([
                sys.executable,
                '-m',
                'pip',
                'install',
                *missing_packages
            ])
            print("✅ Зависимости установлены успешно!\n")
        except subprocess.CalledProcessError as e:
            print(f"❌ Ошибка установки зависимостей: {e}")
            print("   Попробуйте установить вручную:")
            print(f"   pip install {' '.join(missing_packages)}")
            sys.exit(1)

# Проверяем и устанавливаем зависимости при импорте
install_dependencies()


# ═══════════════════════════════════════════════════════════
# 🌐 АВТОМАТИЧЕСКИЙ ЗАПУСК SERVEO (ТУННЕЛИРОВАНИЕ)
# ═══════════════════════════════════════════════════════════

def start_serveo(port):
    """
    Запускает Serveo туннель как альтернативу ngrok.
    Возвращает (public_url, process) или (None, None) при ошибке.
    """
    print("🌐 Запускаю Serveo туннель...")
    print(f"   Порт: {port}")

    try:
        import re
        from threading import Thread

        print("   Подключаюсь к serveo.net через SSH...")

        # Запускаем SSH туннель с таймаутом
        serveo_process = subprocess.Popen(
            ['ssh', '-o', 'StrictHostKeyChecking=no',
             '-o', 'ConnectTimeout=10',
             '-o', 'ServerAliveInterval=30',
             '-o', 'ServerAliveCountMax=3',
             '-R', f'80:localhost:{port}', 'serveo.net'],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            stdin=subprocess.PIPE,
            text=True,
            bufsize=1,
            creationflags=subprocess.CREATE_NO_WINDOW if platform.system() == 'Windows' else 0
        )

        serveo_url = None
        print("   Жду ответ от Serveo (макс 15 сек)...")

        # Читаем вывод и ищем URL (макс 15 секунд)
        for i in range(30):  # 30 * 0.5 = 15 секунд
            # Проверяем что процесс жив
            if serveo_process.poll() is not None:
                print("   ⚠️  Serveo процесс завершился")
                # Читаем ошибку
                output = serveo_process.stdout.read() if serveo_process.stdout else ""
                if output:
                    print(f"   Вывод: {output[:200]}")
                break

            line = serveo_process.stdout.readline()
            if line:
                # Показываем отладочную информацию
                if i < 5:  # Первые несколько строк
                    print(f"   Debug: {line.strip()[:80]}")

                # Ищем URL в формате https://xxxxx.serveo.net
                match = re.search(r'https://[a-zA-Z0-9\-]+\.serveo\.net', line)
                if match:
                    serveo_url = match.group(0)
                    break
            else:
                time.sleep(0.5)

        if serveo_url:
            print(f"✅ Serveo туннель активен!")
            print(f"🌍 Публичный URL: {serveo_url}\n")

            # Продолжаем читать вывод в фоне
            def read_output():
                try:
                    for line in serveo_process.stdout:
                        pass
                except:
                    pass

            Thread(target=read_output, daemon=True).start()

            return serveo_url, serveo_process
        else:
            print("❌ Не удалось получить URL от Serveo (таймаут 15 сек)")
            print("   Возможно Serveo перегружен или недоступен")
            try:
                serveo_process.kill()
            except:
                pass
            return None, None

    except FileNotFoundError:
        print("❌ SSH клиент не найден!")
        print("   Windows 10/11: Параметры → Приложения → Дополнительные компоненты → OpenSSH Client")
        return None, None
    except Exception as e:
        print(f"❌ Ошибка Serveo: {e}")
        import traceback
        traceback.print_exc()
        return None, None


import asyncio
import json
import logging
import math
from urllib.parse import quote
from aiohttp import web

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, WebAppInfo, ReplyKeyboardMarkup, KeyboardButton


# ═══════════════════════════════════════════════════════════
# 🔧 УТИЛИТЫ
# ═══════════════════════════════════════════════════════════

def kill_process_on_port(port):
    """Убивает все процессы которые используют указанный порт"""
    try:
        # Находим процессы на порту
        result = subprocess.run(
            f'netstat -ano | findstr :{port}',
            shell=True,
            capture_output=True,
            text=True
        )

        if not result.stdout.strip():
            return True  # Порт свободен

        # Извлекаем PID процессов
        pids = set()
        for line in result.stdout.strip().split('\n'):
            parts = line.split()
            if len(parts) >= 5:
                pid = parts[-1]
                if pid.isdigit():
                    pids.add(pid)

        if not pids:
            return True  # Порт свободен

        # Убиваем каждый процесс
        for pid in pids:
            try:
                subprocess.run(
                    f'taskkill /PID {pid} /F',
                    shell=True,
                    capture_output=True,
                    check=True
                )
                print(f"❌ Остановил процесс на порту {port} (PID {pid})")
            except subprocess.CalledProcessError:
                pass  # Процесс уже завершён

        return True

    except Exception:
        return False


# ═══════════════════════════════════════════════════════════
# ⚙️  НАСТРОЙКИ
# ═══════════════════════════════════════════════════════════

BOT_TOKEN = "8529662300:AAHnb8e8Qh93INgnC_x3rkDc1QC20c3ulFM"

# Админы (могут загружать каталог)
ADMIN_USERNAMES = ["AlexeyBakaev", "musyanya", "GussionHovo"]

WEBAPP_HOST = "0.0.0.0"
WEBAPP_PORT = 8080

# Режим работы:
# - "auto" = автоматический туннель через Serveo (бесплатно, без регистрации)
# - "manual" = ручной режим, нужно указать свой URL ниже
MODE = "manual"

# Если MODE = "manual", вставь сюда свой HTTPS URL от Serveo/LocalTunnel/etc
MANUAL_WEBAPP_URL = "https://nimblicatalog-alexey20031986.amvera.io"

# WEBAPP_URL будет установлен автоматически
WEBAPP_URL = None

# ═══════════════════════════════════════════════════════════
# 📦 КАТАЛОГ ТОВАРОВ (можно редактировать)
# ═══════════════════════════════════════════════════════════

# Стандартные товары (используются если нет Excel файла)
PRODUCTS_DEFAULT = [
    {
        "id": 1,
        "name": "Футболка Premium",
        "price": 1500,
        "image": "👕",
    },
    {
        "id": 2,
        "name": "Кроссовки Sport",
        "price": 4500,
        "image": "👟",
    },
    {
        "id": 3,
        "name": "Рюкзак Urban",
        "price": 2800,
        "image": "🎒",
    },
    {
        "id": 4,
        "name": "Наушники Pro",
        "price": 6000,
        "image": "🎧",
    },
    {
        "id": 5,
        "name": "Смарт-часы",
        "price": 8500,
        "image": "⌚",
    },
    {
        "id": 6,
        "name": "Кепка Classic",
        "price": 900,
        "image": "🧢",
    },
]

PRODUCTS = []  # Будет загружено из Excel или использованы стандартные


def get_images_dir():
    """
    Определяет путь к папке images в зависимости от окружения.

    Приоритет:
    1. /data/images/ (если /data существует) - постоянное хранилище Amvera
    2. script_dir/images/ - локальная разработка (fallback)
    """
    # Проверяем /data/ на Amvera
    data_path = Path('/data')
    if data_path.exists() and data_path.is_dir():
        # На Amvera - всегда используем /data/images
        data_images_dir = data_path / 'images'
        data_images_dir.mkdir(exist_ok=True)
        return data_images_dir

    # Локальная разработка - используем папку со скриптом
    script_dir = Path(__file__).parent
    images_dir = script_dir / 'images'
    images_dir.mkdir(exist_ok=True)
    return images_dir


def load_products_from_excel(file_path=None):
    """Загружает товары из Excel файла."""
    global PRODUCTS

    # Если путь не указан, ищем сначала в /data (Amvera), потом локально
    if file_path is None:
        # Проверяем /data/products_links.xlsx (persistenceMount на Amvera)
        data_path = Path('/data')
        if data_path.exists() and data_path.is_dir():
            data_excel = data_path / "products_links.xlsx"
            if data_excel.exists():
                file_path = data_excel
            else:
                # Fallback: папка со скриптом
                script_dir = Path(__file__).parent
                file_path = script_dir / "products_links.xlsx"
        else:
            # Локальная разработка
            script_dir = Path(__file__).parent
            file_path = script_dir / "products_links.xlsx"
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"📦 Excel файл не найден: {file_path}")
        print("   Используются стандартные товары")
        print("   Для управления товарами через Excel:")
        print("   1. Запусти: python parser_gui.py")
        print("   2. Создай шаблон и заполни ссылки")
        print("   3. Спарси товары")
        print("   4. Перезапусти мини-апп\n")
        PRODUCTS = PRODUCTS_DEFAULT
        return

    try:
        # Проверяем openpyxl
        try:
            import openpyxl
        except ImportError:
            print("📦 Устанавливаю openpyxl...")
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            import openpyxl

        from openpyxl import load_workbook

        # Загружаем двумя способами:
        # data_only=True — кэшированные значения формул (если файл сохранён из Excel)
        # data_only=False — сырые значения (числа или формулы как строки)
        wb_data = load_workbook(file_path, data_only=True)
        wb_raw = load_workbook(file_path, data_only=False)
        ws_data = wb_data.active
        ws_raw = wb_raw.active

        products = []

        for row_num in range(2, ws_data.max_row + 1):
            name = ws_data.cell(row_num, 2).value          # B: Название
            category = ws_data.cell(row_num, 4).value      # D: Группа
            subcategory = ws_data.cell(row_num, 5).value       # E: Подгруппа
            product_category = ws_data.cell(row_num, 6).value  # F: Категория товара
            brand = ws_data.cell(row_num, 20).value            # T: Бренд
            gender = ws_data.cell(row_num, 21).value           # U: Пол
            balance = ws_data.cell(row_num, 22).value          # V: Баланс
            priority = ws_data.cell(row_num, 23).value         # W: Приоритет
            image_urls = ws_data.cell(row_num, 7).value        # G: URL фото
            local_images = ws_data.cell(row_num, 8).value      # H: Локальное фото
            sizes = ws_data.cell(row_num, 9).value             # I: Размеры

            # Берём цену: P(кэш) → P(сырое) → C(€)
            price_cached = ws_data.cell(row_num, 16).value  # P: кэш формулы
            price_raw = ws_raw.cell(row_num, 16).value      # P: сырое значение
            price_eur = ws_data.cell(row_num, 3).value       # C: Цена (€)

            price = None
            # 1. Кэшированное значение формулы (файл сохранён из Excel)
            if price_cached and isinstance(price_cached, (int, float)) and price_cached > 0:
                price = int(price_cached)
            # 2. Сырое число в ячейке (не формула)
            elif price_raw and isinstance(price_raw, (int, float)) and price_raw > 0:
                price = int(price_raw)
            # 3. Фоллбэк на цену в евро
            elif price_eur and isinstance(price_eur, (int, float)) and price_eur > 0:
                price = int(price_eur)

            # Пропускаем строки без данных
            if not name or not price:
                continue

            # Определяем изображение для показа
            image_to_use = "📦"  # По умолчанию placeholder эмодзи
            all_images = []  # Все фотографии для галереи

            # Если есть локальные фотографии, используем первую
            if local_images:
                # Локальные фото могут быть разделены запятыми
                local_photos = [img.strip() for img in local_images.split(',')]
                if local_photos:
                    # Создаем массив всех локальных фотографий для галереи
                    for photo in local_photos:
                        # Убираем префикс "images\" или "images/" если он есть
                        photo_path = photo.replace('images\\', '').replace('images/', '')
                        all_images.append(f"/images/{photo_path}")

                    # Используем первую фотографию как основную
                    image_to_use = all_images[0]

            # Парсим размеры в массив
            sizes_array = []
            if sizes:
                sizes_array = [s.strip() for s in str(sizes).split(',') if s.strip()]

            products.append({
                "id": row_num - 1,
                "name": name,
                "price": int(price) if price else 0,
                "image": image_to_use,
                "images": all_images if all_images else [image_to_use],
                "sizes": sizes_array,
                "category": category or "",
                "subcategory": subcategory or "",
                "product_category": product_category or "",
                "brand": brand or "",
                "gender": {"женские": "Женский", "мужские": "Мужской", "девочки": "Женский", "мальчики": "Мужской"}.get((gender or "").strip().lower(), gender) or "Унисекс",
                "balance": balance or "",
                "priority": int(priority) if priority and isinstance(priority, (int, float)) else 999,
            })

        wb_data.close()
        wb_raw.close()

        # Сортируем по приоритету (1 первым, 999 = без приоритета — в конец)
        products.sort(key=lambda p: p['priority'])

        if products:
            PRODUCTS = products
            print(f"✅ Загружено товаров из Excel: {len(products)}")

            # Подсчитываем товары с фотографиями
            with_photos = sum(1 for p in products if p['image'].startswith('/images/'))
            print(f"   📸 Товаров с фотографиями: {with_photos}")
            print(f"   📦 Товаров с эмодзи: {len(products) - with_photos}\n")
        else:
            print("⚠️  Excel файл пустой, используются стандартные товары\n")
            PRODUCTS = PRODUCTS_DEFAULT

    except Exception as e:
        print(f"❌ Ошибка загрузки Excel: {e}")
        print("   Используются стандартные товары\n")
        PRODUCTS = PRODUCTS_DEFAULT

# ═══════════════════════════════════════════════════════════
# 🤖 TELEGRAM БОТ
# ═══════════════════════════════════════════════════════════

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    """Команда /start - показывает приветствие и кнопку магазина."""
    # Проверяем параметр (из канала: /start catalog)
    args = message.text.split(maxsplit=1)
    from_channel = len(args) > 1 and args[1] == "catalog"

    # Клавиатура с кнопкой каталога внизу
    reply_keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(
                    text="🛍 Каталог",
                    web_app=WebAppInfo(url=WEBAPP_URL),
                )
            ]
        ],
        resize_keyboard=True,
        is_persistent=True,
    )

    if from_channel:
        # Пришёл из канала — сразу показываем inline кнопку с Mini App
        inline_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="🛍 Открыть каталог",
                        web_app=WebAppInfo(url=WEBAPP_URL),
                    )
                ]
            ]
        )
        await message.answer(
            "🏐 <b>Добро пожаловать в NIMBLI!</b>\n\n"
            "Нажми кнопку чтобы открыть каталог! 👇",
            reply_markup=inline_keyboard,
            parse_mode="HTML",
        )
        # Также ставим кнопку каталога внизу
        await message.answer(
            "Или используй кнопку внизу экрана ⬇️",
            reply_markup=reply_keyboard,
            parse_mode="HTML",
        )
    else:
        await message.answer(
            "🏐 <b>Добро пожаловать в NIMBLI!</b>\n\n"
            "⚡ Твой спортивный магазин!\n"
            "Нажми кнопку <b>🛍 Каталог</b> внизу экрана!",
            reply_markup=reply_keyboard,
            parse_mode="HTML",
        )


@dp.message(Command("shop"))
async def cmd_shop(message: types.Message):
    """Команда /shop - открывает магазин."""
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🛍 Открыть магазин",
                    web_app=WebAppInfo(url=WEBAPP_URL),
                )
            ]
        ]
    )
    await message.answer(
        "Нажми кнопку, чтобы открыть каталог:",
        reply_markup=keyboard,
    )


CHANNEL_USERNAME = "@nimbli_sport"

@dp.message(Command("post"))
async def cmd_post(message: types.Message):
    """Команда /post - публикует пост в канал. Писать боту в личку."""
    username = message.from_user.username
    if username not in ADMIN_USERNAMES:
        return

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🛍 Открыть каталог",
                    url="https://t.me/Catalog_Alex_bot?start=catalog",
                )
            ]
        ]
    )

    try:
        await bot.send_message(
            chat_id=CHANNEL_USERNAME,
            text="<b>NIMBLI</b> | Нишевый спорт из Европы 🏀\n\n"
                 "Падел | Волейбол | Теннис | Бег 🏸🏐🎾\n"
                 "— оригинальные бренды\n"
                 "— прямая поставка из Европы\n"
                 "— цены ниже рынка\n"
                 "— доставка почти до вашей двери 🔥",
            reply_markup=keyboard,
            parse_mode="HTML",
        )
        await message.answer("✅ Пост опубликован в канал!")
    except Exception as e:
        await message.answer(f"❌ Ошибка публикации: {e}")


@dp.message(Command("reload"))
async def cmd_reload(message: types.Message):
    """Команда /reload - перезагружает каталог товаров из Excel."""
    await message.answer("🔄 Перезагружаю каталог товаров...")

    try:
        load_products_from_excel()
        await message.answer(
            f"✅ Каталог обновлён!\n\n"
            f"📦 Товаров: {len(PRODUCTS)}\n"
            f"📸 С фото: {sum(1 for p in PRODUCTS if p['image'].startswith('/images/'))}"
        )
    except Exception as e:
        await message.answer(f"❌ Ошибка обновления каталога:\n{str(e)}")


@dp.message(F.document)
async def handle_document(message: types.Message):
    """Обрабатывает загрузку архивов с каталогом товаров."""
    # Проверяем права админа
    username = message.from_user.username
    if username not in ADMIN_USERNAMES:
        await message.answer(
            "🚫 <b>Доступ запрещён!</b>\n\n"
            "Загружать каталог могут только администраторы.\n"
            "Обратитесь к @AlexeyBakaev, @musyanya или @GussionHovo",
            parse_mode="HTML"
        )
        logger.warning(
            f"⚠️ Попытка загрузки каталога от неавторизованного пользователя: "
            f"@{username} ({message.from_user.full_name})"
        )
        return

    document = message.document

    # Проверяем расширение файла (только ZIP для простоты)
    if not document.file_name.endswith('.zip'):
        await message.answer(
            "⚠️ Пожалуйста, отправь ZIP архив с каталогом.\n\n"
            "📝 Как создать архив:\n"
            "  1. Положи в одну папку:\n"
            "     • products_links.xlsx\n"
            "     • папку images/\n"
            "  2. Выдели оба → ПКМ → Отправить → Сжатая ZIP-папка\n\n"
            "Структура архива:\n"
            "  📁 catalog.zip\n"
            "     ├── 📄 products_links.xlsx\n"
            "     └── 📁 images/\n"
            "          ├── 🖼 product_1.webp\n"
            "          ├── 🖼 product_2.webp\n"
            "          └── ..."
        )
        return

    try:
        await message.answer("📥 Скачиваю архив...")

        # Определяем где сохранять файлы (приоритет /data для Amvera)
        data_path = Path('/data')
        if data_path.exists() and data_path.is_dir():
            # На Amvera - сохраняем в /data (persistenceMount)
            extract_dir = data_path
        else:
            # Локально - сохраняем в папку со скриптом
            extract_dir = Path(__file__).parent

        archive_path = extract_dir / document.file_name

        await bot.download(document, destination=archive_path)
        await message.answer("✅ Архив скачан, распаковываю...")

        # Распаковываем ZIP
        import zipfile
        with zipfile.ZipFile(archive_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)

        # Удаляем архив
        archive_path.unlink()

        await message.answer("✅ Архив распакован, обновляю каталог...")

        # Перезагружаем товары
        load_products_from_excel()

        await message.answer(
            f"🎉 Каталог успешно обновлён!\n\n"
            f"📦 Товаров: {len(PRODUCTS)}\n"
            f"📸 С фотографиями: {sum(1 for p in PRODUCTS if p['image'].startswith('/images/'))}\n\n"
            f"Используй /shop чтобы открыть магазин"
        )

    except zipfile.BadZipFile:
        await message.answer("❌ Ошибка: файл повреждён или это не ZIP архив")
    except Exception as e:
        logger.error("Ошибка обработки архива: %s", e)
        await message.answer(f"❌ Ошибка обработки архива:\n{str(e)}")


@dp.message(F.web_app_data)
async def handle_web_app_data(message: types.Message):
    """Обрабатывает данные из Mini App (консультация)."""
    print(f"\n🎯 WEB_APP_DATA ПОЛУЧЕН! Данные: {message.web_app_data.data[:200]}\n", flush=True)
    try:
        data = json.loads(message.web_app_data.data)
        print(f"📦 Распарсено: {data}\n", flush=True)
        action = data.get("action", "order")
        items = data.get("items", [])
        total = data.get("total", 0)

        if not items:
            await message.answer("❌ Вы не выбрали ни одного интересного товара!")
            return

        # Формируем список товаров для сообщения менеджеру
        products_list = ""
        for item in items:
            rounded_price = math.ceil(item['price'] / 100) * 100
            products_list += f"• {item['name']} — {rounded_price:,.0f} ₽\n".replace(',', ' ')

        # Формируем текст для предзаполнения в личке
        rounded_total = math.ceil(total / 100) * 100
        prefilled_text = f"Здравствуйте, подскажите о наличии товара:\n\n{products_list}\n💰 Общая стоимость: {rounded_total:,.0f} ₽".replace(',', ' ')
        encoded_text = quote(prefilled_text)

        # Показываем сообщение пользователю с выбором менеджера
        message_text = "⭐ <b>Вас заинтересовали следующие товары:</b>\n\n"

        for item in items:
            rounded_price = math.ceil(item['price'] / 100) * 100
            formatted_price = f"{rounded_price:,.0f}".replace(',', ' ')
            message_text += (
                f"<b>{item['name']}</b>\n"
                f"💰 Цена: {formatted_price} ₽\n\n"
            )

        formatted_total = f"{rounded_total:,.0f}".replace(',', ' ')
        message_text += f"📊 <b>Общая стоимость: {formatted_total} ₽</b>\n\n"
        message_text += (
            "💬 <b>Выберите менеджера для консультации:</b>"
        )

        # Кнопки для связи с менеджерами (с предзаполненным текстом)
        keyboard = types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(
                text="👤 @AlexeyBakaev",
                url=f"https://t.me/AlexeyBakaev?text={encoded_text}"
            )],
            [types.InlineKeyboardButton(
                text="👤 @musyanya",
                url=f"https://t.me/musyanya?text={encoded_text}"
            )]
        ])

        await message.answer(message_text, parse_mode="HTML", reply_markup=keyboard)

        # Логируем в консоль
        logger.info(
            "Запрос консультации от %s (@%s): %d товаров на %d ₽",
            message.from_user.full_name,
            message.from_user.username or "без username",
            len(items),
            total,
        )

    except (json.JSONDecodeError, KeyError) as e:
        logger.error("Ошибка обработки данных: %s", e)
        await message.answer("❌ Произошла ошибка при обработке данных.")


# ═══════════════════════════════════════════════════════════
# 🌐 ВЕБ-СЕРВЕР (раздаёт HTML и API)
# ═══════════════════════════════════════════════════════════

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Магазин</title>
    <script src="https://telegram.org/js/telegram-web-app.js"></script>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        @keyframes pulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.05); }
        }

        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(100%);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        @keyframes gradientShift {
            0% {
                background-position: 0% 50%;
            }
            50% {
                background-position: 100% 50%;
            }
            100% {
                background-position: 0% 50%;
            }
        }

        @keyframes float {
            0%, 100% {
                transform: translateY(0px);
            }
            50% {
                transform: translateY(-20px);
            }
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', sans-serif;
            background: linear-gradient(-45deg, #e2e6ea, #d6dae0, #caced4, #dfe3e7);
            background-size: 400% 400%;
            animation: gradientShift 15s ease infinite;
            color: #1a1a1a;
            padding: 20px;
            padding-bottom: 100px;
            min-height: 100vh;
            position: relative;
            overflow-x: hidden;
        }

        body::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(255, 255, 255, 0.2);
            backdrop-filter: blur(100px);
            pointer-events: none;
            z-index: 0;
        }

        .header {
            text-align: center;
            margin-bottom: 24px;
            animation: fadeInUp 0.6s ease-out;
            position: relative;
            z-index: 1;
        }

        h1 {
            font-size: 52px;
            font-weight: 900;
            margin-bottom: 8px;
            color: white;
            text-shadow:
                0 2px 10px rgba(0,0,0,0.3),
                0 4px 20px rgba(0,0,0,0.2),
                0 0 40px rgba(255,215,0,0.3);
            letter-spacing: 5px;
            animation: float 3s ease-in-out infinite;
            text-transform: uppercase;
        }

        .subtitle {
            color: rgba(30, 30, 30, 0.85);
            margin-bottom: 0;
            font-size: 16px;
            font-weight: 600;
            text-shadow: 0 2px 10px rgba(255,255,255,0.5);
            letter-spacing: 2px;
        }

        .categories-container {
            margin: 32px auto 16px;
            max-width: 100%;
            animation: fadeInUp 0.8s ease-out 0.2s both;
        }

        .categories-tabs {
            display: flex;
            justify-content: center;
            gap: 10px;
            overflow-x: auto;
            padding: 4px 20px 12px;
            scrollbar-width: thin;
            scrollbar-color: rgba(102, 126, 234, 0.3) transparent;
        }

        .categories-tabs::-webkit-scrollbar {
            height: 4px;
        }

        .categories-tabs::-webkit-scrollbar-track {
            background: transparent;
        }

        .categories-tabs::-webkit-scrollbar-thumb {
            background: rgba(102, 126, 234, 0.3);
            border-radius: 2px;
        }

        .category-tab {
            flex-shrink: 0;
            padding: 10px 20px;
            border: none;
            border-radius: 20px;
            background: rgba(255, 255, 255, 0.7);
            backdrop-filter: blur(10px);
            color: #2d3748;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
            white-space: nowrap;
        }

        .category-tab:hover {
            background: rgba(255, 255, 255, 0.9);
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.12);
        }

        .category-tab.active {
            background: linear-gradient(135deg, #FFD700 0%, #DAA520 100%);
            color: #2d3748;
            box-shadow: 0 4px 15px rgba(255, 215, 0, 0.5);
        }

        .subcategories-container {
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.4s ease, opacity 0.3s ease;
            opacity: 0;
        }

        .subcategories-container.visible {
            max-height: 60px;
            opacity: 1;
        }

        .subcategories-tabs {
            display: flex;
            gap: 8px;
            padding: 0 20px 10px;
            overflow-x: auto;
            scrollbar-width: none;
        }
        .subcategories-tabs::-webkit-scrollbar {
            display: none;
        }

        .subcategory-tab {
            flex-shrink: 0;
            padding: 6px 16px;
            border: none;
            border-radius: 16px;
            background: rgba(255, 255, 255, 0.5);
            color: #555;
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            white-space: nowrap;
        }

        .subcategory-tab:hover {
            background: rgba(255, 255, 255, 0.8);
        }

        .subcategory-tab.active {
            background: rgba(255, 215, 0, 0.3);
            color: #2d3748;
            border: 1.5px solid rgba(218, 165, 32, 0.5);
        }

        .brands-container {
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.4s ease, opacity 0.3s ease;
            opacity: 0;
        }

        .brands-container.visible {
            max-height: 50px;
            opacity: 1;
        }

        .brands-tabs {
            display: flex;
            justify-content: center;
            gap: 6px;
            padding: 0 20px 8px;
            overflow-x: auto;
            scrollbar-width: none;
        }

        .brands-tabs::-webkit-scrollbar {
            display: none;
        }

        .brand-tab {
            flex-shrink: 0;
            padding: 5px 12px;
            border: 1px solid rgba(0,0,0,0.1);
            border-radius: 14px;
            background: rgba(255, 255, 255, 0.5);
            color: #666;
            font-size: 11px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            white-space: nowrap;
        }

        .brand-tab:hover {
            background: rgba(255, 255, 255, 0.8);
        }

        .brand-tab.active {
            background: #2d3748;
            color: white;
            border-color: #2d3748;
        }

        /* Фильтр размеров */
        .size-filter-container {
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.4s ease, opacity 0.3s ease;
            opacity: 0;
            margin-top: 8px;
        }
        .size-filter-container.visible {
            max-height: 200px;
            opacity: 1;
            overflow-y: auto;
        }
        .size-filter-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 0 4px;
            margin-bottom: 8px;
        }
        .size-filter-title {
            font-size: 12px;
            font-weight: 600;
            color: var(--tg-theme-hint-color, #999);
        }
        .size-filter-reset {
            font-size: 11px;
            color: #667eea;
            cursor: pointer;
            border: none;
            background: none;
            padding: 2px 6px;
        }
        .size-filter-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(65px, 1fr));
            gap: 6px;
            padding: 0 4px;
        }
        .size-filter-btn {
            padding: 7px 4px;
            border: 1.5px solid rgba(0,0,0,0.12);
            border-radius: 8px;
            background: rgba(255,255,255,0.6);
            font-size: 12px;
            font-weight: 500;
            color: #333;
            cursor: pointer;
            text-align: center;
            transition: all 0.2s ease;
        }
        .size-filter-btn:hover {
            border-color: rgba(0,0,0,0.25);
        }
        .size-filter-btn.active {
            background: #2d3748;
            color: white;
            border-color: #2d3748;
        }

        /* Фильтр по полу */
        .gender-filter-container {
            display: flex;
            gap: 8px;
            justify-content: center;
            margin-top: 10px;
            padding: 0 4px;
            overflow-x: auto;
            scrollbar-width: none;
        }
        .gender-filter-container::-webkit-scrollbar {
            display: none;
        }
        .gender-tab {
            flex-shrink: 0;
            padding: 5px 14px;
            border: 1px solid rgba(0,0,0,0.1);
            border-radius: 14px;
            background: rgba(255,255,255,0.5);
            font-size: 12px;
            font-weight: 500;
            color: #555;
            cursor: pointer;
            transition: all 0.2s ease;
        }
        .gender-tab.active {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-color: transparent;
        }

        /* Бейдж баланса на карточке */
        .priority-badge {
            position: absolute;
            top: 8px;
            left: 8px;
            padding: 3px 8px;
            border-radius: 6px;
            font-size: 10px;
            font-weight: 700;
            z-index: 10;
            color: white;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .priority-badge.hot {
            background: linear-gradient(135deg, #ff416c 0%, #ff4b2b 100%);
        }
        .priority-badge.new {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }

        .search-container {
            margin: 20px auto 24px;
            max-width: 600px;
            animation: fadeInUp 0.8s ease-out 0.3s both;
        }

        .search-wrapper {
            position: relative;
            display: flex;
            align-items: center;
        }

        .search-icon {
            position: absolute;
            left: 18px;
            font-size: 20px;
            opacity: 0.6;
            pointer-events: none;
            z-index: 2;
        }

        .search-input {
            width: 100%;
            padding: 16px 50px 16px 52px;
            font-size: 16px;
            border: none;
            border-radius: 20px;
            background: rgba(255, 255, 255, 0.85);
            backdrop-filter: blur(10px);
            box-shadow:
                0 4px 20px rgba(0, 0, 0, 0.1),
                inset 0 1px 0 rgba(255, 255, 255, 0.9);
            transition: all 0.3s ease;
            font-family: inherit;
            color: #333;
        }

        .search-input:focus {
            outline: none;
            background: rgba(255, 255, 255, 0.95);
            box-shadow:
                0 6px 30px rgba(0, 0, 0, 0.15),
                inset 0 1px 0 rgba(255, 255, 255, 1),
                0 0 0 3px rgba(129, 212, 250, 0.3);
            transform: translateY(-2px);
        }

        .search-input::placeholder {
            color: rgba(0, 0, 0, 0.4);
        }

        .clear-search {
            position: absolute;
            right: 18px;
            font-size: 20px;
            color: rgba(0, 0, 0, 0.5);
            cursor: pointer;
            padding: 4px 8px;
            border-radius: 50%;
            transition: all 0.2s ease;
            z-index: 2;
        }

        .clear-search:hover {
            color: rgba(0, 0, 0, 0.8);
            background: rgba(0, 0, 0, 0.05);
        }

        .products-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(170px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
            position: relative;
            z-index: 1;
        }

        .product-card {
            background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
            border-radius: 16px;
            padding: 14px;
            cursor: pointer;
            transition: all 0.3s ease;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
            position: relative;
            overflow: hidden;
            animation: fadeInUp 0.6s ease-out;
            animation-fill-mode: both;
            border: 1px solid rgba(0, 0, 0, 0.06);
        }

        .product-card:hover {
            transform: translateY(-4px) scale(1.03);
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.12);
        }

        .product-card:active {
            transform: scale(0.98);
        }

        .product-card.in-cart {
            border-color: #667eea;
            box-shadow: 0 4px 16px rgba(102, 126, 234, 0.2);
        }

        .product-image {
            width: 100%;
            height: 140px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: linear-gradient(135deg, #f0f4f8 0%, #e2e8f0 100%);
            border-radius: 12px;
            margin-bottom: 12px;
            overflow: hidden;
        }

        .product-image img {
            max-width: 100%;
            max-height: 120px;
            object-fit: contain;
        }

        .product-image div {
            font-size: 56px;
        }

        .product-badge {
            position: absolute;
            top: 8px;
            right: 8px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 10px;
            font-weight: 700;
            z-index: 3;
            box-shadow: 0 2px 8px rgba(102, 126, 234, 0.4);
            display: none;
        }

        .product-card.in-cart .product-badge {
            display: block;
            animation: pulse 2s infinite;
        }

        .product-name {
            font-weight: 600;
            font-size: 14px;
            margin-bottom: 8px;
            color: #2d3748;
            line-height: 1.4;
            display: -webkit-box;
            -webkit-line-clamp: 2;
            -webkit-box-orient: vertical;
            overflow: hidden;
            min-height: 40px;
        }

        .product-price {
            font-size: 20px;
            font-weight: 800;
            color: #667eea;
            letter-spacing: -0.5px;
        }
        .price-delivery-hint {
            font-size: 10px;
            font-weight: 400;
            color: #999;
            letter-spacing: 0;
        }

        .product-quantity {
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-top: 10px;
            gap: 10px;
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
            padding: 6px;
            border-radius: 12px;
        }

        .quantity-btn {
            width: 36px;
            height: 36px;
            border-radius: 10px;
            border: none;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-size: 20px;
            font-weight: bold;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.3s ease;
            box-shadow: 0 2px 8px rgba(102, 126, 234, 0.3);
        }

        .quantity-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
        }

        .quantity-btn:active {
            transform: scale(0.95);
        }

        .quantity-display {
            font-weight: 700;
            font-size: 18px;
            min-width: 30px;
            text-align: center;
            color: var(--tg-theme-text-color, #212529);
        }

        .cart-footer {
            position: fixed;
            bottom: 0;
            left: 0;
            right: 0;
            background: rgba(255, 255, 255, 0.25);
            backdrop-filter: blur(30px) saturate(180%);
            -webkit-backdrop-filter: blur(30px) saturate(180%);
            padding: 20px;
            box-shadow: 0 -8px 32px rgba(0, 0, 0, 0.2);
            display: none;
            border-top: 1px solid rgba(255, 255, 255, 0.4);
            z-index: 100;
        }

        .cart-footer.visible {
            display: block;
            animation: slideUp 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        }

        .cart-summary {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 14px;
            font-size: 16px;
            font-weight: 700;
            color: #2d2d2d;
            text-shadow: 0 1px 3px rgba(255,255,255,0.5);
        }

        .cart-total {
            font-size: 26px;
            font-weight: 900;
            background: linear-gradient(135deg, #FF6B35, #FFD93D);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            filter: drop-shadow(0 2px 4px rgba(255,107,53,0.3));
        }

        .order-btn {
            width: 100%;
            padding: 18px;
            border-radius: 16px;
            border: none;
            background: linear-gradient(135deg, #FF6B35 0%, #FFD93D 100%);
            color: white;
            font-size: 18px;
            font-weight: 800;
            cursor: pointer;
            box-shadow: 0 8px 24px rgba(255,107,53,0.4);
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            letter-spacing: 1px;
            text-transform: uppercase;
            position: relative;
            overflow: hidden;
        }

        .order-btn::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
            transition: left 0.5s;
        }

        .order-btn:hover::before {
            left: 100%;
        }

        .order-btn:hover {
            transform: translateY(-3px) scale(1.02);
            box-shadow: 0 12px 32px rgba(255,107,53,0.5);
        }

        .order-btn:active {
            transform: scale(0.98);
        }

        .empty-cart {
            text-align: center;
            padding: 60px 20px;
            color: var(--tg-theme-hint-color, #6c757d);
        }

        .empty-cart-icon {
            font-size: 72px;
            margin-bottom: 20px;
            opacity: 0.5;
            animation: pulse 3s infinite;
        }

        /* Модальное окно товара */
        .modal-overlay {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.7);
            backdrop-filter: blur(10px);
            display: none;
            align-items: center;
            justify-content: center;
            z-index: 1000;
            padding: 20px;
            animation: fadeIn 0.3s ease;
        }

        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }

        .modal-overlay.active {
            display: flex;
        }

        .modal-content {
            background: #ffffff;
            border-radius: 20px;
            max-width: 500px;
            width: 100%;
            max-height: 90vh;
            overflow-y: auto;
            position: relative;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.3);
            animation: slideInUp 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        }

        @keyframes slideInUp {
            from {
                opacity: 0;
                transform: translateY(50px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        .modal-close {
            position: absolute;
            top: 16px;
            right: 16px;
            width: 36px;
            height: 36px;
            border-radius: 50%;
            background: rgba(0, 0, 0, 0.1);
            border: none;
            font-size: 24px;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            z-index: 1;
            transition: all 0.3s ease;
        }

        .modal-close:hover {
            background: rgba(0, 0, 0, 0.2);
            transform: rotate(90deg);
        }

        .modal-image-container {
            position: relative;
            width: 100%;
            height: 300px;
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
            border-radius: 20px 20px 0 0;
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
        }

        .modal-image {
            width: 100%;
            height: 100%;
            object-fit: cover;
            transition: opacity 0.3s ease;
        }

        .gallery-nav {
            position: absolute;
            top: 50%;
            transform: translateY(-50%);
            width: 40px;
            height: 40px;
            background: rgba(255, 255, 255, 0.9);
            border: none;
            border-radius: 50%;
            font-size: 24px;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.2);
            transition: all 0.3s ease;
            z-index: 10;
        }

        .gallery-nav:hover {
            background: white;
            transform: translateY(-50%) scale(1.1);
        }

        .gallery-nav-prev {
            left: 10px;
        }

        .gallery-nav-next {
            right: 10px;
        }

        .gallery-counter {
            position: absolute;
            bottom: 10px;
            left: 50%;
            transform: translateX(-50%);
            background: rgba(0, 0, 0, 0.7);
            color: white;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            z-index: 10;
        }

        .sizes-section {
            margin-bottom: 20px;
        }

        .sizes-title {
            font-size: 16px;
            font-weight: 700;
            margin-bottom: 12px;
            color: var(--tg-theme-text-color, #212529);
        }

        .sizes-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(60px, 1fr));
            gap: 8px;
        }

        .size-badge {
            padding: 10px;
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
            border: 2px solid rgba(102, 126, 234, 0.3);
            border-radius: 10px;
            text-align: center;
            font-weight: 600;
            font-size: 14px;
            color: #212529;
            transition: all 0.3s ease;
        }

        .size-badge:hover {
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.2) 0%, rgba(118, 75, 162, 0.2) 100%);
            border-color: #667eea;
            transform: translateY(-2px);
        }

        .modal-body {
            padding: 24px;
        }

        .modal-title {
            font-size: 24px;
            font-weight: 800;
            margin-bottom: 12px;
            color: #212529;
            line-height: 1.3;
        }

        .modal-price-section {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 16px;
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
            border-radius: 14px;
            margin-bottom: 20px;
        }

        .modal-price-label {
            font-size: 14px;
            font-weight: 600;
            color: #6c757d;
        }

        .modal-price {
            font-size: 32px;
            font-weight: 800;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .modal-info-section {
            margin-bottom: 20px;
        }

        .modal-info-title {
            font-size: 16px;
            font-weight: 700;
            margin-bottom: 12px;
            color: #212529;
        }

        .modal-info-item {
            display: flex;
            justify-content: space-between;
            padding: 10px 0;
            border-bottom: 1px solid rgba(0, 0, 0, 0.05);
        }

        .modal-info-item:last-child {
            border-bottom: none;
        }

        .modal-info-label {
            font-size: 14px;
            color: #6c757d;
        }

        .modal-info-value {
            font-size: 14px;
            font-weight: 600;
            color: #212529;
        }

        .modal-actions {
            display: flex;
            gap: 12px;
        }

        .modal-btn {
            flex: 1;
            padding: 16px;
            border-radius: 14px;
            border: none;
            font-size: 16px;
            font-weight: 700;
            cursor: pointer;
            transition: all 0.3s ease;
        }

        .modal-btn-primary {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            box-shadow: 0 4px 16px rgba(102, 126, 234, 0.4);
        }

        .modal-btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.5);
        }

        .modal-btn-secondary {
            background: rgba(102, 126, 234, 0.1);
            color: #667eea;
        }

        .modal-btn-secondary:hover {
            background: rgba(102, 126, 234, 0.2);
        }
        /* Particles на фоне */
        .particle {
            position: fixed;
            width: 4px;
            height: 4px;
            background: rgba(255, 255, 255, 0.5);
            border-radius: 50%;
            pointer-events: none;
            animation: float-particle 8s infinite;
        }

        @keyframes float-particle {
            0%, 100% {
                transform: translateY(0) translateX(0);
                opacity: 0;
            }
            10% {
                opacity: 1;
            }
            90% {
                opacity: 1;
            }
            100% {
                transform: translateY(-100vh) translateX(20px);
                opacity: 0;
            }
        }

        .badge {
            position: absolute;
            top: 12px;
            right: 12px;
            background: linear-gradient(135deg, #a18cd1, #fbc2eb);
            color: white;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 9px;
            font-weight: 600;
            z-index: 2;
        }

        .badge.new {
            background: linear-gradient(135deg, #4facfe, #00f2fe);
            box-shadow: 0 4px 15px rgba(79, 172, 254, 0.4);
        }

        .gender-badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 8px;
            font-size: 10px;
            font-weight: 600;
            margin-top: 4px;
        }
        .gender-badge.male {
            background: rgba(66, 133, 244, 0.15);
            color: #4285f4;
        }
        .gender-badge.female {
            background: rgba(234, 67, 149, 0.15);
            color: #ea4395;
        }
        .gender-badge.unisex {
            background: rgba(102, 126, 234, 0.15);
            color: #667eea;
        }

    </style>
</head>
<body>
    <!-- Animated particles -->
    <div class="particles-container"></div>

    <div class="header">
        <h1>🏐 NIMBLI</h1>
        <p class="subtitle">Твой спорт • Твоя победа • Твой успех</p>
    </div>

    <div class="categories-container">
        <div class="categories-tabs" id="categoriesTabs"></div>
        <div class="subcategories-container" id="subcategoriesContainer">
            <div class="subcategories-tabs" id="subcategoriesTabs"></div>
        </div>
        <div class="brands-container" id="brandsContainer">
            <div class="brands-tabs" id="brandsTabs"></div>
        </div>
        <div class="size-filter-container" id="sizeFilterContainer">
            <div class="size-filter-header">
                <span class="size-filter-title">Размер (EU)</span>
                <button class="size-filter-reset" id="sizeFilterReset" style="display:none;">Сбросить</button>
            </div>
            <div class="size-filter-grid" id="sizeFilterGrid"></div>
        </div>
        <div class="gender-filter-container" id="genderFilterContainer"></div>
    </div>

    <div class="search-container">
        <div class="search-wrapper">
            <span class="search-icon">🔍</span>
            <input
                type="text"
                id="searchInput"
                class="search-input"
                placeholder="Поиск товаров..."
                autocomplete="off"
            />
            <span class="clear-search" id="clearSearch" style="display: none;">✕</span>
        </div>
    </div>

    <div class="products-grid" id="productsGrid"></div>

    <div class="cart-footer" id="cartFooter">
        <div class="cart-summary">
            <span>⭐ Интересных: <span id="cartCount">0</span></span>
            <span class="cart-total"><span id="cartTotal">0</span> ₽</span>
        </div>
        <button class="order-btn" id="orderBtn">💬 Консультация</button>
    </div>

    <!-- Модальное окно детального просмотра -->
    <div class="modal-overlay" id="productModal">
        <div class="modal-content">
            <button class="modal-close" onclick="closeProductModal()">×</button>
            <div class="modal-image-container">
                <img class="modal-image" id="modalImage" src="" alt="">
            </div>
            <div class="modal-body">
                <h2 class="modal-title" id="modalTitle"></h2>

                <div class="modal-price-section">
                    <span class="modal-price-label">Цена</span>
                    <span class="modal-price" id="modalPrice"></span>
                </div>

                <div id="modalGenderSection" style="display: none; margin-bottom: 12px;">
                    <span id="modalGenderBadge"></span>
                </div>

                <div class="sizes-section" id="sizesSection" style="display: none;">
                    <h3 class="sizes-title">👟 Размеры (EU)</h3>
                    <div class="sizes-grid" id="sizesGrid"></div>
                </div>

                <div class="modal-actions">
                    <button class="modal-btn modal-btn-secondary" onclick="closeProductModal()">Закрыть</button>
                    <button class="modal-btn modal-btn-primary" id="modalAddBtn" onclick="">В корзину</button>
                </div>
            </div>
        </div>
    </div>

    <script>
        // Создаём floating particles
        function createParticles() {
            const particleCount = 25;
            for (let i = 0; i < particleCount; i++) {
                const particle = document.createElement('div');
                particle.className = 'particle';
                particle.style.left = Math.random() * 100 + '%';
                particle.style.animationDelay = Math.random() * 8 + 's';
                particle.style.animationDuration = (8 + Math.random() * 4) + 's';
                document.body.appendChild(particle);
            }
        }

        const tg = window.Telegram.WebApp;
        tg.expand();
        tg.MainButton.hide();

        let cart = {};  // Теперь это список интересных товаров
        let products = [];
        let currentCategory = null;  // Текущая выбранная группа
        let currentSubcategory = null;  // Текущая выбранная подгруппа
        let currentBrand = null;  // Текущий выбранный бренд
        let selectedSizes = new Set();  // Выбранные размеры для фильтра
        let currentGender = null;  // Текущий выбранный пол

        // Форматирование цены с пробелами (22000 → 22 000)
        function formatPrice(price) {
            return (Math.ceil(price / 100) * 100).toString().replace(/\B(?=(\d{3})+(?!\d))/g, ' ');
        }

        // Инициализация particles при загрузке
        createParticles();
        let currentProduct = null;  // Текущий товар в модальном окне

        // Загружаем товары с сервера
        fetch('/api/products')
            .then(res => res.json())
            .then(data => {
                products = data;
                renderCategories();
                renderProducts();
            });

        // Поиск товаров
        const searchInput = document.getElementById('searchInput');
        const clearSearch = document.getElementById('clearSearch');

        searchInput.addEventListener('input', (e) => {
            const query = e.target.value;
            renderProducts(query);
            clearSearch.style.display = query ? 'block' : 'none';
        });

        clearSearch.addEventListener('click', () => {
            searchInput.value = '';
            clearSearch.style.display = 'none';
            renderProducts('');
            searchInput.focus();
        });

        // Открытие модального окна
        function openProductModal(productId) {
            currentProduct = products.find(p => p.id === productId);
            if (!currentProduct) return;

            const modal = document.getElementById('productModal');
            const modalImage = document.getElementById('modalImage');
            const modalTitle = document.getElementById('modalTitle');
            const modalPrice = document.getElementById('modalPrice');
            const modalAddBtn = document.getElementById('modalAddBtn');
            const sizesSection = document.getElementById('sizesSection');
            const sizesGrid = document.getElementById('sizesGrid');

            // Устанавливаем изображение (только первое)
            if (currentProduct.image && currentProduct.image.startsWith('/images/')) {
                modalImage.src = currentProduct.image;
                modalImage.style.display = 'block';
            } else {
                modalImage.style.display = 'none';
            }

            modalTitle.textContent = currentProduct.name;
            modalPrice.textContent = formatPrice(currentProduct.price) + ' ₽';

            // Показываем пол, если указан
            const genderSection = document.getElementById('modalGenderSection');
            const genderBadge = document.getElementById('modalGenderBadge');
            if (currentProduct.gender) {
                const gl = currentProduct.gender.toLowerCase();
                let genderClass = 'unisex';
                if (gl.includes('мужск') || gl.includes('male') || gl.includes('man') || gl.includes('men')) genderClass = 'male';
                else if (gl.includes('женск') || gl.includes('female') || gl.includes('woman') || gl.includes('women')) genderClass = 'female';
                genderBadge.className = 'gender-badge ' + genderClass;
                genderBadge.textContent = currentProduct.gender;
                genderSection.style.display = 'block';
            } else {
                genderSection.style.display = 'none';
            }

            // Показываем размеры, если они есть
            if (currentProduct.sizes && currentProduct.sizes.length > 0) {
                sizesSection.style.display = 'block';
                sizesGrid.innerHTML = '';
                currentProduct.sizes.forEach(size => {
                    const sizeBadge = document.createElement('div');
                    sizeBadge.className = 'size-badge';
                    sizeBadge.textContent = size;
                    sizesGrid.appendChild(sizeBadge);
                });
            } else {
                sizesSection.style.display = 'none';
            }

            // Обновляем кнопку
            const isInteresting = cart[productId] && cart[productId] > 0;
            modalAddBtn.textContent = isInteresting ? '✓ Убрать из интересного' : '⭐ Интересно';
            modalAddBtn.onclick = () => toggleInteresting(productId);

            modal.classList.add('active');
            tg.HapticFeedback.impactOccurred('medium');
        }

        // Закрытие модального окна
        function closeProductModal() {
            const modal = document.getElementById('productModal');
            modal.classList.remove('active');
            currentProduct = null;
        }

        // Переключение "Интересно"
        function toggleInteresting(productId) {
            if (cart[productId] && cart[productId] > 0) {
                delete cart[productId];
            } else {
                cart[productId] = 1;
            }
            renderProducts();
            updateCartFooter();
            closeProductModal();
            tg.HapticFeedback.impactOccurred('light');
        }

        // Закрытие модального окна при клике на overlay
        document.getElementById('productModal').addEventListener('click', function(e) {
            if (e.target === this) {
                closeProductModal();
            }
        });

        function renderProducts(searchQuery = '') {
            const grid = document.getElementById('productsGrid');
            grid.innerHTML = '';

            // Фильтруем товары по группе, подгруппе и поисковому запросу
            const filteredProducts = products.filter(product => {
                // Фильтр по группе
                if (currentCategory && product.category !== currentCategory) {
                    return false;
                }
                // Фильтр по подгруппе
                if (currentSubcategory && product.subcategory !== currentSubcategory) {
                    return false;
                }
                // Фильтр по бренду
                if (currentBrand && product.brand !== currentBrand) {
                    return false;
                }
                // Фильтр по размерам
                if (selectedSizes.size > 0 && product.sizes && product.sizes.length > 0) {
                    const hasSize = product.sizes.some(s => selectedSizes.has(s));
                    if (!hasSize) return false;
                }
                if (selectedSizes.size > 0 && (!product.sizes || product.sizes.length === 0)) {
                    return false;
                }
                // Фильтр по полу
                if (currentGender && product.gender !== currentGender) {
                    return false;
                }

                // Фильтр по поисковому запросу
                if (!searchQuery) return true;
                const query = searchQuery.toLowerCase();
                return product.name.toLowerCase().includes(query);
            });

            // Если ничего не найдено
            if (filteredProducts.length === 0) {
                grid.innerHTML = `
                    <div style="grid-column: 1/-1; text-align: center; padding: 60px 20px; color: rgba(0,0,0,0.5);">
                        <div style="font-size: 48px; margin-bottom: 16px;">🔍</div>
                        <div style="font-size: 18px; font-weight: 600;">Ничего не найдено</div>
                        <div style="font-size: 14px; margin-top: 8px;">Попробуйте изменить запрос</div>
                    </div>
                `;
                return;
            }

            filteredProducts.forEach(product => {
                const card = document.createElement('div');
                card.className = 'product-card';
                card.dataset.id = product.id;
                if (cart[product.id]) {
                    card.classList.add('in-cart');
                }

                const quantity = cart[product.id] || 0;

                // Определяем как показывать изображение
                let imageHtml;
                if (product.image.startsWith('/images/')) {
                    // Реальная фотография
                    imageHtml = `<img src="${product.image}" alt="${product.name}" onerror="this.outerHTML='<div>📦</div>'">`;
                } else {
                    // Placeholder эмодзи
                    imageHtml = `<div>${product.image}</div>`;
                }

                // Определяем класс пола
                let genderBadgeHtml = '';
                if (product.gender) {
                    const gl = product.gender.toLowerCase();
                    let genderClass = 'unisex';
                    if (gl.includes('мужск') || gl.includes('male') || gl.includes('man') || gl.includes('men')) genderClass = 'male';
                    else if (gl.includes('женск') || gl.includes('female') || gl.includes('woman') || gl.includes('women')) genderClass = 'female';
                    genderBadgeHtml = `<div class="gender-badge ${genderClass}">${product.gender}</div>`;
                }

                card.innerHTML = `
                    ${quantity > 0 ? '<div class="product-badge">⭐ Интересно</div>' : ''}
                    ${product.priority === 1 ? '<div class="priority-badge hot">Hot</div>' : product.priority === 2 ? '<div class="priority-badge new">New</div>' : ''}
                    <div class="product-image">${imageHtml}</div>
                    <div class="product-name">${product.name}</div>
                    <div class="product-price">${formatPrice(product.price)} ₽ <span class="price-delivery-hint">с доставкой</span></div>
                    ${genderBadgeHtml}
                `;

                // При клике открываем модальное окно
                card.onclick = () => openProductModal(product.id);

                grid.appendChild(card);
            });

            // Добавляем badges на товары
            setTimeout(() => {
                // Бейджи баланса (Мощность и т.д.) — справа сверху
                const productCards = document.querySelectorAll('.product-card');
                productCards.forEach((card) => {
                    const productId = parseInt(card.dataset.id);
                    const product = products.find(p => p.id === productId);
                    if (product && product.balance && !card.querySelector('.badge')) {
                        const badge = document.createElement('div');
                        badge.className = 'badge';
                        badge.textContent = product.balance;
                        card.appendChild(badge);
                    }
                });
            }, 50);

            updateCartFooter();
        }

        // Рендеринг вкладок категорий (группы)
        function renderCategories() {
            const categoriesContainer = document.getElementById('categoriesTabs');
            categoriesContainer.innerHTML = '';

            const categories = [...new Set(products.map(p => p.category).filter(c => c && c.trim()))];

            if (products.length === 0) return;

            if (categories.length === 0) {
                const allTab = document.createElement('button');
                allTab.className = 'category-tab active';
                allTab.textContent = 'Все';
                categoriesContainer.appendChild(allTab);
                renderSubcategories();
                renderBrands();
                return;
            }

            // Вкладка "Все"
            const allTab = document.createElement('button');
            allTab.className = 'category-tab' + (!currentCategory ? ' active' : '');
            allTab.textContent = 'Все';
            allTab.onclick = () => {
                if (!currentCategory) return;
                currentCategory = null;
                currentSubcategory = null;
                currentBrand = null;
                selectedSizes.clear();
                currentGender = null;
                renderCategories();
                renderSubcategories();
                renderBrands();
                renderProducts(searchInput.value);
            };
            categoriesContainer.appendChild(allTab);

            // Вкладки групп
            categories.sort().forEach(category => {
                const tab = document.createElement('button');
                tab.className = 'category-tab' + (currentCategory === category ? ' active' : '');
                tab.textContent = category;
                tab.onclick = () => {
                    if (currentCategory === category) {
                        currentCategory = null;
                        currentSubcategory = null;
                    } else {
                        currentCategory = category;
                        currentSubcategory = null;
                    }
                    currentBrand = null;
                    selectedSizes.clear();
                currentGender = null;
                    renderCategories();
                    renderSubcategories();
                    renderBrands();
                    renderProducts(searchInput.value);
                };
                categoriesContainer.appendChild(tab);
            });

            renderSubcategories();
            renderBrands();
        }

        // Рендеринг подвкладок (подгруппы)
        function renderSubcategories() {
            const container = document.getElementById('subcategoriesContainer');
            const tabsContainer = document.getElementById('subcategoriesTabs');
            tabsContainer.innerHTML = '';

            // Если группа не выбрана — скрываем подвкладки
            if (!currentCategory) {
                container.classList.remove('visible');
                return;
            }

            // Получаем подгруппы для выбранной группы
            const subcategories = [...new Set(
                products
                    .filter(p => p.category === currentCategory)
                    .map(p => p.subcategory)
                    .filter(s => s && s.trim())
            )];

            // Если подгрупп нет или только одна — не показываем
            if (subcategories.length <= 1) {
                container.classList.remove('visible');
                return;
            }

            // Показываем контейнер
            container.classList.add('visible');

            // Вкладка "Все" для подгрупп
            const allTab = document.createElement('button');
            allTab.className = 'subcategory-tab' + (!currentSubcategory ? ' active' : '');
            allTab.textContent = 'Все';
            allTab.onclick = () => {
                if (!currentSubcategory) return;
                currentSubcategory = null;
                currentBrand = null;
                selectedSizes.clear();
                currentGender = null;
                renderSubcategories();
                renderBrands();
                renderProducts(searchInput.value);
            };
            tabsContainer.appendChild(allTab);

            // Вкладки подгрупп
            subcategories.sort().forEach(sub => {
                const tab = document.createElement('button');
                tab.className = 'subcategory-tab' + (currentSubcategory === sub ? ' active' : '');
                tab.textContent = sub;
                tab.onclick = () => {
                    if (currentSubcategory === sub) {
                        currentSubcategory = null;
                    } else {
                        currentSubcategory = sub;
                    }
                    currentBrand = null;
                    selectedSizes.clear();
                currentGender = null;
                    renderSubcategories();
                    renderBrands();
                    renderProducts(searchInput.value);
                };
                tabsContainer.appendChild(tab);
            });
        }

        // Рендеринг фильтра по брендам
        function renderBrands() {
            const container = document.getElementById('brandsContainer');
            const tabsContainer = document.getElementById('brandsTabs');
            tabsContainer.innerHTML = '';

            // Получаем бренды для текущей выборки (с учётом группы и подгруппы)
            const filteredForBrands = products.filter(p => {
                if (currentCategory && p.category !== currentCategory) return false;
                if (currentSubcategory && p.subcategory !== currentSubcategory) return false;
                return true;
            });

            const brands = [...new Set(filteredForBrands.map(p => p.brand).filter(b => b && b.trim()))];

            // Если брендов нет или только один — не показываем
            if (brands.length <= 1) {
                container.classList.remove('visible');
                currentBrand = null;
                renderSizeFilter();
                return;
            }

            container.classList.add('visible');

            // Вкладка "Все бренды"
            const allTab = document.createElement('button');
            allTab.className = 'brand-tab' + (!currentBrand ? ' active' : '');
            allTab.textContent = 'Все';
            allTab.onclick = () => {
                if (!currentBrand) return;
                currentBrand = null;
                selectedSizes.clear();
                currentGender = null;
                renderBrands();
                renderSizeFilter();
                renderProducts(searchInput.value);
            };
            tabsContainer.appendChild(allTab);

            // Вкладки брендов
            brands.sort().forEach(brand => {
                // Считаем количество товаров этого бренда
                const count = filteredForBrands.filter(p => p.brand === brand).length;
                const tab = document.createElement('button');
                tab.className = 'brand-tab' + (currentBrand === brand ? ' active' : '');
                tab.textContent = `${brand} (${count})`;
                tab.onclick = () => {
                    if (currentBrand === brand) {
                        currentBrand = null;
                    } else {
                        currentBrand = brand;
                    }
                    selectedSizes.clear();
                currentGender = null;
                    renderBrands();
                    renderSizeFilter();
                    renderProducts(searchInput.value);
                };
                tabsContainer.appendChild(tab);
            });

            renderSizeFilter();
        }

        // Рендеринг фильтра по размерам
        function renderSizeFilter() {
            const container = document.getElementById('sizeFilterContainer');
            const grid = document.getElementById('sizeFilterGrid');
            const resetBtn = document.getElementById('sizeFilterReset');
            grid.innerHTML = '';

            // Собираем все размеры из текущей выборки
            const filteredForSizes = products.filter(p => {
                if (currentCategory && p.category !== currentCategory) return false;
                if (currentSubcategory && p.subcategory !== currentSubcategory) return false;
                if (currentBrand && p.brand !== currentBrand) return false;
                return true;
            });

            const allSizes = new Set();
            filteredForSizes.forEach(p => {
                if (p.sizes) p.sizes.forEach(s => allSizes.add(s));
            });

            if (allSizes.size <= 1) {
                container.classList.remove('visible');
                selectedSizes.clear();
                currentGender = null;
                renderGenderFilter();
                return;
            }

            container.classList.add('visible');

            // Сортировка размеров (числовая)
            const sortedSizes = [...allSizes].sort((a, b) => {
                const numA = parseFloat(a.replace(/[^\d.,]/g, '').replace(',', '.'));
                const numB = parseFloat(b.replace(/[^\d.,]/g, '').replace(',', '.'));
                if (!isNaN(numA) && !isNaN(numB)) return numA - numB;
                return a.localeCompare(b);
            });

            // Кнопка сброса
            resetBtn.style.display = selectedSizes.size > 0 ? 'block' : 'none';
            resetBtn.onclick = () => {
                selectedSizes.clear();
                currentGender = null;
                renderSizeFilter();
                renderProducts(searchInput.value);
            };

            sortedSizes.forEach(size => {
                const btn = document.createElement('button');
                btn.className = 'size-filter-btn' + (selectedSizes.has(size) ? ' active' : '');
                btn.textContent = size;
                btn.onclick = () => {
                    if (selectedSizes.has(size)) {
                        selectedSizes.delete(size);
                    } else {
                        selectedSizes.add(size);
                    }
                    renderSizeFilter();
                    renderProducts(searchInput.value);
                };
                grid.appendChild(btn);
            });

            renderGenderFilter();
        }

        // Рендеринг фильтра по полу
        function renderGenderFilter() {
            const container = document.getElementById('genderFilterContainer');
            container.innerHTML = '';

            // Собираем полы из текущей выборки
            const filtered = products.filter(p => {
                if (currentCategory && p.category !== currentCategory) return false;
                if (currentSubcategory && p.subcategory !== currentSubcategory) return false;
                if (currentBrand && p.brand !== currentBrand) return false;
                return true;
            });

            const genders = [...new Set(filtered.map(p => p.gender).filter(g => g && g.trim()))];

            if (genders.length <= 1) {
                currentGender = null;
                return;
            }

            // Вкладка "Все"
            const allTab = document.createElement('button');
            allTab.className = 'gender-tab' + (!currentGender ? ' active' : '');
            allTab.textContent = 'Все';
            allTab.onclick = () => {
                if (!currentGender) return;
                currentGender = null;
                renderGenderFilter();
                renderProducts(searchInput.value);
            };
            container.appendChild(allTab);

            genders.sort().forEach(g => {
                const tab = document.createElement('button');
                tab.className = 'gender-tab' + (currentGender === g ? ' active' : '');
                tab.textContent = g;
                tab.onclick = () => {
                    currentGender = currentGender === g ? null : g;
                    renderGenderFilter();
                    renderProducts(searchInput.value);
                };
                container.appendChild(tab);
            });
        }

        function changeQuantity(productId, delta) {
            if (!cart[productId]) {
                cart[productId] = 0;
            }

            cart[productId] += delta;

            if (cart[productId] <= 0) {
                delete cart[productId];
            }

            renderProducts();
            tg.HapticFeedback.impactOccurred('light');
        }

        function updateCartFooter() {
            const footer = document.getElementById('cartFooter');
            const cartCount = document.getElementById('cartCount');
            const cartTotal = document.getElementById('cartTotal');

            let totalItems = 0;
            let totalPrice = 0;

            for (const [productId, quantity] of Object.entries(cart)) {
                const product = products.find(p => p.id === parseInt(productId));
                if (product) {
                    totalItems += quantity;
                    totalPrice += product.price * quantity;
                }
            }

            if (totalItems > 0) {
                footer.classList.add('visible');
                cartCount.textContent = totalItems;
                cartTotal.textContent = formatPrice(totalPrice);
            } else {
                footer.classList.remove('visible');
            }
        }

        // Подготовка данных для отправки
        function prepareConsultationData() {
            const items = [];
            let total = 0;

            for (const [productId, quantity] of Object.entries(cart)) {
                const product = products.find(p => p.id === parseInt(productId));
                if (product) {
                    items.push({
                        id: product.id,
                        name: product.name,
                        price: product.price,
                        quantity: quantity,
                        image: product.image
                    });
                    total += product.price * quantity;
                }
            }

            return {
                action: 'consultation',
                items: items,
                total: total
            };
        }

        // Кнопка консультации - открывает выбор менеджера
        document.getElementById('orderBtn').addEventListener('click', () => {
            const data = prepareConsultationData();

            if (data.items.length === 0) {
                tg.showAlert('Добавьте хотя бы один товар в интересное!');
                return;
            }

            // Формируем текст для отправки менеджеру
            let messageText = 'Здравствуйте, подскажите о наличии товара:\\n\\n';
            data.items.forEach(item => {
                messageText += `• ${item.name} — ${formatPrice(item.price)} ₽\\n`;
            });
            messageText += `\\n💰 Общая стоимость: ${formatPrice(data.total)} ₽`;

            // Случайно выбираем менеджера
            const managers = ['AlexeyBakaev', 'musyanya'];
            const username = managers[Math.floor(Math.random() * managers.length)];
            const url = `https://t.me/${username}?text=${encodeURIComponent(messageText)}`;

            // Открываем чат с менеджером
            tg.openTelegramLink(url);
        });
    </script>
</body>
</html>
"""


async def handle_index(request: web.Request) -> web.Response:
    """Отдаёт HTML страницу Mini App."""
    return web.Response(text=HTML_TEMPLATE, content_type="text/html")


async def handle_products(request: web.Request) -> web.Response:
    """API: список товаров в формате JSON."""
    return web.json_response(PRODUCTS)


async def handle_webhook(request: web.Request) -> web.Response:
    """Обработчик webhook от Telegram."""
    try:
        update_data = await request.json()
        print(f"\n📥 WEBHOOK: {json.dumps(update_data, ensure_ascii=False)[:500]}\n", flush=True)
        from aiogram.types import Update
        update = Update(**update_data)
        await dp.feed_update(bot, update)
        return web.Response(text="OK")
    except Exception as e:
        print(f"\n❌ WEBHOOK ERROR: {e}\n", flush=True)
        import traceback
        print(traceback.format_exc(), flush=True)
        return web.Response(status=500, text=str(e))


def create_web_app() -> web.Application:
    """Создаёт веб-приложение aiohttp."""
    app = web.Application()
    app.router.add_get("/", handle_index)
    app.router.add_get("/api/products", handle_products)
    app.router.add_post("/webhook", handle_webhook)  # Webhook endpoint

    # Раздаём статические файлы (фотографии товаров)
    images_dir = get_images_dir()
    if images_dir.exists():
        app.router.add_static("/images/", path=images_dir, name="images")
        logger.info(f"📁 Раздация изображений из: {images_dir}")

    return app


# ═══════════════════════════════════════════════════════════
# 🚀 ЗАПУСК
# ═══════════════════════════════════════════════════════════

async def main():
    """Запускает бота и веб-сервер одновременно."""
    global WEBAPP_URL

    # Освобождаем порт перед запуском
    print(f"🔍 Проверяю порт {WEBAPP_PORT}...")
    kill_process_on_port(WEBAPP_PORT)
    print(f"✅ Порт {WEBAPP_PORT} готов к использованию\n")

    # Загружаем товары из Excel
    load_products_from_excel()

    tunnel_process = None

    # 1. Настраиваем публичный URL
    if MODE == "auto":
        # Автоматический режим с Serveo
        print("🔧 Режим: Автоматический (Serveo)\n")
        WEBAPP_URL, tunnel_process = start_serveo(WEBAPP_PORT)

        if not WEBAPP_URL:
            # Serveo не сработал - запускаемся локально
            print("\n" + "=" * 60)
            print("⚠️  SERVEO НЕДОСТУПЕН - ЗАПУСК В ЛОКАЛЬНОМ РЕЖИМЕ")
            print("=" * 60)
            print()
            print("🏠 Бот запущен локально на http://localhost:8080")
            print()
            print("⚠️  ВАЖНО:")
            print("   • Telegram Mini App НЕ БУДЕТ РАБОТАТЬ")
            print("   • Можно открыть http://localhost:8080 в браузере")
            print("   • Для полной работы нужен публичный HTTPS URL")
            print()
            print("💡 Как получить публичный URL:")
            print()
            print("   ВАРИАНТ 1: Serveo (ручной режим)")
            print("     1. Открой новый терминал")
            print(f"     2. Запусти: ssh -R 80:localhost:{WEBAPP_PORT} serveo.net")
            print("     3. Скопируй полученный URL")
            print("     4. Вставь URL в mini_app.py (строка 205):")
            print('        MANUAL_WEBAPP_URL = "твой_url"')
            print("     5. Измени MODE = \"manual\" (строка 202)")
            print("     6. Перезапусти бота")
            print()
            print("   ВАРИАНТ 2: LocalTunnel")
            print(f"     npx localtunnel --port {WEBAPP_PORT}")
            print()
            print("   ВАРИАНТ 3: Деплой на облако (Railway, Render)")
            print("     Бот будет работать 24/7 с автоматическим HTTPS")
            print()
            print("=" * 60)
            print()

            # Запускаемся локально для тестирования
            WEBAPP_URL = f"http://localhost:{WEBAPP_PORT}"
            print(f"▶️  Запускаю в локальном режиме...")
            print(f"   Адрес: {WEBAPP_URL}")
            print()

            # Автоматически открываем браузер через 3 секунды
            import webbrowser
            from threading import Timer
            def open_browser():
                try:
                    webbrowser.open(WEBAPP_URL)
                    print("🌐 Открыл веб-интерфейс в браузере")
                except:
                    pass
            Timer(3.0, open_browser).start()

    else:
        # Ручной режим - используем указанный URL
        WEBAPP_URL = MANUAL_WEBAPP_URL
        print("📌 Ручной режим: используется URL из настроек")
        print(f"🌍 URL: {WEBAPP_URL}\n")

    # 2. Запускаем веб-сервер
    web_app = create_web_app()
    runner = web.AppRunner(web_app)
    await runner.setup()
    site = web.TCPSite(runner, WEBAPP_HOST, WEBAPP_PORT)
    await site.start()

    logger.info("=" * 60)
    logger.info("🌐 Локальный сервер: http://%s:%s", WEBAPP_HOST, WEBAPP_PORT)
    logger.info("🌍 Публичный URL (Mini App): %s", WEBAPP_URL)
    logger.info("=" * 60)

    # 3. Запускаем бота
    logger.info("🤖 Telegram бот запущен!")
    logger.info("💬 Напиши боту /start чтобы открыть магазин!\n")

    # Определяем режим работы
    use_webhook = WEBAPP_URL and ("amvera.io" in WEBAPP_URL or WEBAPP_URL.startswith("https://"))

    try:
        if use_webhook:
            # Webhook mode для продакшена (Amvera и др.)
            webhook_url = f"{WEBAPP_URL}/webhook"
            logger.info("🔗 Режим: WEBHOOK")
            logger.info(f"📍 Webhook URL: {webhook_url}")
            # Устанавливаем webhook с поддержкой групповых сообщений
            await bot.set_webhook(
                url=webhook_url,
                allowed_updates=["message", "callback_query", "inline_query", "web_app_data"],
            )
            logger.info("✅ Webhook установлен")
            await asyncio.Event().wait()  # Бесконечное ожидание
        else:
            # Polling mode для локальной разработки
            logger.info("🔄 Режим: POLLING (локальная разработка)")
            await dp.start_polling(bot)
    finally:
        # Останавливаем всё при выходе
        logger.info("Останавливаю сервер...")
        await runner.cleanup()
        if tunnel_process:
            logger.info("Останавливаю туннель...")
            tunnel_process.kill()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Остановка бота...")
